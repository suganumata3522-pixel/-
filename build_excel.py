# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ---- 共通スタイル ----
header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True, size=11)
title_font = Font(bold=True, size=14, color="1F4E78")
note_font = Font(italic=True, size=9, color="808080")
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
alt_fill = PatternFill("solid", fgColor="EAF1FB")

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

def style_body(ws, start_row, end_row, ncols):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = wrap
            cell.border = border
            if (r - start_row) % 2 == 1:
                cell.fill = alt_fill

# =====================================================================
# シート1：応募先機関リスト
# =====================================================================
ws = wb.active
ws.title = "応募先機関リスト"

ws["A1"] = "住宅性能評価・構造審査 在宅副業 応募先機関リスト"
ws["A1"].font = title_font
ws["A2"] = "作成日：2026-06-22 ／ 構造設計一級建築士の強みを活かせる「構造分野審査」を◎で表示"
ws["A2"].font = note_font

headers = ["No", "機関名", "区分", "委託の主な業務内容", "在宅可否",
           "構造枠の有無", "応募・問合せ窓口（URL）", "応募方法・備考"]
ws.append([])  # row3 spacer
hr = 4
for i, h in enumerate(headers, 1):
    ws.cell(row=hr, column=i, value=h)
style_header(ws, hr, len(headers))

rows = [
    [1, "日本ERI株式会社", "指定確認検査機関／登録評価機関",
     "設計住宅性能評価（意匠・構造の図面審査）を自宅PCで実施。意匠／構造で分業",
     "◎ 在宅可", "◎ 構造審査枠あり",
     "https://www.j-eri.co.jp/corporate/corp_recruit_outsrc.html",
     "「業務委託」ページから応募。業界最大手、案件量が安定"],
    [2, "ビューローベリタスジャパン株式会社", "指定確認検査機関／登録評価機関",
     "設計図面評価（意匠 or 構造）＋現場検査（直行直帰）。副業歓迎を明記",
     "◎ 設計評価は在宅可", "◎ 構造選択可",
     "https://hrmos.co/pages/bureauveritas/jobs/01_01_122_999_80",
     "採用サイト(hrmos)から応募。構造一級を前面に出すと構造枠で優遇"],
    [3, "ハウスプラス確認検査株式会社", "指定確認検査機関／登録評価機関",
     "構造審査・意匠審査・高層評定・省エネ評価・構造評価(任意評価)など",
     "△ 要相談", "◎ 構造審査・構造評価あり",
     "https://www.houseplus.co.jp/hpa/recruit/index.html",
     "履歴書・職務経歴書・資格証(写)を郵送。構造評価部門あり＝構造一級向き"],
    [4, "株式会社東日本住宅評価センター", "登録住宅性能評価機関",
     "設計評価・建設評価の評価員委託（既存住宅の建設評価含む）",
     "△ 一部在宅", "○",
     "https://www.e-hyoka.co.jp/outsourcing/",
     "委託募集ページから応募 → 書類選考＋面接の2段階"],
    [5, "株式会社日本住宅保証検査機構（JIO）", "登録評価機関／保険法人",
     "住宅性能評価員（設計評価・検査）。研修後に在宅勤務へ移行",
     "◎ 研修後に在宅可", "○",
     "https://www.jio-kensa.co.jp/recruit/info/evaluation-judge.html",
     "募集要項ページから応募。月給型は20〜40万円提示の例"],
    [6, "一般財団法人ベターリビング", "登録評価機関（一般財団法人）",
     "住宅性能評価業務（業務委託契約の評価員）、技術系職員",
     "△ 要相談", "○",
     "https://www.cbl.or.jp/career/index",
     "採用情報＞キャリア採用から応募"],
    [7, "確認サービス株式会社 ほか地場の登録評価機関", "登録住宅性能評価機関",
     "設計評価の図面審査・建設評価。地場機関は柔軟な委託条件が多い",
     "△〜◎ 機関による", "機関による",
     "https://www.kakunin-s.com/gyomu/hyouka/",
     "居住地＋『登録住宅性能評価機関』で検索し地場機関に直接打診"],
    [8, "（横断検索）住宅性能評価・表示協会", "機関検索サイト",
     "全国の登録住宅性能評価機関 約123機関を地域・業務で検索可能",
     "—", "—",
     "https://www.hyoukakyoukai.or.jp/kikan/hyouka_search.php",
     "まずここで居住地の機関を洗い出すのが効率的"],
]
for r in rows:
    ws.append(r)
end_row = hr + len(rows)
style_body(ws, hr + 1, end_row, len(headers))

# 列幅
widths = [4, 26, 22, 38, 12, 16, 42, 34]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 22

# 注記
nr = end_row + 2
ws.cell(row=nr, column=1,
        value="※ 報酬・在宅条件・案件量は各機関とも非公開。応募時に①構造分野審査を希望②在宅可否③1件単価か月額か の3点を確認すると効率的。")
ws.cell(row=nr, column=1).font = note_font

# =====================================================================
# シート2：報酬・作業時間の目安
# =====================================================================
ws2 = wb.create_sheet("報酬・作業時間の目安")
ws2["A1"] = "報酬・作業時間・業務内容の目安"
ws2["A1"].font = title_font
ws2["A2"] = "※各機関が単価非公開のため、求人・実務者情報からの相場感。確定額は応募時に確認のこと"
ws2["A2"].font = note_font

h2 = ["項目", "目安・内容", "補足"]
r0 = 4
for i, h in enumerate(h2, 1):
    ws2.cell(row=r0, column=i, value=h)
style_header(ws2, r0, len(h2))

data2 = [
    ["業務内容（設計評価）", "品確法の10分野について図面・仕様・構造計算書をチェックし評価書ドラフト作成、申請者への照会対応", "在宅向き。デスクワーク中心"],
    ["業務内容（建設評価）", "工事段階での現地検査（標準4回程度の立会い）", "現地必須・直行直帰型"],
    ["報酬（出来高制）", "戸建1件あたり 数千円〜1万円台が中心", "構造分野・長期優良併願など複雑案件は高め"],
    ["報酬（月額・委託）", "評価員で月20〜40万円提示の例。副業実務者は月5〜15万円程度の事例が多い", "稼働量による"],
    ["作業時間（設計評価）", "戸建1件あたり概ね1〜3時間", "慣れと難易度による"],
    ["作業時間（建設評価）", "1日完結型が多く単価は高め", "現地移動を含む"],
    ["構造分野審査", "耐震等級など構造審査は人手不足で、意匠審査より高単価になりやすい", "構造設計一級の最大の強み"],
]
for d in data2:
    ws2.append(d)
e2 = r0 + len(data2)
style_body(ws2, r0 + 1, e2, len(h2))
for i, w in enumerate([22, 60, 34], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# =====================================================================
# シート3：構造設計一級建築士を活かす＋資格・手続き
# =====================================================================
ws3 = wb.create_sheet("構造一級の活かし方・資格")
ws3["A1"] = "構造設計一級建築士を活かせる仕事と必要資格・手続き"
ws3["A1"].font = title_font

h3 = ["優先度", "仕事の種類", "内容", "必要な資格・手続き", "在宅性／報酬感"]
r3 = 3
for i, h in enumerate(h3, 1):
    ws3.cell(row=r3, column=i, value=h)
style_header(ws3, r3, len(h3))

data3 = [
    ["①最短", "住宅性能評価の構造分野審査", "耐震等級等の構造図面・計算書審査を在宅受託",
     "一級建築士＋各機関での評価員講習修了・登録（資格面は充足済）", "在宅◎／意匠より高単価"],
    ["②中期", "構造計算適合性判定員（適判員）", "確認審査と並行する構造計算の適合性判定。指定適判機関・都道府県が委嘱",
     "『構造計算適合判定資格者検定』(日本建築防災協会)に合格→登録が必要。受検資格は充足", "通所＋持帰り併用／1件数万円規模と高単価"],
    ["③併用", "長期優良住宅・省エネ適合性判定(BELS等)の技術審査", "性能・省エネ性の技術審査",
     "評価機関での登録（構造一級が有利）", "在宅○／案件次第"],
    ["③併用", "構造設計・構造計算の外注受託", "設計事務所等からの構造計算・図面の受託",
     "実務経験（資格充足）。クラウドワークス等でも案件あり", "在宅◎／単価は個別交渉"],
    ["③併用", "確認検査機関での構造審査補助", "建築確認の構造審査の補助業務",
     "一級建築士（建築基準適合判定資格者があればなお可）", "機関による"],
]
for d in data3:
    ws3.append(d)
e3 = r3 + len(data3)
style_body(ws3, r3 + 1, e3, len(h3))
for i, w in enumerate([10, 30, 36, 44, 24], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# 適判員の補足ブロック
br = e3 + 2
ws3.cell(row=br, column=1, value="【重要】構造計算適合性判定員になる手順").font = Font(bold=True, color="C00000")
steps = [
    "1. 『構造計算適合判定資格者検定』を受検（実施：一般財団法人 日本建築防災協会）。構造設計一級建築士は受検資格を満たす。",
    "2. 検定合格後、構造計算適合判定資格者として登録（国交省／都道府県）。",
    "3. 指定構造計算適合性判定機関・都道府県の適判員として委嘱を受ける（非常勤・案件委嘱型あり）。",
    "→ 構造設計一級だけでは自動的に適判員にはなれない点に注意。検定合格＋登録が前提。",
]
for k, s in enumerate(steps, 1):
    ws3.cell(row=br + k, column=1, value=s)
    ws3.cell(row=br + k, column=1).alignment = wrap

# =====================================================================
# シート4：進め方チェックリスト
# =====================================================================
ws4 = wb.create_sheet("進め方チェックリスト")
ws4["A1"] = "進め方・確認事項チェックリスト"
ws4["A1"].font = title_font

h4 = ["□", "ステップ", "内容"]
r4 = 3
for i, h in enumerate(h4, 1):
    ws4.cell(row=r4, column=i, value=h)
style_header(ws4, r4, len(h4))

data4 = [
    ["", "STEP1 機関の洗い出し", "評価協会の検索サイトで居住地の登録評価機関をリスト化"],
    ["", "STEP2 応募", "日本ERI・ビューローベリタス等に『設計評価の構造分野・在宅業務委託』で応募"],
    ["", "STEP3 強みの明示", "応募時に構造設計一級建築士を前面に出し、構造審査枠を希望"],
    ["", "STEP4 条件確認", "①構造分野審査の可否 ②在宅可否 ③1件単価か月額か の3点を必ず確認"],
    ["", "STEP5 評価員登録", "採用後、各機関で評価員講習修了・登録手続き"],
    ["", "STEP6 中期目標", "構造計算適合判定資格者検定を受検・合格 → 適判員登録で高単価案件へ展開"],
    ["", "応募書類", "履歴書・職務経歴書・資格証(写)。郵送指定の機関あり(ハウスプラス等)"],
]
for d in data4:
    ws4.append(d)
e4 = r4 + len(data4)
style_body(ws4, r4 + 1, e4, len(h4))
for i, w in enumerate([5, 26, 70], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

# 出典シート
ws5 = wb.create_sheet("出典")
ws5["A1"] = "出典・参考リンク"
ws5["A1"].font = title_font
src = [
    "東日本住宅評価センター｜委託募集： https://www.e-hyoka.co.jp/outsourcing/",
    "日本ERI株式会社｜業務委託： https://www.j-eri.co.jp/corporate/corp_recruit_outsrc.html",
    "ビューローベリタスジャパン｜住宅性能評価員 業務委託： https://hrmos.co/pages/bureauveritas/jobs/01_01_122_999_80",
    "ハウスプラス住宅保証｜採用情報： https://www.houseplus.co.jp/hpa/recruit/index.html",
    "JIO 日本住宅保証検査機構｜評価員 募集要項： https://www.jio-kensa.co.jp/recruit/info/evaluation-judge.html",
    "ベターリビング｜キャリア採用： https://www.cbl.or.jp/career/index",
    "確認サービス｜住宅性能評価業務： https://www.kakunin-s.com/gyomu/hyouka/",
    "住宅性能評価・表示協会｜登録評価機関検索： https://www.hyoukakyoukai.or.jp/kikan/hyouka_search.php",
    "建築技術教育普及センター｜構造設計一級建築士講習： https://www.jaeic.or.jp/koshuannai/koshu/s1k/s1k-seidozenpan/index.html",
    "国土交通省｜構造計算適合判定資格者検定の実施について： https://www.mlit.go.jp/jutakukentiku/build/jutakukentiku_house_fr_000075.html",
    "東京都都市整備局｜構造計算適合判定資格者登録について： https://www.toshiseibi.metro.tokyo.lg.jp/kenchiku/kouzou_shikaku/touroku.html",
    "※報酬の具体額は各機関が単価非公開のため、求人・実務者情報からの相場感です。確定額は各機関へ直接確認してください。",
]
for k, s in enumerate(src, 3):
    ws5.cell(row=k, column=1, value=s).alignment = wrap
ws5.column_dimensions["A"].width = 110

out = "/home/user/-/住宅性能評価_在宅副業_応募先リスト.xlsx"
wb.save(out)
print("saved:", out)
