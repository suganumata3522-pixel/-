# -*- coding: utf-8 -*-
"""鉄筋の定着 問題集（図つき）Excel 生成スクリプト。
出力: docs/rebar_anchorage/鉄筋定着問題集.xlsx
No.6-1 応力伝達機構 / 6-2 定着長L2・L2h・La・Lb / 6-3 L2・L2h暗記
6-4 計算仮定条件 / 6-5 部材内定着・仕口内定着
RC造マンションの設計担当を想定。
"""
import os
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib import font_manager as fm

FONT_PATH = "/usr/share/fonts/truetype/fonts-japanese-gothic.ttf"
jp = fm.FontProperties(fname=FONT_PATH)
fm.fontManager.addfont(FONT_PATH)
plt.rcParams["font.family"] = jp.get_name()
plt.rcParams["axes.unicode_minus"] = False

OUT = "docs/rebar_anchorage"
FIG = os.path.join(OUT, "figures")
os.makedirs(FIG, exist_ok=True)


def save(fig, name):
    p = os.path.join(FIG, name)
    fig.savefig(p, dpi=130, bbox_inches="tight")
    plt.close(fig)
    return p


# ===========================================================================
# 図 6-1: 付着による応力伝達機構
# ===========================================================================
def fig_6_1():
    fig, axes = plt.subplots(2, 1, figsize=(11, 7.2),
                             gridspec_kw={"height_ratios": [1, 1]})
    # --- (a) 付着応力の伝達イメージ ---
    ax = axes[0]
    # コンクリート
    ax.add_patch(mpatches.Rectangle((0, 0), 11, 2.4, fc="#e8e8e8",
                                     ec="k", lw=1.0))
    # 鉄筋
    ax.add_patch(mpatches.Rectangle((-1.2, 1.0), 8.5, 0.4, fc="#888",
                                     ec="k", lw=0.8, zorder=3))
    # 節（リブ）
    for x in np.arange(0.2, 7.2, 0.55):
        ax.add_patch(mpatches.Rectangle((x, 0.92), 0.08, 0.56, fc="#555",
                                         zorder=4))
    # 引張力（鉄筋端）
    ax.annotate("", xy=(-1.2, 1.2), xytext=(-3.0, 1.2),
                arrowprops=dict(arrowstyle="-|>", color="#c00000", lw=3))
    ax.text(-3.2, 1.2, "T\n(引張力)", fontproperties=jp, ha="right",
            va="center", fontsize=10, color="#c00000")
    # 付着応力（コンクリートが鉄筋を握る方向）
    for x in np.arange(0.5, 7.0, 0.9):
        ax.annotate("", xy=(x, 1.55), xytext=(x + 0.5, 1.75),
                    arrowprops=dict(arrowstyle="-|>", color="#1f7a1f", lw=1.3))
        ax.annotate("", xy=(x, 0.85), xytext=(x + 0.5, 0.65),
                    arrowprops=dict(arrowstyle="-|>", color="#1f7a1f", lw=1.3))
    ax.text(5.5, 2.05, "付着応力 τ（コンクリートと鉄筋の境界で伝達）",
            fontproperties=jp, ha="center", fontsize=9, color="#1f7a1f")
    ax.text(3.3, -0.6, "鉄筋の引張力 T は、節（リブ）の支圧と付着応力 τ により"
                       "定着長にわたって徐々にコンクリートへ伝わる",
            fontproperties=jp, ha="center", fontsize=9.5, color="#444")
    ax.set_xlim(-4.0, 11.5)
    ax.set_ylim(-1.2, 2.6)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(a) 付着による応力伝達（直線定着）",
                 fontproperties=jp, fontsize=11, fontweight="bold")

    # --- (b) 鉄筋応力と付着応力の分布 ---
    ax = axes[1]
    L = 7.0
    x = np.linspace(0, L, 100)
    sigma = (1 - x / L)  # 端で最大、定着端で0
    tau = np.full_like(x, 1.0 / L) * L * 0.14  # ほぼ一様
    ax.fill_between(x, 0, sigma, color="#c00000", alpha=0.25)
    ax.plot(x, sigma, color="#c00000", lw=2, label="鉄筋応力 σs（端で最大→定着端で0）")
    ax.plot(x, tau, color="#1f7a1f", lw=2, ls="--",
            label="付着応力 τ（ほぼ一様と仮定）")
    ax.set_xlabel("定着端からの距離", fontproperties=jp)
    ax.set_ylabel("応力（相対）", fontproperties=jp)
    ax.set_xlim(0, L)
    ax.set_ylim(0, 1.15)
    ax.legend(prop=jp, fontsize=9, loc="upper right")
    ax.text(L / 2, -0.32, "必要定着長 = 鉄筋の引張力をコンクリートに伝えきるのに要する長さ",
            fontproperties=jp, ha="center", fontsize=9.5, color="#1f4e79")
    ax.set_title("(b) 鉄筋応力と付着応力の分布",
                 fontproperties=jp, fontsize=11, fontweight="bold")
    ax.grid(alpha=0.3)
    fig.suptitle("図 6-1  鉄筋の定着 ── 付着による応力伝達機構",
                 fontproperties=jp, fontsize=13, fontweight="bold")
    return save(fig, "fig6-1_bond.png")


# ===========================================================================
# 図 6-2: 各種定着長 L2, L2h, La, Lb
# ===========================================================================
def fig_6_2():
    fig, axes = plt.subplots(2, 2, figsize=(13.5, 9.0))

    # (a) L2: 直線定着（小梁・スラブ筋の梁への定着など）
    ax = axes[0, 0]
    ax.add_patch(mpatches.Rectangle((0, 0), 3.0, 5.0, fc="#d9d9d9",
                                     ec="k", lw=1.0))
    ax.text(1.5, 4.5, "梁・柱など\n(定着先)", fontproperties=jp, ha="center",
            fontsize=8, color="#444")
    ax.add_patch(mpatches.Rectangle((3.0, 2.3), 5.0, 0.35, fc="#bcd2ea",
                                     ec="k", lw=0.6))
    # 定着鉄筋（直線）
    ax.plot([0.5, 6.5], [2.0, 2.0], color="#c00000", lw=3)
    ax.annotate("", xy=(0.5, 1.5), xytext=(3.0, 1.5),
                arrowprops=dict(arrowstyle="<|-|>", color="#1f4e79", lw=1.3))
    ax.text(1.75, 1.1, "L2", fontproperties=jp, ha="center", fontsize=12,
            color="#1f4e79", fontweight="bold")
    ax.text(4.0, 0.3, "L2：直線定着長\n（フックなしでまっすぐ埋め込む）",
            fontproperties=jp, ha="center", fontsize=9, color="#444")
    ax.set_xlim(-0.3, 8.3)
    ax.set_ylim(-0.5, 5.3)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(a) L2：直線定着", fontproperties=jp,
                 fontsize=11, fontweight="bold")

    # (b) L2h: フック付き定着
    ax = axes[0, 1]
    ax.add_patch(mpatches.Rectangle((0, 0), 3.0, 5.0, fc="#d9d9d9",
                                     ec="k", lw=1.0))
    ax.add_patch(mpatches.Rectangle((3.0, 2.3), 5.0, 0.35, fc="#bcd2ea",
                                     ec="k", lw=0.6))
    # フック付き鉄筋（水平→下に折り曲げ）
    ax.plot([1.0, 6.5], [2.0, 2.0], color="#c00000", lw=3)
    ax.plot([1.0, 1.0], [2.0, 0.6], color="#c00000", lw=3)
    ax.annotate("", xy=(1.0, 1.5), xytext=(3.0, 1.5),
                arrowprops=dict(arrowstyle="<|-|>", color="#1f4e79", lw=1.3))
    ax.text(2.0, 1.1, "L2h", fontproperties=jp, ha="center", fontsize=12,
            color="#1f4e79", fontweight="bold")
    ax.text(0.55, 1.0, "フック\n(余長)", fontproperties=jp, ha="center",
            fontsize=8, color="#c00000")
    ax.text(4.0, 0.0, "L2h：フック付き定着長\n（先端を折り曲げ＝定着長を短縮できる）",
            fontproperties=jp, ha="center", fontsize=9, color="#444")
    ax.set_xlim(-0.3, 8.3)
    ax.set_ylim(-0.7, 5.3)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(b) L2h：フック付き定着", fontproperties=jp,
                 fontsize=11, fontweight="bold")

    # (c) La: 小梁・スラブの定着（曲げ余長の取れない直線）/ 投影定着長
    ax = axes[1, 0]
    # 柱
    ax.add_patch(mpatches.Rectangle((0, 0), 2.6, 5.0, fc="#d9d9d9",
                                     ec="k", lw=1.0))
    ax.text(1.3, 4.5, "柱", fontproperties=jp, ha="center", fontsize=9)
    # 梁主筋が柱に折り曲げ定着（投影長 La）
    ax.plot([5.5, 1.4], [2.6, 2.6], color="#c00000", lw=3)
    ax.plot([1.4, 1.4], [2.6, 0.7], color="#c00000", lw=3)
    ax.add_patch(mpatches.Rectangle((2.6, 2.4), 4.0, 0.35, fc="#bcd2ea",
                                     ec="k", lw=0.6))
    ax.text(4.5, 3.0, "梁", fontproperties=jp, ha="center", fontsize=8)
    # 投影定着長 La（柱フェイスから折り曲げ起点まで＋余長）
    ax.annotate("", xy=(2.6, 1.9), xytext=(1.4, 1.9),
                arrowprops=dict(arrowstyle="<|-|>", color="#1f7a1f", lw=1.3))
    ax.text(2.0, 1.5, "La\n(投影定着長)", fontproperties=jp, ha="center",
            fontsize=9, color="#1f7a1f", fontweight="bold")
    ax.text(3.5, -0.3, "La：仕口面（柱フェイス）から測る\n折り曲げ定着の投影長さ",
            fontproperties=jp, ha="center", fontsize=9, color="#444")
    ax.set_xlim(-0.3, 7.0)
    ax.set_ylim(-0.9, 5.3)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(c) La：折り曲げ定着の投影定着長",
                 fontproperties=jp, fontsize=11, fontweight="bold")

    # (d) Lb: 基本定着長
    ax = axes[1, 1]
    ax.text(0.5, 0.78, "Lb：基本定着長", transform=ax.transAxes,
            ha="center", fontproperties=jp, fontsize=13, fontweight="bold",
            color="#1f4e79")
    ax.text(0.5, 0.52,
            "Lb = (σt / (10·fb)) · db\n\n"
            "σt：鉄筋の引張応力度（許容 or 降伏）\n"
            "fb：許容付着応力度（コンクリート強度・位置で決まる）\n"
            "db：鉄筋径",
            transform=ax.transAxes, ha="center", fontproperties=jp,
            fontsize=10, color="#444")
    ax.text(0.5, 0.16,
            "L2・L2h・La は、この基本定着長 Lb を元に\n"
            "フックや余長を考慮して定めた『設計用の定着長』",
            transform=ax.transAxes, ha="center", fontproperties=jp,
            fontsize=9.5, color="#c00000")
    ax.axis("off")
    ax.set_title("(d) Lb：基本定着長（すべての元）",
                 fontproperties=jp, fontsize=11, fontweight="bold")
    fig.suptitle("図 6-2  各種定着長 L2・L2h・La・Lb",
                 fontproperties=jp, fontsize=13, fontweight="bold")
    return save(fig, "fig6-2_lengths.png")


# ===========================================================================
# 図 6-5: 部材内定着と仕口内定着
# ===========================================================================
def fig_6_5():
    fig, axes = plt.subplots(1, 2, figsize=(13.5, 5.8))
    # --- (a) 仕口内定着（柱梁接合部に折り曲げ定着） ---
    ax = axes[0]
    # 柱
    ax.add_patch(mpatches.Rectangle((0, 0), 2.6, 6.0, fc="#d9d9d9",
                                     ec="k", lw=1.2))
    ax.text(1.3, 5.5, "柱", fontproperties=jp, ha="center", fontsize=10)
    # 仕口（パネルゾーン）
    ax.add_patch(mpatches.Rectangle((0, 2.4), 2.6, 1.6, fc="#c9d9c0",
                                     ec="k", lw=0.8, hatch="..", zorder=1))
    ax.text(1.3, 0.4, "仕口（柱梁接合部）", fontproperties=jp, ha="center",
            fontsize=8, color="#1f7a1f")
    # 梁
    ax.add_patch(mpatches.Rectangle((2.6, 2.4), 4.5, 1.6, fc="#bcd2ea",
                                     ec="k", lw=1.0))
    ax.text(5.0, 3.2, "梁", fontproperties=jp, ha="center", fontsize=10)
    # 梁上端筋が仕口内に折り曲げ定着
    ax.plot([7.0, 0.7], [3.7, 3.7], color="#c00000", lw=3)
    ax.plot([0.7, 0.7], [3.7, 1.5], color="#c00000", lw=3)
    # 梁下端筋
    ax.plot([7.0, 0.7], [2.7, 2.7], color="#1f7a1f", lw=3)
    ax.plot([0.7, 0.7], [2.7, 4.6], color="#1f7a1f", lw=3)
    ax.annotate("梁主筋を仕口内に\n折り曲げ定着（La）",
                xy=(0.7, 1.8), xytext=(2.4, 0.9),
                fontproperties=jp, fontsize=9, color="#c00000",
                arrowprops=dict(arrowstyle="->", color="#c00000"))
    ax.set_xlim(-0.3, 7.5)
    ax.set_ylim(-0.3, 6.3)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(a) 仕口内定着（柱梁接合部に定着）",
                 fontproperties=jp, fontsize=11, fontweight="bold")

    # --- (b) 部材内定着（梁内に小梁・スラブ筋を直線定着） ---
    ax = axes[1]
    # 大梁
    ax.add_patch(mpatches.Rectangle((0, 0), 2.6, 6.0, fc="#bcd2ea",
                                     ec="k", lw=1.2))
    ax.text(1.3, 5.5, "大梁\n(定着先)", fontproperties=jp, ha="center",
            fontsize=9)
    # 小梁
    ax.add_patch(mpatches.Rectangle((2.6, 2.6), 4.5, 1.2, fc="#cfe0c0",
                                     ec="k", lw=1.0))
    ax.text(5.0, 3.2, "小梁・スラブ", fontproperties=jp, ha="center",
            fontsize=9)
    # 小梁上端筋を大梁内に直線定着（L2）
    ax.plot([7.0, 0.5], [3.5, 3.5], color="#c00000", lw=3)
    ax.annotate("", xy=(0.5, 4.2), xytext=(2.6, 4.2),
                arrowprops=dict(arrowstyle="<|-|>", color="#1f4e79", lw=1.3))
    ax.text(1.55, 4.5, "L2（直線定着）", fontproperties=jp, ha="center",
            fontsize=9, color="#1f4e79", fontweight="bold")
    ax.text(3.5, 0.5, "部材（大梁）の中に\n小梁・スラブ筋を直線定着",
            fontproperties=jp, ha="center", fontsize=9, color="#444")
    ax.set_xlim(-0.3, 7.5)
    ax.set_ylim(-0.3, 6.3)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(b) 部材内定着（大梁内に直線定着）",
                 fontproperties=jp, fontsize=11, fontweight="bold")
    fig.suptitle("図 6-5  部材内定着 と 仕口内定着",
                 fontproperties=jp, fontsize=13, fontweight="bold")
    return save(fig, "fig6-5_joint.png")


figs = {
    "f1": fig_6_1(),
    "f2": fig_6_2(),
    "f5": fig_6_5(),
}
print("figures:", list(figs.keys()))

# ===========================================================================
# Excel 構築
# ===========================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

wb = Workbook()
C_TITLE = "1F4E79"; C_HEAD = "2E75B6"; C_ANS = "E2EFDA"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
f_title = Font(name="MS PGothic", size=15, bold=True, color="FFFFFF")
f_head = Font(name="MS PGothic", size=11, bold=True, color="FFFFFF")
f_body = Font(name="MS PGothic", size=10)
f_ans = Font(name="MS PGothic", size=10, color="375623")
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)


def setup(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False


def title_row(ws, row, text, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = f_title; c.fill = PatternFill("solid", fgColor=C_TITLE)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 30


def head(ws, row, text, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = f_head; c.fill = PatternFill("solid", fgColor=C_HEAD)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 22


def body(ws, row, text, span=8, ans=False, h=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = f_ans if ans else f_body
    c.alignment = wrap
    if ans:
        c.fill = PatternFill("solid", fgColor=C_ANS)
    if h:
        ws.row_dimensions[row].height = h


def table(ws, start_row, headers, rows, col1=1):
    r = start_row
    for j, htxt in enumerate(headers):
        c = ws.cell(r, col1 + j, htxt)
        c.font = Font(name="MS PGothic", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=C_HEAD)
        c.alignment = center; c.border = border
    for data in rows:
        r += 1
        for j, v in enumerate(data):
            c = ws.cell(r, col1 + j, v)
            c.font = f_body
            c.alignment = center if j > 0 else wrap
            c.border = border
            if r % 2 == 0:
                c.fill = PatternFill("solid", fgColor="F2F7FC")
    return r


def put_img(ws, path, anchor, w=None):
    img = XLImage(path)
    if w:
        ratio = w / img.width
        img.width = w; img.height = int(img.height * ratio)
    ws.add_image(img, anchor)


# ---- 目次 ----
ws = wb.active
ws.title = "目次"
setup(ws, [4, 22, 60, 14])
title_row(ws, 1, "鉄筋の定着 問題集（RC マンション設計担当・新人向け）", span=4)
body(ws, 2,
     "テーマ No.6：RC 部材の鉄筋定着。付着による応力伝達機構から、定着長 L2・L2h・La・Lb の"
     "使い分け・算出、一覧表の暗記、計算仮定、部材内/仕口内定着までを通す。"
     "各シートに図、最後に解答。",
     span=4, h=44)
r = table(ws, 4,
          ["No.", "シート", "到達目標", "図"],
          [["6-1", "6-1 応力伝達機構",
            "定着の応力伝達機構を説明できる", "付着・応力分布"],
           ["6-2", "6-2 定着長の種類",
            "L2・L2h・La・Lb の説明・算出ができる", "各定着長"],
           ["6-3", "6-3 L2・L2h 暗記",
            "定着長一覧表の L2・L2h を暗記している", "一覧表"],
           ["6-4", "6-4 計算仮定条件",
            "一覧表の計算仮定条件を理解している", "—"],
           ["6-5", "6-5 部材内と仕口内",
            "部材内定着・仕口内定着を説明できる", "接合部・大梁"]])
body(ws, r + 2,
     "凡例：L2=直線定着長、L2h=フック付き定着長、La=折り曲げ定着の投影定着長、"
     "Lb=基本定着長。数値は『RC 規準』『公共建築工事標準仕様書』等に基づく目安"
     "（fc・鉄筋種別で変わる）。最新版規準・標準図・自社基準を必ず併用すること。",
     span=4, h=58)

# ---- 6-1 ----
ws = wb.create_sheet("6-1 応力伝達機構")
setup(ws, [8, 16, 18, 16, 14, 12, 12, 12])
title_row(ws, 1, "6-1  鉄筋定着の応力伝達機構")
head(ws, 3, "■ 図 6-1  付着による応力伝達")
put_img(ws, figs["f1"], "A4", w=720)
head(ws, 30, "■ 問題 1  機構の穴埋め")
body(ws, 31, "鉄筋に生じた引張力 T は、鉄筋表面の【 ① 】（リブ）の支圧と、鉄筋とコンクリートの"
             "境界に生じる【 ② 】応力 τ によって、定着長にわたって徐々に【 ③ 】へ伝達される。"
             "鉄筋応力 σs は引張端で【 ④ 】、定着端で【 ⑤ 】になる。", h=44)
head(ws, 33, "■ 問題 2  なぜ定着が必要か")
body(ws, 34, "鉄筋の引張力をコンクリートに伝えきれない（定着不足）と、何が起こるか 2 つ挙げよ"
             "（抜け出し／付着割裂ひび割れ）。それぞれが部材性能にどう影響するか 1 行で。",
     h=44)
head(ws, 36, "■ 問題 3  付着を高める要素")
body(ws, 37, "付着性能（許容付着応力度 fb）を高める／低下させる要素を整理せよ。",
     h=20)
r = table(ws, 39,
          ["要素", "付着への影響（記入）"],
          [["コンクリート強度 fc が高い", ""],
           ["鉄筋が異形（節あり）", ""],
           ["かぶり厚さが大きい", ""],
           ["上端筋（コンクリート上部）", ""],
           ["横補強筋（あばら筋・帯筋）が多い", ""]])
body(ws, r + 2, "『上端筋』の付着が下端筋より不利になる理由を、コンクリート打設時の"
                "ブリーディング（水の上昇）から説明せよ。", h=40)

# ---- 6-2 ----
ws = wb.create_sheet("6-2 定着長の種類")
setup(ws, [8, 16, 18, 16, 14, 12, 12, 12])
title_row(ws, 1, "6-2  各種定着長 L2・L2h・La・Lb の説明と算出")
head(ws, 3, "■ 図 6-2  L2・L2h・La・Lb")
put_img(ws, figs["f2"], "A4", w=760)
head(ws, 32, "■ 問題 1  定着長の対応（線結び）")
r = table(ws, 33,
          ["記号", "意味（記入）"],
          [["L2", ""],
           ["L2h", ""],
           ["La", ""],
           ["Lb", ""]])
body(ws, r + 2, "選択肢：A 直線定着長／B フック付き定着長／"
                "C 折り曲げ定着の投影定着長／D 基本定着長（式で計算する元の長さ）", h=32)
head(ws, r + 4, "■ 問題 2  基本定着長 Lb の計算")
body(ws, r + 5, "Lb = (σt / (10·fb)) · db。"
                "σt=295 N/mm²（SD295 の許容引張、短期）、fb=2.0 N/mm²、db=22mm（D22）のとき "
                "Lb を求めよ（mm）。", h=32)
head(ws, r + 7, "■ 問題 3  L2・L2h の関係")
body(ws, r + 8, "(1) フック付き L2h が直線 L2 より短くできる理由を述べよ"
                "（フックの機械的抵抗）。", h=20)
body(ws, r + 9, "(2) 一般に L2h ≒ 0.75·L2 程度とされる。L2=40db のとき L2h は db の何倍か。",
     h=20)
body(ws, r + 10, "(3) フックには『余長（フック先端の直線部）』が必要。"
                 "余長を含めても投影定着長 La に算入しない理由を 1 行で。", h=32)
head(ws, r + 12, "■ 問題 4  鉄筋径と定着長")
body(ws, r + 13, "定着長は鉄筋径 db に比例する（L2 = ○○·db の形）。"
                 "D19（19mm）と D25（25mm）で L2=40db のとき、それぞれの L2 を mm で求め、"
                 "太径ほど定着長が長くなることを確認せよ。", h=40)

# ---- 6-3 ----
ws = wb.create_sheet("6-3 L2・L2h 暗記")
setup(ws, [10, 14, 12, 12, 12, 12, 14])
title_row(ws, 1, "6-3  定着長一覧表（L2・L2h）の暗記", span=7)
head(ws, 3, "■ 定着長一覧表（fc 別・鉄筋種別、db の倍数）", span=7)
body(ws, 4, "下表は『直線定着 L2 / フック付き L2h』の代表値（db の倍数）。"
            "RC 規準・標準仕様書ベースの目安。実務はこの表を暗記して使う。", span=7, h=32)
r = table(ws, 6,
          ["Fc (N/mm²)", "SD295 L2", "SD295 L2h",
           "SD345 L2", "SD345 L2h", "SD390 L2", "SD390 L2h"],
          [["18", "40db", "30db", "40db", "30db", "—", "—"],
           ["21", "35db", "25db", "40db", "30db", "45db", "35db"],
           ["24〜27", "30db", "20db", "35db", "25db", "40db", "30db"],
           ["30〜36", "30db", "20db", "30db", "20db", "35db", "25db"],
           ["39〜45", "25db", "15db", "30db", "20db", "35db", "25db"]])
body(ws, r + 2,
     "※ 上表は『一般的な目安値』。実際の設計では使用する規準・告示・標準図"
     "（例：公共建築工事標準仕様書、各社標準図）の値を用いること。"
     "数値は版・条件で異なるため、必ず最新の表で確認する。",
     span=7, h=44)
head(ws, r + 4, "■ 問題 1  表の穴埋め暗記", span=7)
body(ws, r + 5, "次の条件の L2・L2h を上表から答えよ（db の倍数）。", span=7, h=20)
r2 = table(ws, r + 7,
           ["No.", "Fc", "鉄筋", "L2（記入）", "L2h（記入）"],
           [["(1)", "21", "SD345", "", ""],
            ["(2)", "24", "SD295", "", ""],
            ["(3)", "30", "SD345", "", ""],
            ["(4)", "21", "SD390", "", ""],
            ["(5)", "36", "SD390", "", ""]])
head(ws, r2 + 2, "■ 問題 2  実寸法の算出", span=7)
body(ws, r2 + 3, "Fc=24、SD345、D22（db=22mm）の大梁主筋を直線定着する。"
                 "L2（mm）を求めよ。フック付き L2h なら何 mm か。", span=7, h=32)
head(ws, r2 + 5, "■ 問題 3  傾向の理解", span=7)
body(ws, r2 + 6, "表から読み取れる傾向を 3 つ挙げよ："
                 "(1) Fc が大きいほど定着長は？　(2) 鉄筋強度が高いほど定着長は？　"
                 "(3) フックを付けると定着長は？", span=7, h=40)

# ---- 6-4 ----
ws = wb.create_sheet("6-4 計算仮定条件")
setup(ws, [8, 18, 20, 16, 14, 12, 12])
title_row(ws, 1, "6-4  定着長一覧表における計算仮定条件")
head(ws, 3, "■ 問題 1  一覧表の前提（穴埋め）")
body(ws, 4, "定着長一覧表の数値は、ある仮定条件のもとで Lb を計算し丸めたもの。"
            "主な仮定条件を確認する。", h=32)
r = table(ws, 6,
          ["項目", "一般的な仮定", "理解（記入）"],
          [["鉄筋の引張応力度 σt", "鉄筋の許容引張応力度（または降伏強度）を全強で見込む",
            ""],
           ["許容付着応力度 fb", "Fc と鉄筋位置（上端/その他）で決まる規準値", ""],
           ["鉄筋位置", "『上端筋』と『下端筋・その他』で fb が異なる", ""],
           ["フック", "標準フック（90°/135°/180°）を前提", ""],
           ["かぶり・あき", "標準的なかぶり厚・鉄筋あきを前提", ""],
           ["横補強筋", "標準的な配置を前提", ""]])
head(ws, r + 2, "■ 問題 2  上端筋の割増し")
body(ws, r + 3, "(1) 一覧表で『上端筋』の定着長が他より長く設定される理由を述べよ"
                "（ブリーディングによる付着低下）。", h=32)
body(ws, r + 4, "(2) ここでいう『上端筋』とは、コンクリート打設面から下に何 mm 以上の位置に"
                "ある水平鉄筋を指すか（一般的な目安）。", h=32)
head(ws, r + 6, "■ 問題 3  仮定から外れる場合")
body(ws, r + 7, "次の場合、一覧表の値をそのまま使ってよいか、割増し等が必要かを答えよ。",
     h=20)
r = table(ws, r + 9,
          ["No.", "状況", "扱い（記入）"],
          [["(1)", "かぶり厚さが標準より小さい", ""],
           ["(2)", "軽量コンクリートを使用", ""],
           ["(3)", "鉄筋を密に配置しあきが小さい", ""],
           ["(4)", "エポキシ塗装鉄筋を使用", ""]])
head(ws, r + 2, "■ 問題 4  なぜ仮定を理解する必要があるか")
body(ws, r + 3, "新人が「一覧表の数字だけ」を使うと危険な理由を、"
                "上の仮定条件をふまえて述べよ（標準条件を外れるケースの見落とし）。",
     h=44)

# ---- 6-5 ----
ws = wb.create_sheet("6-5 部材内と仕口内")
setup(ws, [8, 16, 18, 16, 14, 12, 12, 12])
title_row(ws, 1, "6-5  部材内定着 と 仕口内定着")
head(ws, 3, "■ 図 6-5  仕口内定着（左）と部材内定着（右）")
put_img(ws, figs["f5"], "A4", w=760)
head(ws, 28, "■ 問題 1  用語の説明")
body(ws, 29, "(1) 『仕口内定着』とは何か。どの部材のどの鉄筋を、どこに定着することか説明せよ。",
     h=32)
body(ws, 30, "(2) 『部材内定着』とは何か。代表例（小梁・スラブ筋の大梁への定着）で説明せよ。",
     h=32)
head(ws, 32, "■ 問題 2  使い分け")
r = table(ws, 33,
          ["定着の種類", "代表例", "定着長（記入）", "備考（記入）"],
          [["仕口内定着", "梁主筋を柱梁接合部に折り曲げ", "", ""],
           ["部材内定着", "小梁・スラブ筋を大梁内に直線", "", ""]])
head(ws, r + 2, "■ 問題 3  仕口内定着の納まり")
body(ws, r + 3, "(1) 梁主筋を仕口（柱梁接合部）に折り曲げ定着する際、投影定着長 La は"
                "どこから測るか（起点）。", h=32)
body(ws, r + 4, "(2) 大地震時、梁主筋は仕口面で降伏する可能性がある。"
                "仕口内定着で『折り曲げ後の余長（縦の部分）』が重要になる理由を述べよ。",
     h=40)
body(ws, r + 5, "(3) 最上階の柱頭や、外柱の仕口で定着が特に難しくなる理由を 1 行で"
                "（直交方向の梁主筋との干渉、定着長の確保）。", h=32)
head(ws, r + 7, "■ 問題 4  通し配筋との比較")
body(ws, r + 8, "中柱の仕口では、梁主筋を『仕口内に定着』する代わりに『柱を貫通して通し配筋』"
                "とすることもある。通し配筋のメリットと、適用できる条件を述べよ。", h=44)
head(ws, r + 10, "■ 問題 5  マンション実務")
body(ws, r + 11, "RC マンションのラーメン部で、(1) 大梁主筋を外柱に定着するとき、"
                 "(2) 小梁主筋を大梁に定着するとき、それぞれ L2/L2h/La のどれを使うか。"
                 "また定着長不足になりがちな箇所と対策（フック化・余長確保・梁せい調整）を述べよ。",
     h=58)

# ---- 解答 ----
ws = wb.create_sheet("解答")
setup(ws, [4, 22, 60, 14])
title_row(ws, 1, "解答・解説", span=4)
row = 2


def ah(t):
    global row
    head(ws, row, t, span=4); row += 1


def an(t, h=None):
    global row
    body(ws, row, t, span=4, ans=True, h=h); row += 1


ah("6-1  応力伝達機構")
an("問1：①節（リブ） ②付着 ③コンクリート ④最大 ⑤ゼロ（0）。"
   "異形鉄筋は節の支圧が付着の主役。", h=44)
an("問2：①抜け出し＝鉄筋がコンクリートから滑り出て応力を伝えられず、部材の曲げ耐力が出ない。"
   "②付着割裂ひび割れ＝かぶりコンクリートが割れて剥落し、急激な耐力低下（脆性的）。"
   "いずれも設計で期待した耐力・靭性が確保できなくなる。", h=58)
an("問3：fc 高い→付着↑。異形→付着↑（節の支圧）。かぶり大→付着↑（割裂しにくい）。"
   "上端筋→付着↓（ブリーディングで下部に空隙）。横補強筋多い→付着↑（割裂を拘束）。"
   "上端筋：打設したコンクリートの水分が上昇（ブリーディング）し、上端の鉄筋下面に"
   "水膜・空隙ができて付着が低下するため、定着長を割増す。", h=72)

ah("6-2  定着長の種類")
an("問1：L2=A 直線定着長、L2h=B フック付き定着長、La=C 折り曲げ定着の投影定着長、"
   "Lb=D 基本定着長。", h=32)
an("問2：Lb = (σt/(10·fb))·db = (295/(10·2.0))·22 = (295/20)·22 = 14.75·22 = 324.5 ≒ 325mm。",
   h=44)
an("問3：(1)フック先端が機械的に引っかかり、付着＋支圧で抵抗するため直線より短くできる。"
   "(2)L2h=0.75·40db=30db。(3)余長はフックの形状を保つための最小直線部で、"
   "定着耐力としては投影長 La で評価するため二重計上しない。", h=58)
an("問4：L2=40db。D19→40×19=760mm、D25→40×25=1,000mm。"
   "太径ほど定着長が長くなり、納まり（梁せい・柱せい）に効く。", h=44)

ah("6-3  L2・L2h 暗記")
an("問1：(1)Fc21/SD345：L2=40db、L2h=30db。(2)Fc24/SD295：L2=30db、L2h=20db。"
   "(3)Fc30/SD345：L2=30db、L2h=20db。(4)Fc21/SD390：L2=45db、L2h=35db。"
   "(5)Fc36/SD390：L2=35db、L2h=25db。", h=58)
an("問2：Fc24/SD345/D22：L2=35db=35×22=770mm。L2h=25db=25×22=550mm。", h=32)
an("問3：(1)Fc 大→定着長 短（付着が増す）。(2)鉄筋強度 高→定着長 長（伝える力が大きい）。"
   "(3)フック付き→定着長 短（機械的抵抗が加わる）。", h=44)

ah("6-4  計算仮定条件")
an("問1：σt＝鉄筋の許容引張（短期は F、または降伏強度）を見込む。"
   "fb＝Fc と鉄筋位置で決まる規準値。上端筋とその他で fb が違う。"
   "標準フック・標準かぶり・標準的な横補強を前提に表が作られている。", h=58)
an("問2：(1)コンクリート打設時のブリーディングで上端鉄筋の下面に水・空隙ができ付着が"
   "低下するため、上端筋は定着長を割増す。(2)一般に、その鉄筋の下に 300mm 以上の"
   "コンクリートが打ち込まれる位置の水平鉄筋を『上端筋』として扱う（規準による）。",
   h=58)
an("問3：(1)かぶり小→割裂しやすく付着低下→割増し必要。(2)軽量コンクリート→付着が"
   "普通コンクリートより小さく割増し必要。(3)あき小→割裂しやすい→割増し or あき確保。"
   "(4)エポキシ塗装→付着低下し割増し必要。いずれも一覧表（標準条件）をそのまま使えない。",
   h=72)
an("問4：一覧表は『標準条件で計算した値』なので、軽量コン・小かぶり・密配筋・塗装鉄筋など"
   "条件を外れるケースでは危険側になる。仮定を理解していないと、外れた条件を見落として"
   "定着不足の設計をしてしまう。表は『前提つきの早見』と認識する。", h=58)

ah("6-5  部材内/仕口内")
an("問1：(1)仕口内定着＝柱梁接合部（仕口・パネルゾーン）の中に、梁主筋（や柱主筋）を"
   "折り曲げて定着すること。(2)部材内定着＝ある部材（大梁等）の断面の中に、"
   "別の部材の鉄筋（小梁・スラブ筋）を直線などで定着すること。", h=58)
an("問2：仕口内定着＝梁主筋を柱梁接合部に折り曲げ→投影定着長 La（フェイスから測る）。"
   "部材内定着＝小梁・スラブ筋を大梁内に直線→L2（直線定着長）。"
   "仕口内は折り曲げ＋余長で納め、部材内は直線 or フックで納める。", h=58)
an("問3：(1)La は仕口面（柱フェイス）から折り曲げ起点までの水平投影長で測る。"
   "(2)大地震時に梁主筋が仕口面で降伏・伸びるため、折り曲げ後の縦の余長が"
   "抜け出し防止に効く（折り曲げ定着の信頼性を担保）。"
   "(3)外柱・最上階は梁が片側のみ・直交梁主筋と干渉し、所定の定着長・余長の確保が"
   "難しいため（柱せい不足）。", h=72)
an("問4：通し配筋のメリット＝定着の納まりが不要、施工が単純、接合部の信頼性が高い。"
   "条件＝柱せいが梁主筋径の一定倍（例：柱せい ≥ 20db 程度）以上あり、貫通させても"
   "付着・定着が確保できること。中柱で梁が両側に連続する場合に有効。", h=58)
an("問5：(1)大梁主筋を外柱に定着＝仕口内に折り曲げ定着 La（必要に応じフック L2h）。"
   "(2)小梁主筋を大梁に定着＝部材内に直線定着 L2（余長取れなければ L2h）。"
   "定着不足になりがちな箇所＝外柱・最上階柱頭（柱せい不足）、小梁が浅く L2 が取れない箇所。"
   "対策＝フック化（L2h で短縮）、定着余長の確保、梁せい・柱せいの調整、定着板の使用。",
   h=72)

XLSX = os.path.join(OUT, "鉄筋定着問題集.xlsx")
wb.save(XLSX)
print("saved:", XLSX)
