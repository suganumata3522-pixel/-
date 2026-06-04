# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "アンケート"
ws.sheet_view.showGridLines = False

# ===== カラーパレット =====
NAVY    = "1B2A4A"   # 濃紺（タイトル）
BLUE    = "2F6FED"   # アクセントブルー
LBLUE   = "EAF1FF"   # 設問見出し背景
PALE    = "F7F9FC"   # 本文背景
LINE    = "D6DEE8"   # 罫線
INK     = "1B2A4A"   # 文字濃
GRAY    = "5B6B7F"   # 補足文字
ANSBG   = "FFFFFF"   # 回答欄背景
ANSEDGE = "BFD0EA"

# ===== フォント =====
F = "Yu Gothic UI"
f_title = Font(name=F, size=18, bold=True, color="FFFFFF")
f_sub   = Font(name=F, size=10, color="D7E2F5")
f_lead  = Font(name=F, size=10, color=GRAY)
f_qno   = Font(name=F, size=15, bold=True, color=BLUE)
f_qttl  = Font(name=F, size=12, bold=True, color=INK)
f_body  = Font(name=F, size=10, color=INK)
f_opt   = Font(name=F, size=10.5, color=INK)
f_note  = Font(name=F, size=9, color=GRAY)
f_ansl  = Font(name=F, size=9, bold=True, color=BLUE)
f_foot  = Font(name=F, size=10, color=GRAY)

# ===== 罫線 =====
thin   = Side(style="thin", color=LINE)
accent = Side(style="thick", color=BLUE)
ansb   = Side(style="thin", color=ANSEDGE)
box      = Border(left=thin, right=thin, top=thin, bottom=thin)
ans_box  = Border(left=ansb, right=ansb, top=ansb, bottom=ansb)

# ===== 配置 =====
wrap   = Alignment(wrap_text=True, vertical="top")
wrapc  = Alignment(wrap_text=True, vertical="center")
ctr    = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_c = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ===== 列幅（A:余白 / B:設問No / C..F:本文 / G:余白）=====
widths = {"A": 2.5, "B": 6, "C": 22, "D": 22, "E": 22, "F": 16, "G": 2.5}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

def fillrange(r1, c1, r2, c2, color):
    pf = PatternFill("solid", fgColor=color)
    for rr in range(r1, r2 + 1):
        for cc in range(c1, c2 + 1):
            ws.cell(rr, cc).fill = pf

def border_range(r1, c1, r2, c2, bd):
    for rr in range(r1, r2 + 1):
        for cc in range(c1, c2 + 1):
            ws.cell(rr, cc).border = bd

r = 1

# ================= ヘッダーバンド =================
ws.row_dimensions[r].height = 8           # 上余白
r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r+2, end_column=6)
fillrange(r, 2, r+2, 6, NAVY)
c = ws.cell(r, 2, "RC標準雑詳細図WG　方針検討アンケート")
c.font = f_title; c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
ws.row_dimensions[r].height = 14
ws.row_dimensions[r+1].height = 18
ws.row_dimensions[r+2].height = 10
# サブタイトル（同じバンド内の下段）
c2 = ws.cell(r+2, 2)  # already merged; set via top-left only -> use separate line
r += 3

# サブコピー帯
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
fillrange(r, 2, r, 6, BLUE)
c = ws.cell(r, 2, "Standard Detail Drawing  /  Policy Survey")
c.font = f_sub; c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
ws.row_dimensions[r].height = 16
r += 2

# ================= リード文 =================
ws.merge_cells(start_row=r, start_column=2, end_row=r+1, end_column=6)
c = ws.cell(r, 2,
    "お疲れ様です。本WGでは今後の作図方針を決定するにあたり、部内の皆様のご意見を伺いたく存じます。\n"
    "別紙資料をご参照のうえ、各設問の回答欄へのご記入をお願いいたします。")
c.font = f_lead; c.alignment = wrap
ws.row_dimensions[r].height = 16; ws.row_dimensions[r+1].height = 16
r += 2

# 回答者欄
ws.cell(r, 2, "回答者").font = f_ansl
ws.cell(r, 2).alignment = ctr
ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
fillrange(r, 3, r, 6, ANSBG)
border_range(r, 3, r, 6, ans_box)
ws.row_dimensions[r].height = 22
r += 2

dv_list = []  # data validations to add later

def heading(no, title):
    """設問見出し（No + タイトル）"""
    global r
    fillrange(r, 2, r, 6, LBLUE)
    border_range(r, 2, r, 6, Border(bottom=accent))
    ws.cell(r, 2, no).font = f_qno
    ws.cell(r, 2).alignment = ctr
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    c = ws.cell(r, 3, title)
    c.font = f_qttl; c.alignment = Alignment(vertical="center")
    ws.row_dimensions[r].height = 28
    r += 1

def bodytext(text, h=34):
    global r
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    fillrange(r, 2, r, 6, PALE)
    c = ws.cell(r, 2, text)
    c.font = f_body
    c.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
    ws.row_dimensions[r].height = h
    r += 1

def options(items, h=22):
    """選択肢を1行ずつ。items=[(label,text),...]"""
    global r
    for label, text in items:
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        ws.cell(r, 2, label).font = f_opt
        ws.cell(r, 2).alignment = ctr
        c = ws.cell(r, 3, text)
        c.font = f_opt; c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = h
        r += 1

def answer(label="回答", h=46, dropdown=None):
    """設問直下の回答欄"""
    global r
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    c = ws.cell(r, 2, "▼ " + label)
    c.font = f_ansl
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[r].height = 16
    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    fillrange(r, 2, r, 6, ANSBG)
    border_range(r, 2, r, 6, ans_box)
    ws.cell(r, 2).alignment = wrap
    if dropdown:
        dv = DataValidation(type="list",
                            formula1='"%s"' % ",".join(dropdown),
                            allow_blank=True)
        dv.prompt = "選択してください"
        dv.promptTitle = "回答"
        ws.add_data_validation(dv)
        dv.add(ws.cell(r, 2))
    ws.row_dimensions[r].height = h
    r += 1

def spacer(h=8):
    global r
    ws.row_dimensions[r].height = h
    r += 1

# ===================== 設問① =====================
heading("①", "雑詳細図の利用形式について")
bodytext("雑詳細図の運用方法として、以下のどちらが望ましいとお考えですか。", h=22)
options([
    ("A", "必要な納まりを都度抜粋して使う「カタログ的な利用」とする"),
    ("B", "あらかじめレイアウトまで作り込み、物件で使用しない箇所は斜線で消す形式とする"),
])
answer("回答（A / B）とその理由", dropdown=["A", "B"])
spacer()

# ===================== 設問② =====================
heading("②", "標準図・部材リストとの情報の重複について")
bodytext("伝えたいポイントを明確にするため、標準図や部材リストから読み取れる情報は極力省略する方向で検討しています。"
         "（省略により情報量が極端に薄くなる場合は記載を残すことを想定）", h=34)
options([
    ("A", "標準図・部材リストと重複する情報は省略してよい"),
    ("B", "これまで通り、重複する情報も記載しておくべき"),
])
answer("回答（A / B）とその理由", dropdown=["A", "B"])
spacer()

# ===================== 設問③ =====================
heading("③", "拡大図の縮尺について")
bodytext("拡大図の縮尺として、最も見やすいと感じるものを一つお選びください。", h=22)
options([
    ("□", "1/20"),
    ("□", "1/15"),
    ("□", "1/12"),
    ("□", "1/10"),
], h=18)
answer("回答（縮尺）とその理由", dropdown=["1/20", "1/15", "1/12", "1/10"])
spacer()

# ===================== 設問④ =====================
heading("④", "拡大図の記載方針について")
bodytext("各種手摺配筋詳細図、片持ちスラブ先端壁・手摺配筋詳細図、パラペット配筋詳細図には既に拡大図を記載しています。"
         "図面間の濃度差を統一する観点から、以降の図面についても下記方針で進めたいと考えています。", h=40)
options([
    ("(1)", "配筋が細かく分かりにくい図面に拡大図を併記する方針に賛成か／より限定的にすべきか"),
    ("(2)", "「拡大図が必要」と判断する基準（例：鉄筋本数・定着継手の有無・納まりの複雑さ・縮尺 等）"),
])
answer("(1) 方針への賛否", h=30, dropdown=["方針どおりで良い", "より限定的にすべき"])
answer("(2) 必要と判断する基準についてのご意見", h=46)
spacer()

# ===================== フッター =====================
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
c = ws.cell(r, 2, "ご協力のほどよろしくお願い申し上げます。")
c.font = f_foot; c.alignment = Alignment(vertical="center")
ws.row_dimensions[r].height = 24
r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
fillrange(r, 2, r, 6, NAVY)
ws.row_dimensions[r].height = 5

# ===== 印刷設定 =====
ws.print_options.horizontalCentered = True
ws.page_setup.orientation = "portrait"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = ws.page_margins.right = 0.3
ws.page_margins.top = ws.page_margins.bottom = 0.4

wb.save("RC標準雑詳細図WG_方針アンケート.xlsx")
print("saved")
