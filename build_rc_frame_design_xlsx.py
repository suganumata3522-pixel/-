# -*- coding: utf-8 -*-
"""架構の耐震設計とモデル化 問題集（図つき）Excel 生成スクリプト。
出力: docs/rc_frame_design/架構モデル化問題集.xlsx
No.4-1 ラーメン架構の耐震設計 / 4-2 柱梁のモデル化(線材置換・剛域)
No.4-3 耐震壁架構の耐震設計 / 4-4 耐震壁のモデル化(エレメント置換・開口)
No.4-5 耐震要素の平面・立面バランス
RC造マンションの設計担当を想定。
"""
import os
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib import font_manager as fm

FONT_PATH = "/usr/share/fonts/truetype/fonts-japanese-gothic.ttf"
jp = fm.FontProperties(fname=FONT_PATH)
fm.fontManager.addfont(FONT_PATH)
plt.rcParams["font.family"] = jp.get_name()
plt.rcParams["axes.unicode_minus"] = False

OUT = "docs/rc_frame_design"
FIG = os.path.join(OUT, "figures")
os.makedirs(FIG, exist_ok=True)


def save(fig, name):
    p = os.path.join(FIG, name)
    fig.savefig(p, dpi=130, bbox_inches="tight")
    plt.close(fig)
    return p


# ===========================================================================
# 図 4-1: ラーメン架構の地震時挙動と崩壊形
# ===========================================================================
def fig_4_1():
    fig, axes = plt.subplots(1, 3, figsize=(15, 5.4))
    bays = [0, 6, 12]
    h = 3.2
    ns = 3

    def grid(ax, lw=0.8, color="lightgray", ls="--", off=None):
        off = off or [0] * (ns + 1)
        for x in bays:
            pts = [(x + off[i], i * h) for i in range(ns + 1)]
            ax.plot(*zip(*pts), color=color, lw=lw, ls=ls, zorder=1)
        for i in range(1, ns + 1):
            ax.plot([b + off[i] for b in bays], [i * h] * 3,
                    color=color, lw=lw, ls=ls, zorder=1)

    # --- (a) 水平力と変形 ---
    ax = axes[0]
    grid(ax)  # 無変形
    off = [0, 0.55, 0.95, 1.2]
    for x in bays:
        pts = [(x + off[i], i * h) for i in range(ns + 1)]
        ax.plot(*zip(*pts), color="#1f4e79", lw=2.2, zorder=3)
    for i in range(1, ns + 1):
        ax.plot([b + off[i] for b in bays], [i * h] * 3,
                color="#1f4e79", lw=2.2, zorder=3)
    # 地震力（上ほど大）
    for i, ln in zip(range(1, ns + 1), [1.2, 1.8, 2.4]):
        y = i * h
        ax.annotate("", xy=(off[i] - 0.1, y), xytext=(off[i] - 0.1 - ln, y),
                    arrowprops=dict(arrowstyle="-|>", color="#c00000", lw=2))
        ax.text(off[i] - 0.3 - ln, y, f"P{i}", fontproperties=jp,
                fontsize=9, color="#c00000", ha="right", va="center")
    ax.text(6, -1.4, "地震力の流れ：床スラブ → 大梁 → 柱 → 基礎 → 地盤",
            fontproperties=jp, ha="center", fontsize=9, color="#444")
    ax.text(6, -2.3, "柱・梁の曲げ剛性と剛接合の節点で水平力に抵抗",
            fontproperties=jp, ha="center", fontsize=9, color="#444")
    ax.set_xlim(-4.5, 14.5)
    ax.set_ylim(-3.0, ns * h + 1.2)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(a) 水平力を受けるラーメンの変形",
                 fontproperties=jp, fontsize=11, fontweight="bold")

    # --- (b) 全体崩壊形（GOOD） ---
    ax = axes[1]
    grid(ax, color="#999", ls="-", lw=1.6)
    # 梁端ヒンジ
    for i in range(1, ns + 1):
        for x in bays:
            for dx in ([0.5] if x == 0 else [-0.5] if x == 12 else [-0.5, 0.5]):
                ax.plot(x + dx, i * h, "o", mfc="white", mec="#c00000",
                        ms=9, mew=2, zorder=5)
    # 柱脚ヒンジ
    for x in bays:
        ax.plot(x, 0.15, "o", mfc="white", mec="#c00000", ms=9, mew=2, zorder=5)
    ax.text(6, -1.4, "ヒンジ（○）が梁端＋柱脚に分散\n→ 建物全体でエネルギー吸収（粘り強い）",
            fontproperties=jp, ha="center", fontsize=9, color="#1f7a1f")
    ax.set_xlim(-2.5, 14.5)
    ax.set_ylim(-2.6, ns * h + 1.2)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(b) GOOD 梁降伏先行・全体崩壊形",
                 fontproperties=jp, fontsize=11, fontweight="bold")

    # --- (c) 層崩壊形（NG） ---
    ax = axes[2]
    # 1階だけ大きく変形、上は剛体移動
    d1 = 1.6
    off2 = [0, d1, d1, d1]
    for x in bays:
        pts = [(x + off2[i], i * h) for i in range(ns + 1)]
        ax.plot(*zip(*pts), color="#999", lw=1.6, zorder=2)
    for i in range(1, ns + 1):
        ax.plot([b + off2[i] for b in bays], [i * h] * 3,
                color="#999", lw=1.6, zorder=2)
    # 1階柱頭・柱脚ヒンジ
    for x in bays:
        ax.plot(x, 0.15, "o", mfc="#c00000", mec="#c00000", ms=9, zorder=5)
        ax.plot(x + d1, h - 0.15, "o", mfc="#c00000", mec="#c00000",
                ms=9, zorder=5)
    ax.text(6, -1.4, "ヒンジが1階柱の上下に集中\n→ 1階のみ変形が集中し層崩壊（脆い）",
            fontproperties=jp, ha="center", fontsize=9, color="#c00000")
    ax.set_xlim(-2.5, 14.5)
    ax.set_ylim(-2.6, ns * h + 1.2)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(c) NG 柱降伏・層崩壊形",
                 fontproperties=jp, fontsize=11, fontweight="bold")
    fig.suptitle("図 4-1  ラーメン架構の耐震設計 ── 変形と崩壊メカニズム",
                 fontproperties=jp, fontsize=13, fontweight="bold")
    return save(fig, "fig4-1_frame.png")


# ===========================================================================
# 図 4-2: 線材置換と剛域、スラブ協力幅
# ===========================================================================
def fig_4_2():
    fig, axes = plt.subplots(1, 2, figsize=(13.5, 5.6),
                              gridspec_kw={"width_ratios": [1.5, 1.0]})
    # --- (a) 実形状 → 線材モデル + 剛域 ---
    ax = axes[0]
    L = 7.0          # 軸線スパン
    cb = 0.8         # 柱せい 800
    bd = 0.7         # 梁せい 700
    yb = 4.6         # 実形状の梁中心高さ
    # 実形状（上段）
    for x in [0, L]:
        ax.add_patch(mpatches.Rectangle((x - cb / 2, yb - 2.0), cb, 4.0,
                                         fc="#d9d9d9", ec="k", lw=1.0))
    ax.add_patch(mpatches.Rectangle((cb / 2, yb - bd / 2), L - cb, bd,
                                     fc="#bcd2ea", ec="k", lw=1.0))
    # パネルゾーン
    for x in [0, L]:
        ax.add_patch(mpatches.Rectangle((x - cb / 2, yb - bd / 2), cb, bd,
                                         fc="#8c8c8c", ec="k", lw=0.8,
                                         hatch="xx", zorder=3))
    ax.text(L / 2, yb + 1.0, "実形状：柱 800角、梁 500×700、スパン L=7,000",
            fontproperties=jp, ha="center", fontsize=9)
    ax.annotate("パネルゾーン\n(柱梁重複部)", xy=(0, yb + bd / 2),
                xytext=(-2.1, yb + 1.5), fontproperties=jp, fontsize=8,
                color="#444",
                arrowprops=dict(arrowstyle="->", color="#444"))
    # 線材モデル（下段）
    ym = 1.0
    ax.plot([0, L], [ym, ym], color="#1f4e79", lw=1.8)
    # 剛域（太線）長さ 0.225
    rz = cb / 2 - bd / 4  # 0.4 - 0.175 = 0.225
    for x, s in [(0, 1), (L, -1)]:
        ax.plot([x, x + s * rz], [ym, ym], color="k", lw=6,
                solid_capstyle="butt", zorder=4)
        ax.plot([x, x], [ym - 0.5, ym + 0.5], color="#1f4e79", lw=1.8)
        ax.plot(x, ym, "s", color="#1f4e79", ms=6)
    # 寸法注記
    ax.annotate("剛域（黒太線）= 柱せい/2 − 梁せい/4 = 400 − 175 = 225",
                xy=(rz / 2, ym), xytext=(1.4, ym - 1.3),
                fontproperties=jp, fontsize=8.5, color="#c00000",
                arrowprops=dict(arrowstyle="->", color="#c00000"))
    ax.annotate("", xy=(rz, ym + 0.55), xytext=(L - rz, ym + 0.55),
                arrowprops=dict(arrowstyle="<|-|>", color="#1f7a1f", lw=1.2))
    ax.text(L / 2, ym + 0.85, "可とう長さ（曲げ変形する部分）= 7,000 − 2×225 = 6,550",
            fontproperties=jp, ha="center", fontsize=8.5, color="#1f7a1f")
    # フェイスとD/4 の説明（実形状側）
    ax.annotate("柱フェイス", xy=(cb / 2, yb - bd / 2 - 0.05),
                xytext=(1.6, yb - 1.7), fontproperties=jp, fontsize=8,
                arrowprops=dict(arrowstyle="->", color="#444"))
    ax.annotate("剛域端 = フェイスから\n梁せい/4 (=175) 内側",
                xy=(cb / 2 + bd / 4, yb), xytext=(4.2, yb - 1.9),
                fontproperties=jp, fontsize=8, color="#c00000",
                arrowprops=dict(arrowstyle="->", color="#c00000"))
    ax.annotate("", xy=(-1.3, ym), xytext=(-1.3, yb),
                arrowprops=dict(arrowstyle="-|>", color="#777", lw=1.5))
    ax.text(-1.6, (ym + yb) / 2, "線材置換\n(軸線でモデル化)", fontproperties=jp,
            rotation=90, fontsize=8.5, va="center", ha="center", color="#777")
    ax.set_xlim(-2.3, L + 1.3)
    ax.set_ylim(-0.9, yb + 2.4)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(a) 線材置換と剛域（柱800角・梁せい700 の例）",
                 fontproperties=jp, fontsize=11, fontweight="bold")

    # --- (b) スラブ協力幅（T形梁） ---
    ax = axes[1]
    bw, D, t, B = 1.0, 1.4, 0.36, 3.4   # 比率で描画
    x0 = 0
    # スラブ
    ax.add_patch(mpatches.Rectangle((x0 - B / 2, D - t), B, t,
                                     fc="#bcd2ea", ec="k", lw=1.0))
    # ウェブ
    ax.add_patch(mpatches.Rectangle((x0 - bw / 2, 0), bw, D - t,
                                     fc="#d9d9d9", ec="k", lw=1.0))
    # 寸法
    ax.annotate("", xy=(-B / 2, D + 0.25), xytext=(B / 2, D + 0.25),
                arrowprops=dict(arrowstyle="<|-|>", color="#c00000", lw=1.2))
    ax.text(0, D + 0.42, "協力幅 B（スラブが梁と一体で効く幅）",
            fontproperties=jp, ha="center", fontsize=8.5, color="#c00000")
    ax.annotate("", xy=(B / 2 + 0.25, D - t), xytext=(B / 2 + 0.25, D),
                arrowprops=dict(arrowstyle="<|-|>", color="k", lw=1))
    ax.text(B / 2 + 0.4, D - t / 2, "スラブ t", fontproperties=jp,
            fontsize=8, va="center")
    ax.annotate("", xy=(bw / 2 + 0.25, 0), xytext=(bw / 2 + 0.25, D - t),
                arrowprops=dict(arrowstyle="<|-|>", color="k", lw=1))
    ax.text(bw / 2 + 0.4, (D - t) / 2, "梁せい D", fontproperties=jp,
            fontsize=8, va="center")
    ax.text(0, -0.45, "長方形梁に対する剛性増大率 φ ≒ 1.5〜2.0\n"
                      "（両側スラブ付き大梁の目安）",
            fontproperties=jp, ha="center", fontsize=9, color="#1f4e79")
    ax.set_xlim(-2.4, 3.0)
    ax.set_ylim(-1.0, 2.4)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(b) スラブ協力幅と T 形梁の剛性増大",
                 fontproperties=jp, fontsize=11, fontweight="bold")
    fig.suptitle("図 4-2  柱・梁のモデル化 ── 線材置換・剛域・スラブ協力幅",
                 fontproperties=jp, fontsize=13, fontweight="bold")
    return save(fig, "fig4-2_modeling.png")


# ===========================================================================
# 図 4-3: 耐震壁とフレームの変形モード・分担
# ===========================================================================
def fig_4_3():
    fig, axes = plt.subplots(1, 3, figsize=(14.5, 5.6),
                              gridspec_kw={"width_ratios": [1, 1, 1.3]})
    ns, h = 6, 1.0
    HT = ns * h

    # --- (a) 連層壁：曲げ型 ---
    ax = axes[0]
    ax.add_patch(mpatches.Rectangle((-0.5, 0), 1.0, HT, fc="#e8c9ce",
                                     ec="k", lw=1.2, hatch="//"))
    ys = np.linspace(0, HT, 60)
    xs = 1.8 * (ys / HT) ** 2
    ax.plot(xs + 0.9, ys, color="#c00000", lw=2.2)
    ax.annotate("", xy=(2.9, HT), xytext=(0.9, HT),
                arrowprops=dict(arrowstyle="-|>", color="#c00000", lw=1.4))
    ax.text(0.4, -0.8, "曲げ型：上層ほど層間変形が大きい\n（片持ち梁の曲げと同じ）",
            fontproperties=jp, ha="center", fontsize=9, color="#444")
    ax.set_xlim(-1.6, 3.4)
    ax.set_ylim(-1.5, HT + 0.8)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(a) 連層耐震壁＝曲げ型", fontproperties=jp,
                 fontsize=11, fontweight="bold")

    # --- (b) ラーメン：せん断型 ---
    ax = axes[1]
    xs2 = 2.0 * np.sqrt(np.linspace(0, 1, ns + 1))
    for i in range(ns + 1):
        ax.plot([xs2[i] - 0.7, xs2[i] + 0.7], [i * h, i * h],
                color="#1f4e79", lw=1.8)
    for i in range(ns):
        for dx in [-0.7, 0.7]:
            ax.plot([xs2[i] + dx, xs2[i + 1] + dx], [i * h, (i + 1) * h],
                    color="#1f4e79", lw=1.8)
    ax.text(1.0, -0.8, "せん断型：下層ほど層間変形が大きい\n（各層が水平にずれる）",
            fontproperties=jp, ha="center", fontsize=9, color="#444")
    ax.set_xlim(-1.6, 3.6)
    ax.set_ylim(-1.5, HT + 0.8)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(b) ラーメン＝せん断型", fontproperties=jp,
                 fontsize=11, fontweight="bold")

    # --- (c) 併用時の分担 ---
    ax = axes[2]
    floors = np.arange(1, ns + 1)
    wall_share = np.array([88, 82, 74, 64, 50, 30])
    frame_share = 100 - wall_share
    ax.plot(wall_share, floors, "o-", color="#c00000", lw=2, label="耐震壁")
    ax.plot(frame_share, floors, "s-", color="#1f4e79", lw=2, label="ラーメン")
    ax.set_xlabel("層せん断力の負担率 (%)", fontproperties=jp)
    ax.set_ylabel("階", fontproperties=jp)
    ax.set_yticks(floors)
    ax.set_xlim(0, 100)
    ax.grid(alpha=0.3)
    ax.legend(prop=jp, fontsize=9)
    ax.set_title("(c) 壁とフレームの分担（高さ方向）\n"
                 "下層は壁が大半を負担、上層はフレームの負担が増える",
                 fontproperties=jp, fontsize=10, fontweight="bold")
    fig.suptitle("図 4-3  耐震壁架構の挙動 ── 変形モードと荷重分担",
                 fontproperties=jp, fontsize=13, fontweight="bold")
    return save(fig, "fig4-3_wall_frame.png")


# ===========================================================================
# 図 4-4: 壁のエレメント置換と開口周比
# ===========================================================================
def fig_4_4():
    fig, axes = plt.subplots(1, 2, figsize=(13.5, 5.6))
    # --- (a) エレメント置換 ---
    ax = axes[0]
    # 実形状（左）
    x0 = 0
    for lv in range(3):
        y = lv * 3.0
        ax.add_patch(mpatches.Rectangle((x0, y), 0.6, 3.0, fc="#d9d9d9",
                                         ec="k", lw=0.8))
        ax.add_patch(mpatches.Rectangle((x0 + 5.4, y), 0.6, 3.0, fc="#d9d9d9",
                                         ec="k", lw=0.8))
        ax.add_patch(mpatches.Rectangle((x0 + 0.6, y + 0.4), 4.8, 2.2,
                                         fc="#e8c9ce", ec="k", lw=0.8,
                                         hatch="//"))
        ax.add_patch(mpatches.Rectangle((x0, y + 2.6), 6.0, 0.4, fc="#999",
                                         ec="k", lw=0.6))
    ax.text(x0 + 3.0, -0.8, "実形状：両側柱＋壁板＋梁",
            fontproperties=jp, ha="center", fontsize=9)
    # 置換モデル（右）
    xm = 9.5
    for lv in range(3):
        y = lv * 3.0
        # 剛梁
        ax.plot([xm, xm + 6.0], [y + 2.8, y + 2.8], color="k", lw=5,
                solid_capstyle="butt")
    # 側柱
    for dx in [0, 6.0]:
        ax.plot([xm + dx, xm + dx], [0, 9.0], color="#1f4e79", lw=1.8)
    # 壁柱（中央・太線）
    ax.plot([xm + 3.0, xm + 3.0], [0, 9.0], color="#c00000", lw=5)
    ax.text(xm + 3.0, -0.8, "置換モデル：壁柱(太線)＋剛梁(黒太線)",
            fontproperties=jp, ha="center", fontsize=9)
    ax.annotate("壁柱：壁全体の断面性能\n(A・I) を持つ仮想柱",
                xy=(xm + 3.0, 4.6), xytext=(xm + 3.9, 6.6),
                fontproperties=jp, fontsize=8, color="#c00000",
                arrowprops=dict(arrowstyle="->", color="#c00000"))
    ax.annotate("剛梁：梁レベルで\n壁と柱を剛につなぐ",
                xy=(xm + 1.4, 8.8), xytext=(xm + 0.2, 10.2),
                fontproperties=jp, fontsize=8,
                arrowprops=dict(arrowstyle="->", color="#333"))
    ax.annotate("", xy=(xm - 0.7, 4.5), xytext=(x0 + 7.0, 4.5),
                arrowprops=dict(arrowstyle="-|>", color="#777", lw=2))
    ax.text((x0 + 7.0 + xm - 0.7) / 2, 5.0, "エレメント置換",
            fontproperties=jp, ha="center", fontsize=9, color="#777")
    ax.set_xlim(-0.8, 16.5)
    ax.set_ylim(-1.6, 11.0)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(a) 耐震壁のエレメント置換（壁柱＋剛梁）",
                 fontproperties=jp, fontsize=11, fontweight="bold")

    # --- (b) 開口周比 ---
    ax = axes[1]
    l, h = 6.0, 3.0
    l0, h0 = 1.5, 1.8
    ax.add_patch(mpatches.Rectangle((0, 0), l, h, fc="#e8c9ce", ec="k",
                                     lw=1.2, hatch="//"))
    ox, oy = (l - l0) / 2, (h - h0) / 2
    ax.add_patch(mpatches.Rectangle((ox, oy), l0, h0, fc="white", ec="k",
                                     lw=1.2))
    ax.text(l / 2, h / 2, "開口", fontproperties=jp, ha="center",
            va="center", fontsize=10, color="#444")
    # 寸法
    ax.annotate("", xy=(0, -0.4), xytext=(l, -0.4),
                arrowprops=dict(arrowstyle="<|-|>", color="k", lw=1))
    ax.text(l / 2, -0.7, "壁の内法長さ l", fontproperties=jp, ha="center",
            fontsize=9)
    ax.annotate("", xy=(-0.4, 0), xytext=(-0.4, h),
                arrowprops=dict(arrowstyle="<|-|>", color="k", lw=1))
    ax.text(-0.7, h / 2, "内法高さ h", fontproperties=jp, rotation=90,
            va="center", ha="center", fontsize=9)
    ax.annotate("", xy=(ox, oy - 0.25), xytext=(ox + l0, oy - 0.25),
                arrowprops=dict(arrowstyle="<|-|>", color="#c00000", lw=1))
    ax.text(l / 2, oy - 0.55, "l0", fontproperties=jp, ha="center",
            fontsize=9, color="#c00000")
    ax.annotate("", xy=(ox - 0.25, oy), xytext=(ox - 0.25, oy + h0),
                arrowprops=dict(arrowstyle="<|-|>", color="#c00000", lw=1))
    ax.text(ox - 0.55, h / 2, "h0", fontproperties=jp, va="center",
            ha="center", fontsize=9, color="#c00000")
    ax.text(l / 2, h + 0.55, "開口周比 γ0 = √( (h0×l0) / (h×l) )",
            fontproperties=jp, ha="center", fontsize=11, color="#1f4e79",
            fontweight="bold")
    ax.text(l / 2, -1.5, "γ0 が 0.4 以下 → 開口付き耐震壁として扱える\n"
                          "（せん断剛性・耐力は r = 1 − 1.25γ0 で低減）\n"
                          "γ0 が 0.4 超 → 耐震壁とみなせない（フレームで評価）",
            fontproperties=jp, ha="center", fontsize=9, color="#c00000")
    ax.set_xlim(-1.4, l + 1.0)
    ax.set_ylim(-2.6, h + 1.2)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(b) 開口周比 γ0 の定義", fontproperties=jp,
                 fontsize=11, fontweight="bold")
    fig.suptitle("図 4-4  耐震壁のモデル化 ── エレメント置換と開口の影響",
                 fontproperties=jp, fontsize=13, fontweight="bold")
    return save(fig, "fig4-4_wall_model.png")


# ===========================================================================
# 図 4-5: マンションの方向別架構と立面の注意点
# ===========================================================================
def fig_4_5():
    fig, axes = plt.subplots(1, 2, figsize=(14, 5.8),
                              gridspec_kw={"width_ratios": [1.5, 1.0]})
    # --- (a) 基準階平面 ---
    ax = axes[0]
    spx = [5.5] * 5
    spy = [2.0, 8.0, 2.0]   # バルコニー / 住戸 / 廊下
    xl = np.cumsum([0] + spx)
    yl = np.cumsum([0] + spy)
    Wm, Hm = sum(spx), sum(spy)
    ax.add_patch(mpatches.Rectangle((0, 0), Wm, Hm, fc="none", ec="k", lw=1.2))
    ax.fill_between([0, Wm], yl[2], yl[3], color="#f0f0f0", alpha=0.6)
    ax.text(Wm / 2 - 4, yl[2] + 1.0, "共用廊下", fontproperties=jp,
            ha="center", fontsize=8, color="gray")
    ax.fill_between([0, Wm], yl[0], yl[1], color="#fff5e0", alpha=0.6)
    ax.text(Wm / 2, yl[0] + 1.0, "バルコニー（開口大）", fontproperties=jp,
            ha="center", fontsize=8, color="gray")
    # 戸境壁（Y方向連層壁）
    for x in xl:
        ax.add_patch(mpatches.Rectangle((x - 0.22, yl[1]), 0.44,
                                         yl[2] - yl[1],
                                         fc="#c4858f", ec="k", lw=0.7,
                                         hatch="//", alpha=0.9, zorder=3))
    # 妻壁（X方向、両端）
    for y_ in [yl[1], yl[2]]:
        pass
    ax.add_patch(mpatches.Rectangle((0, yl[2] - 0.22), 5.5, 0.44,
                                     fc="#c4858f", ec="k", lw=0.7,
                                     hatch="//", alpha=0.9, zorder=3))
    ax.add_patch(mpatches.Rectangle((Wm - 5.5, yl[2] - 0.22), 5.5, 0.44,
                                     fc="#c4858f", ec="k", lw=0.7,
                                     hatch="//", alpha=0.9, zorder=3))
    # EVコア（右端片寄せ）
    ax.add_patch(mpatches.Rectangle((Wm - 3.0, yl[2]), 3.0, 2.0,
                                     fc="#9ec6e8", ec="k", lw=1.0,
                                     hatch="xx", alpha=0.9, zorder=3))
    ax.text(Wm - 1.5, yl[2] + 1.0, "EV・階段", fontproperties=jp,
            ha="center", va="center", fontsize=8, color="#1f4e79")
    # 柱
    cw = 0.5
    for x in xl:
        for y in yl:
            ax.add_patch(mpatches.Rectangle((x - cw / 2, y - cw / 2), cw, cw,
                                             fc="#bdbdbd", ec="k", lw=0.5,
                                             zorder=5))
    # 方向注記
    ax.annotate("Y方向（梁間）＝連層耐震壁架構\n→ 壁エレメントでモデル化",
                xy=(xl[2], (yl[1] + yl[2]) / 2),
                xytext=(xl[2] - 4.5, Hm + 1.3),
                fontproperties=jp, fontsize=9, color="#7a3b3b",
                arrowprops=dict(arrowstyle="->", color="#7a3b3b"))
    ax.annotate("X方向（桁行）＝開口が多く純ラーメンに近い\n→ 線材（柱・梁）でモデル化",
                xy=(Wm * 0.35, yl[1]), xytext=(2.0, -2.8),
                fontproperties=jp, fontsize=9, color="#1f4e79",
                arrowprops=dict(arrowstyle="->", color="#1f4e79"))
    ax.annotate("コア片寄せ\n→ 偏心に注意", xy=(Wm - 1.5, yl[3]),
                xytext=(Wm + 0.8, Hm + 0.8), fontproperties=jp, fontsize=9,
                color="#c00000",
                arrowprops=dict(arrowstyle="->", color="#c00000"))
    ax.set_xlim(-1.5, Wm + 4.5)
    ax.set_ylim(-3.6, Hm + 3.2)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(a) マンション基準階平面 ── 方向で架構形式が異なる",
                 fontproperties=jp, fontsize=11, fontweight="bold")

    # --- (b) 立面の注意点 ---
    ax = axes[1]
    ns, h, w = 8, 1.0, 6.0
    # 2F以上は壁、1Fはピロティ
    for s in range(ns):
        y = s * h
        ax.plot([0, 0], [y, y + h], color="#555", lw=2)
        ax.plot([w, w], [y, y + h], color="#555", lw=2)
        ax.plot([w / 3, w / 3], [y, y + h], color="#555", lw=1.2)
        ax.plot([2 * w / 3, 2 * w / 3], [y, y + h], color="#555", lw=1.2)
        ax.plot([0, w], [y + h, y + h], color="#555", lw=2)
        if s >= 1:
            ax.add_patch(mpatches.Rectangle((w / 3, y), w / 3, h,
                                             fc="#c4858f", ec="k", lw=0.5,
                                             hatch="//", alpha=0.85))
    ax.text(w / 2, 0.5, "1階：駐車場・店舗\n（壁なし＝ピロティ）",
            fontproperties=jp, ha="center", va="center", fontsize=8.5,
            color="#c00000")
    ax.annotate("剛性率 Rs 低下に注意\n（Rs が 0.6 未満で割増し）",
                xy=(w, 0.5), xytext=(w + 0.5, 2.6), fontproperties=jp,
                fontsize=9, color="#c00000",
                arrowprops=dict(arrowstyle="->", color="#c00000"))
    ax.annotate("連層壁は最下階まで\n通すのが原則",
                xy=(w / 2, 4.5), xytext=(w + 0.5, 5.6), fontproperties=jp,
                fontsize=9, color="#7a3b3b",
                arrowprops=dict(arrowstyle="->", color="#7a3b3b"))
    ax.text(w / 2, -1.0, "細長い建物は塔状比 H/B が 4 を超えると\n転倒・杭引抜きの検討が必要",
            fontproperties=jp, ha="center", fontsize=8.5, color="#444")
    ax.set_xlim(-1.0, w + 5.5)
    ax.set_ylim(-1.8, ns * h + 0.8)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(b) 立面の注意 ── ピロティ・壁抜け",
                 fontproperties=jp, fontsize=11, fontweight="bold")
    fig.suptitle("図 4-5  マンションの耐震要素バランス（平面・立面）",
                 fontproperties=jp, fontsize=13, fontweight="bold")
    return save(fig, "fig4-5_balance.png")


figs = {
    "f1": fig_4_1(),
    "f2": fig_4_2(),
    "f3": fig_4_3(),
    "f4": fig_4_4(),
    "f5": fig_4_5(),
}
print("figures:", list(figs.keys()))

# ===========================================================================
# Excel 構築
# ===========================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

wb = Workbook()
C_TITLE = "1F4E79"; C_HEAD = "2E75B6"; C_ANS = "E2EFDA"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
f_title = Font(name="MS PGothic", size=15, bold=True, color="FFFFFF")
f_head = Font(name="MS PGothic", size=11, bold=True, color="FFFFFF")
f_body = Font(name="MS PGothic", size=10)
f_ans = Font(name="MS PGothic", size=10, color="375623")
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)


def setup(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False


def title_row(ws, row, text, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = f_title; c.fill = PatternFill("solid", fgColor=C_TITLE)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 30


def head(ws, row, text, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = f_head; c.fill = PatternFill("solid", fgColor=C_HEAD)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 22


def body(ws, row, text, span=8, ans=False, h=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = f_ans if ans else f_body
    c.alignment = wrap
    if ans:
        c.fill = PatternFill("solid", fgColor=C_ANS)
    if h:
        ws.row_dimensions[row].height = h


def table(ws, start_row, headers, rows, col1=1):
    r = start_row
    for j, htxt in enumerate(headers):
        c = ws.cell(r, col1 + j, htxt)
        c.font = Font(name="MS PGothic", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=C_HEAD)
        c.alignment = center; c.border = border
    for data in rows:
        r += 1
        for j, v in enumerate(data):
            c = ws.cell(r, col1 + j, v)
            c.font = f_body
            c.alignment = center if j > 0 else wrap
            c.border = border
            if r % 2 == 0:
                c.fill = PatternFill("solid", fgColor="F2F7FC")
    return r


def put_img(ws, path, anchor, w=None):
    img = XLImage(path)
    if w:
        ratio = w / img.width
        img.width = w; img.height = int(img.height * ratio)
    ws.add_image(img, anchor)


# ---- 目次 ----
ws = wb.active
ws.title = "目次"
setup(ws, [4, 24, 58, 14])
title_row(ws, 1, "架構の耐震設計とモデル化 問題集（RC マンション設計担当・新人向け）",
          span=4)
body(ws, 2, "テーマ No.4：RC 造マンションの構造設計を担当する想定で、ラーメン架構・"
            "耐震壁架構の耐震設計の考え方と、一貫計算に入力する解析モデル"
            "（線材置換・剛域・壁エレメント・開口低減）を学ぶ。各シートに図、最後に解答。",
     span=4, h=58)
r = table(ws, 4,
          ["No.", "シート", "到達目標", "図"],
          [["4-1", "No.4-1 ラーメン架構", "ラーメン架構の耐震設計を理解している",
            "変形・崩壊形"],
           ["4-2", "No.4-2 柱梁モデル化", "線材置換・剛域等のモデル化を理解している",
            "剛域・協力幅"],
           ["4-3", "No.4-3 耐震壁架構", "耐震壁架構の耐震設計を理解している",
            "変形モード・分担"],
           ["4-4", "No.4-4 壁モデル化", "エレメント置換・開口影響を理解している",
            "壁柱・開口周比"],
           ["4-5", "No.4-5 バランス", "平面・立面バランスを理解している",
            "方向別架構・立面"]])
body(ws, r + 2,
     "凡例：ヒンジ＝○（白抜き=梁端・塗り=柱）、耐震壁＝ピンク斜線、剛域・剛梁＝黒太線、"
     "壁柱＝赤太線。偏心率・剛性率の数値計算は No.10 教材（バランス配置問題集）も併用のこと。",
     span=4, h=44)

# ---- 4-1 ----
ws = wb.create_sheet("No.4-1 ラーメン架構")
setup(ws, [8, 16, 16, 16, 14, 12, 12, 12])
title_row(ws, 1, "No.4-1  ラーメン架構の耐震設計")
head(ws, 3, "■ 図 4-1  水平力を受けるラーメンの変形と崩壊メカニズム")
put_img(ws, figs["f1"], "A4", w=820)
head(ws, 30, "■ 問題 1  地震力の流れ（穴埋め）")
body(ws, 31, "地震力は、まず【 ① 】に慣性力として生じ、床の面内剛性を介して【 ② 】に伝わり、"
             "節点を通じて【 ③ 】のせん断力・曲げモーメントとなり、最後に【 ④ 】から地盤へ伝わる。",
     h=32)
body(ws, 32, "ラーメン架構は柱と梁を【 ⑤ 】接合（剛接合）した架構で、部材の【 ⑥ 】剛性で"
             "水平力に抵抗する。", h=20)
head(ws, 34, "■ 問題 2  柱の曲げモーメントと反曲点")
body(ws, 35, "(1) 水平力時、ラーメンの柱の曲げモーメント分布はどんな形になるか"
             "（直線／曲線、どこで M=0 か）。", h=20)
body(ws, 36, "(2) 中間階の柱の反曲点はおよそ柱のどの高さにあるか。", h=20)
body(ws, 37, "(3) 1 階柱（柱脚固定）の反曲点は中間階より上下どちらに移動するか。理由とともに。",
     h=32)
head(ws, 39, "■ 問題 3  柱せん断力の配分（D 値法の考え方）")
body(ws, 40, "1 階の層せん断力 Q1 = 600 kN を、柱の D 値（水平剛性の相対値）に比例して"
             "配分する。下表を埋めよ。", h=32)
r = table(ws, 42,
          ["柱", "D 値", "負担率 Di/ΣD (記入)", "負担せん断力 (kN) (記入)"],
          [["柱 A（外柱）", "1.0", "", ""],
           ["柱 B（内柱）", "2.0", "", ""],
           ["柱 C（外柱）", "1.0", "", ""],
           ["合計", "4.0", "1.00", "600"]])
body(ws, r + 2, "内柱の D 値が外柱より大きい理由を、取り付く梁の本数（拘束）の観点で 1 行で。",
     h=20)
head(ws, r + 4, "■ 問題 4  強柱弱梁（崩壊形の設計）")
body(ws, r + 5, "(1) 図 4-1 の (b) 全体崩壊形と (c) 層崩壊形では、どちらを目指して設計するか。"
                "エネルギー吸収の観点で理由を述べよ。", h=32)
body(ws, r + 6, "(2) 「強柱弱梁」を実現するために、柱と梁の何を比較してどう設定するか"
                "（柱梁耐力比）。", h=32)
head(ws, r + 8, "■ 問題 5  マンション実務")
body(ws, r + 9, "マンションの桁行方向（X 方向）が純ラーメンになりやすい理由を住戸計画から説明し、"
                "純ラーメン方向の設計で注意すべき点（変形・梁せいと階高・乾式間仕切への影響）を"
                "2 つ挙げよ。", h=44)

# ---- 4-2 ----
ws = wb.create_sheet("No.4-2 柱梁モデル化")
setup(ws, [8, 16, 18, 16, 14, 12, 12, 12])
title_row(ws, 1, "No.4-2  柱・梁のモデル化（線材置換・剛域）")
head(ws, 3, "■ 図 4-2  線材置換・剛域・スラブ協力幅")
put_img(ws, figs["f2"], "A4", w=820)
head(ws, 30, "■ 問題 1  線材置換（穴埋め）")
body(ws, 31, "一貫構造計算では、柱・梁を断面の【 ① 】を通る線材に置き換える（線材置換）。"
             "線材は断面積 A・断面二次モーメント【 ② 】などの断面性能を持ち、"
             "柱梁の交点は【 ③ 】としてモデル化される。柱梁が重なり合う接合部パネルは"
             "変形が小さいため【 ④ 】として扱う。", h=44)
head(ws, 33, "■ 問題 2  剛域長さの計算")
body(ws, 34, "柱 800×800、大梁 500×700（B×D）、スパン L=7.0m（軸線間）、階高 h=3.0m"
             "（梁は上下とも せい 700）。剛域端は「フェイス位置から部材せいの 1/4 内側」とする。",
     h=32)
r = table(ws, 36,
          ["項目", "式", "値 (mm) (記入)"],
          [["梁の剛域長（片側）", "柱せい/2 − 梁せい/4", ""],
           ["梁の可とう長さ", "L − 2×剛域長", ""],
           ["柱の剛域長（片側）", "梁せい/2 − 柱せい/4", ""],
           ["柱の可とう長さ", "h − 2×剛域長", ""]])
head(ws, r + 2, "■ 問題 3  剛域の意味")
body(ws, r + 3, "(1) 剛域を設けず部材全長を可とう（曲げ変形する）と仮定すると、"
                "建物の水平剛性は実際より大きく出るか小さく出るか。その結果、"
                "変形・周期の算定はどちら側にずれるか。", h=40)
body(ws, r + 4, "(2) 剛域を過大に取る（フェイスまで全部剛とする）と何が起こるか。", h=20)
head(ws, r + 6, "■ 問題 4  スラブ協力幅と剛性増大率")
body(ws, r + 7, "(1) 床スラブが付く大梁の曲げ剛性は、長方形梁単体より大きくなる。"
                "その理由を T 形断面の観点で 1 行で述べよ。", h=20)
body(ws, r + 8, "(2) 両側スラブ付き大梁の剛性増大率 φ の一般的な範囲はいくつか。"
                "剛性増大率を見込まないと応力配分・変形算定はどうずれるか。", h=32)
head(ws, r + 10, "■ 問題 5  断面性能の計算")
body(ws, r + 11, "大梁 500×700 の断面二次モーメント I = bD^3/12 を求めよ（mm4、有効数字 3 桁）。"
                 "剛性増大率 φ=2.0 を見込んだ場合の評価剛性 EI はいくらか"
                 "（Ec = 2.3×10^4 N/mm2 として）。", h=40)

# ---- 4-3 ----
ws = wb.create_sheet("No.4-3 耐震壁架構")
setup(ws, [8, 16, 18, 16, 14, 12, 12, 12])
title_row(ws, 1, "No.4-3  耐震壁架構の耐震設計")
head(ws, 3, "■ 図 4-3  耐震壁とフレームの変形モード・分担")
put_img(ws, figs["f3"], "A4", w=820)
head(ws, 30, "■ 問題 1  変形モードの判別")
body(ws, 31, "次の記述は「曲げ型（連層耐震壁）」「せん断型（ラーメン）」のどちらの説明か。",
     h=20)
r = table(ws, 33,
          ["No.", "記述", "曲げ型/せん断型 (記入)"],
          [["(1)", "下層ほど層間変形角が大きい", ""],
           ["(2)", "上層ほど層間変形角が大きい", ""],
           ["(3)", "片持ち梁を立てたような変形", ""],
           ["(4)", "各層が平行四辺形にずれる変形", ""]])
head(ws, r + 2, "■ 問題 2  壁とフレームの分担計算")
body(ws, r + 3, "1 階の層せん断力 Q = 1,000 kN。連層耐震壁の水平剛性 Kw = 8.0（相対値）、"
                "ラーメン構面が 2 枚で各 Kf = 1.0。剛性比例で配分した場合の負担を求めよ。",
     h=32)
r = table(ws, r + 5,
          ["要素", "剛性 K", "負担率 (記入)", "負担せん断力 (kN) (記入)"],
          [["連層耐震壁", "8.0", "", ""],
           ["ラーメン構面×2", "2.0", "", ""],
           ["合計", "10.0", "1.00", "1,000"]])
head(ws, r + 2, "■ 問題 3  連層耐震壁の設計方針（穴埋め）")
body(ws, r + 3, "連層耐震壁は大地震時に【 ① 】部の【 ② 】降伏が先行するように設計し、"
                "脆性的な【 ③ 】破壊を避ける。また壁の転倒モーメントにより基礎には"
                "【 ④ 】力が生じるため、杭の引抜き抵抗や直接基礎の浮き上がりを検討する。",
     h=40)
head(ws, r + 5, "■ 問題 4  相互作用の理解（○×）")
r = table(ws, r + 6,
          ["No.", "記述", "○× (記入)"],
          [["(1)", "下層では壁が層せん断力の大半を負担する", ""],
           ["(2)", "上層ではフレームの負担率が下層より増える", ""],
           ["(3)", "壁とフレームの分担は全階で一定である", ""],
           ["(4)", "頂部付近では壁のせん断力が負（逆向き）になることがある", ""]])
head(ws, r + 2, "■ 問題 5  マンション実務")
body(ws, r + 3, "妻側の連層耐震壁は大きな転倒モーメントを受ける。基礎設計で何を検討するか"
                "2 つ挙げよ（杭の引抜き／直接基礎の接地圧・浮き上がり）。"
                "また、ワンルーム系の細長いマンションで注意する指標（塔状比）にも触れよ。",
     h=44)

# ---- 4-4 ----
ws = wb.create_sheet("No.4-4 壁モデル化")
setup(ws, [8, 16, 18, 16, 14, 12, 12, 12])
title_row(ws, 1, "No.4-4  耐震壁のモデル化（エレメント置換・開口影響）")
head(ws, 3, "■ 図 4-4  エレメント置換と開口周比")
put_img(ws, figs["f4"], "A4", w=820)
head(ws, 30, "■ 問題 1  エレメント置換（穴埋め）")
body(ws, 31, "壁エレメント置換では、耐震壁を壁全体の断面性能を持つ 1 本の【 ① 】に置き換え、"
             "各階の梁レベルに【 ② 】を仮定して両側柱とつなぐ。"
             "せん断変形と【 ③ 】変形の両方を考慮できるようにする。"
             "その他のモデル化手法として、壁を斜め材に置換する【 ④ 】置換や、"
             "壁を細かい要素に分割する【 ⑤ 】解析がある。", h=44)
head(ws, 33, "■ 問題 2  開口周比の計算")
body(ws, 34, "壁の内法長さ l = 6.0 m、内法高さ h = 3.0 m。次の 2 ケースについて"
             "開口周比 γ0 = √((h0×l0)/(h×l)) を計算し、耐震壁として扱えるか判定せよ"
             "（判定基準：γ0 が 0.4 以下）。", h=32)
r = table(ws, 36,
          ["ケース", "開口寸法 l0×h0", "開口面積/壁面積 (記入)", "γ0 (記入)", "判定 (記入)"],
          [["1：掃き出し窓", "1.5 m × 1.8 m", "", "", ""],
           ["2：大開口", "2.0 m × 2.0 m", "", "", ""]])
head(ws, r + 2, "■ 問題 3  開口低減率")
body(ws, r + 3, "ケース 1 について、せん断剛性・耐力の低減率 r = 1 − 1.25γ0 を計算せよ。"
                "この壁の剛性・耐力は無開口壁の約何 % として扱うことになるか。", h=32)
head(ws, r + 5, "■ 問題 4  開口補強")
body(ws, r + 6, "開口隅部には斜めひび割れが生じやすい。開口周囲に配置する補強筋を 2 種類挙げ、"
                "それぞれの目的を 1 行で述べよ（開口隅部の斜め筋／開口周囲の縦横補強筋）。",
     h=40)
head(ws, r + 8, "■ 問題 5  雑壁の扱い（実務判断）")
body(ws, r + 9, "スリットを切らない雑壁（腰壁・袖壁等）を解析モデルで剛性に見込むか否かで、"
                "周期・応力分配が変わる。「剛性に見込む／見込まない」それぞれの場合に"
                "何が危険側になり得るか、1 行ずつ述べよ。", h=44)

# ---- 4-5 ----
ws = wb.create_sheet("No.4-5 バランス")
setup(ws, [8, 16, 18, 16, 14, 12, 12, 12])
title_row(ws, 1, "No.4-5  耐震要素の平面的・立面的バランス")
head(ws, 3, "■ 図 4-5  マンションの方向別架構と立面の注意点")
put_img(ws, figs["f5"], "A4", w=820)
head(ws, 30, "■ 問題 1  方向別の架構形式とモデル化")
body(ws, 31, "図 4-5(a) のマンションについて、下表を埋めよ。", h=20)
r = table(ws, 33,
          ["方向", "主な耐震要素 (記入)", "架構形式 (記入)", "解析モデル (記入)"],
          [["Y 方向（梁間）", "", "", ""],
           ["X 方向（桁行）", "", "", ""]])
body(ws, r + 2, "方向によって剛性・固有周期が大きく異なる。どちらの方向が周期が長く"
                "（柔らかく）なりやすいか。", h=20)
head(ws, r + 4, "■ 問題 2  平面バランス（偏心）")
body(ws, r + 5, "(1) 図のように EV・階段コアが平面の右端に片寄ると、何が剛心を移動させ、"
                "どんな振動性状が生じるか。", h=32)
body(ws, r + 6, "(2) 改善策を 2 つ挙げよ（コアの分散配置／反対側への壁追加 等）。", h=20)
head(ws, r + 8, "■ 問題 3  立面バランス（ピロティ・壁抜け）")
body(ws, r + 9, "(1) 1 階を駐車場・店舗として壁を抜くと、どの指標が悪化し、"
                "法規上どんなペナルティ（割増し）があるか。", h=32)
body(ws, r + 10, "(2) 連層耐震壁を途中階で止める（壁抜け）と何が起こるか。"
                 "「原則最下階まで通す」理由を応力の流れから説明せよ。", h=40)
head(ws, r + 12, "■ 問題 4  塔状比")
body(ws, r + 13, "塔状比 H/B = 4 を超える細長い建物で必要となる検討を 2 つ挙げよ"
                 "（転倒／杭引抜き・基礎浮き上がり）。"
                 "幅 B=8m のワンルームマンションで H/B=4 となる高さは何 m か。", h=40)
head(ws, r + 15, "■ 問題 5  総合（自由記述）")
body(ws, r + 16, "図 4-5 のマンション計画について、構造計画上の懸念点を 3 つ指摘し、"
                 "それぞれ改善案を 1 行で提案せよ（コア偏心／1 階ピロティ／"
                 "X 方向の壁量不足 などの観点）。", h=44)

# ---- 解答 ----
ws = wb.create_sheet("解答")
setup(ws, [4, 22, 60, 14])
title_row(ws, 1, "解答・解説", span=4)
row = 2


def ah(t):
    global row
    head(ws, row, t, span=4); row += 1


def an(t, h=None):
    global row
    body(ws, row, t, span=4, ans=True, h=h); row += 1


ah("No.4-1  ラーメン架構")
an("問1：①床（スラブ） ②大梁 ③柱 ④基礎 ⑤剛 ⑥曲げ。"
   "床の慣性力→床の面内剛性→大梁→節点→柱→基礎→地盤、の流れを体で覚える。", h=44)
an("問2：(1)直線分布で、部材の中間に M=0 の点（反曲点）ができる。"
   "(2)中間階はほぼ柱中央（h/2 付近）。"
   "(3)1 階柱は柱脚固定で脚部の曲げが大きくなるため、反曲点は中央より上"
   "（h の 0.6〜0.65 倍程度の高さ）に移動する。", h=58)
an("問3：負担率 A:1/4=0.25、B:2/4=0.50、C:0.25。"
   "負担せん断力 A=150 kN、B=300 kN、C=150 kN。"
   "内柱は両側に梁が取り付き回転拘束が強いため D 値（水平剛性）が大きい。", h=44)
an("問4：(1)(b) 全体崩壊形を目指す。ヒンジが多数の梁端に分散し、建物全体で"
   "エネルギーを吸収できるため変形能力が大きい。層崩壊は 1 つの層に変形が集中し"
   "脆性的に倒壊する。(2)節点で「柱の曲げ耐力の和 ＞ 梁の曲げ耐力の和」"
   "（柱梁耐力比を 1 以上、余裕をみて 1.5 程度）とし、梁が先に降伏するようにする。",
   h=72)
an("問5：桁行方向は住戸の開口（バルコニー・廊下・窓）が連続し壁を置けないため"
   "純ラーメンになりやすい。注意点：①変形（層間変形角）が壁方向より大きく、"
   "サッシ・乾式間仕切の変形追従を確認 ②梁せいの確保と階高・天井高の調整"
   "（梁せい大→階高増→コスト増のトレードオフ）。", h=72)

ah("No.4-2  柱梁モデル化")
an("問1：①図心（断面の中心軸） ②断面二次モーメント I ③節点 ④剛域。", h=22)
an("問2：梁の剛域長 = 800/2 − 700/4 = 400 − 175 = 225 mm。"
   "梁の可とう長 = 7,000 − 2×225 = 6,550 mm。"
   "柱の剛域長 = 700/2 − 800/4 = 350 − 200 = 150 mm。"
   "柱の可とう長 = 3,000 − 2×150 = 2,700 mm。"
   "※式が負になる場合は剛域なしとする。", h=72)
an("問3：(1)剛域を無視すると部材が実際より長く曲げ変形するため、剛性を過小評価"
   "→ 変形・周期を過大に算定し、地震力の配分も不正確になる。"
   "(2)逆に剛域を過大（フェイスまで剛）に取ると剛性過大→変形過小・応力過大評価で、"
   "周期も短く算定され危険側・不経済の両方が起こり得る。慣用の D/4 控除はその中間。",
   h=72)
an("問4：(1)スラブが圧縮フランジとして働き、T 形断面として中立軸から遠い位置に"
   "コンクリートが増えるため剛性が増す。"
   "(2)両側スラブ付きで φ≒1.5〜2.0。見込まないと梁剛性過小→柱への応力配分・"
   "節点の固定度・変形算定がすべてずれる（一貫計算では自動考慮が一般的）。", h=58)
an("問5：I = 500×700^3/12 = 1.43×10^10 mm4。"
   "EI = 2.3×10^4 × 1.43×10^10 = 3.29×10^14 N·mm2。"
   "φ=2.0 で評価剛性 EI = 6.58×10^14 N·mm2（I 換算 2.86×10^10 mm4）。", h=44)

ah("No.4-3  耐震壁架構")
an("問1：(1)せん断型 (2)曲げ型 (3)曲げ型 (4)せん断型。", h=22)
an("問2：壁 8/10=0.80 → 800 kN、ラーメン 2/10=0.20 → 200 kN。"
   "下層では壁が約 8 割を負担するのが耐震壁付きラーメンの典型。", h=32)
an("問3：①脚（基礎に近い最下部） ②曲げ ③せん断 ④引抜き（浮き上がり）。"
   "曲げ降伏先行＝粘り強い破壊形式に誘導するのが大原則。", h=32)
an("問4：(1)○ (2)○ (3)×（高さ方向で変化する） (4)○"
   "（曲げ型の壁とせん断型のフレームの変形整合により、頂部付近で壁がフレームに"
   "押し戻され、壁のせん断力が逆転することがある）。", h=44)
an("問5：①杭基礎：押込みだけでなく引抜き耐力の検討（引抜き側の杭・杭頭接合）。"
   "②直接基礎：接地圧の偏り・浮き上がり（接地率）の確認。"
   "塔状比 H/B が 4 を超えると転倒・引抜きの検討が必須となり、"
   "細長いマンションでは妻壁直下の基礎が設計を支配することが多い。", h=58)

ah("No.4-4  壁モデル化")
an("問1：①壁柱（壁エレメント） ②剛梁 ③曲げ ④ブレース ⑤FEM（有限要素）。", h=22)
an("問2：ケース1：開口面積/壁面積 = (1.5×1.8)/(6.0×3.0) = 2.7/18 = 0.15、"
   "γ0 = √0.15 = 0.387 → 0.4 以下 → 開口付き耐震壁として扱える。"
   "ケース2：(2.0×2.0)/18 = 0.222、γ0 = √0.222 = 0.471 → 0.4 超 → 耐震壁と"
   "みなせない（壁を無視するか、袖壁付きフレーム等として別途評価）。", h=72)
an("問3：r = 1 − 1.25×0.387 = 0.516。無開口壁の約 52 % のせん断剛性・耐力として扱う。",
   h=22)
an("問4：①開口隅部の斜め補強筋：隅部に集中する斜め引張（ひび割れ）に直交して抵抗。"
   "②開口周囲の縦・横補強筋：開口で切られた壁筋の応力を肩代わりし周辺に伝達。", h=44)
an("問5：剛性に見込まない場合：実際は雑壁が効いて剛心がずれ、想定外のねじれ・"
   "短柱破壊が起こり得る（剛性評価の危険側）。"
   "見込む場合：雑壁の耐力は不確実なので、耐力まで期待すると保有耐力を過大評価する"
   "恐れ（耐力評価の危険側）。実務は「剛性は見込み、耐力は期待しない」等、"
   "方針を決めて一貫させ、スリットの有無と整合させる。", h=72)

ah("No.4-5  バランス")
an("問1：Y 方向＝戸境壁（連層耐震壁）／耐震壁架構（壁式に近い）／壁エレメント置換。"
   "X 方向＝柱・大梁（開口多く壁が置けない）／純ラーメンに近い架構／線材モデル。"
   "周期は壁の少ない X 方向（桁行）が長く柔らかい。", h=58)
an("問2：(1)コアの壁が右端に集中し剛心が右へ移動、重心との偏心によりねじれ振動が"
   "生じ、コアから遠い左端の柱・壁の変形・損傷が大きくなる。"
   "(2)①コア（階段等）を平面の両端や中央に分散配置 ②反対側（左端）にも"
   "耐震壁を追加し剛心を重心に近づける。", h=58)
an("問3：(1)剛性率 Rs が低下（Rs が 0.6 未満で Fs 割増し＝必要保有水平耐力の増大）。"
   "偏心があれば Fe 割増しも重畳。"
   "(2)壁抜けすると、上階の壁が負担していたせん断力・転倒モーメントが"
   "止めた階の柱・梁に集中し、応力の流れが急変して損傷が集中する。"
   "連層壁は基礎まで通して力をまっすぐ地盤へ流すのが原則。", h=72)
an("問4：①転倒の検討（保有水平耐力時の転倒モーメント） ②杭の引抜き・基礎の"
   "浮き上がりの検討。B=8m なら H = 4×8 = 32 m（約 10 階建て）で塔状比 4 に達する。",
   h=44)
an("問5（例）：①コア片寄せ→偏心：左端妻側に壁を追加、またはコアを分散。"
   "②1 階ピロティ→剛性率低下：1 階に壁を復活、または鉄骨ブレース・柱断面増で"
   "剛性確保し Rs を 0.6 以上に。"
   "③X 方向壁量不足→純ラーメンの変形大：EV コアの X 方向壁を活用、"
   "梁せい確保、層間変形角 1/200 以内を確認。", h=72)

XLSX = os.path.join(OUT, "架構モデル化問題集.xlsx")
wb.save(XLSX)
print("saved:", XLSX)
