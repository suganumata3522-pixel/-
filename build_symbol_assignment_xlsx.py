# -*- coding: utf-8 -*-
"""符号割 問題集（図つき）Excel 生成スクリプト。
出力: docs/symbol_assignment/符号割問題集.xlsx
図:   matplotlib(IPAGothic) で 平面図・軸組図・接合部詳細図 を生成し各シートに埋込。
"""
import os
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib import font_manager as fm

FONT_PATH = "/usr/share/fonts/truetype/fonts-japanese-gothic.ttf"
jp = fm.FontProperties(fname=FONT_PATH)
fm.fontManager.addfont(FONT_PATH)
plt.rcParams["font.family"] = jp.get_name()
plt.rcParams["axes.unicode_minus"] = False

OUT = "docs/symbol_assignment"
FIG = os.path.join(OUT, "figures")
os.makedirs(FIG, exist_ok=True)

# ========== 共通平面：3スパン×3スパン (X=6m × 3, Y=7m × 3) ==========
SPX = [6, 6, 6]
SPY = [7, 7, 7]
XL = np.cumsum([0] + SPX)
YL = np.cumsum([0] + SPY)
W = sum(SPX); H = sum(SPY)
NX, NY = len(XL), len(YL)


def col_letter(ix, iy):
    return chr(ord("a") + iy * NX + ix)


def col_category(ix, iy):
    """(category_label, fill_color)"""
    if ix in [0, NX - 1] and iy in [0, NY - 1]:
        return "隅", "#f4a261"
    if iy in [0, NY - 1]:
        return "妻X", "#ffd180"
    if ix in [0, NX - 1]:
        return "妻Y", "#ffd180"
    return "中", "#a8dadc"


def save(fig, name):
    p = os.path.join(FIG, name)
    fig.savefig(p, dpi=130, bbox_inches="tight")
    plt.close(fig)
    return p


# ========== 平面図 描画関数 ==========
def draw_plan(ax, title, *, walls=False, brace_axes=False, letters=False, col_syms=None):
    # グリッド点線
    for x in XL:
        ax.plot([x, x], [-0.3, H + 0.3], color="lightgray", lw=0.6, ls=":", zorder=1)
    for y in YL:
        ax.plot([-0.3, W + 0.3], [y, y], color="lightgray", lw=0.6, ls=":", zorder=1)

    # 軸線丸（X1〜X4, Y1〜Y4）
    for i, x in enumerate(XL):
        for yp in [-1.6, H + 1.0]:
            ax.text(x, yp, f"X{i + 1}", fontproperties=jp, ha="center", va="center",
                    bbox=dict(boxstyle="circle,pad=0.18", fc="white", ec="k", lw=0.8),
                    fontsize=8, zorder=10)
    for i, y in enumerate(YL):
        for xp in [-1.6, W + 1.0]:
            ax.text(xp, y, f"Y{i + 1}", fontproperties=jp, ha="center", va="center",
                    bbox=dict(boxstyle="circle,pad=0.18", fc="white", ec="k", lw=0.8),
                    fontsize=8, zorder=10)

    # スパン寸法 (mm)
    for i, s in enumerate(SPX):
        ax.text((XL[i] + XL[i + 1]) / 2, -2.7, f"{s*1000:,}",
                fontproperties=jp, ha="center", fontsize=7, color="gray")
    for i, s in enumerate(SPY):
        ax.text(-2.8, (YL[i] + YL[i + 1]) / 2, f"{s*1000:,}",
                fontproperties=jp, ha="center", va="center", rotation=90,
                fontsize=7, color="gray")

    # 耐震壁
    if walls:
        ww = 0.45
        for y in [YL[0], YL[-1]]:
            ax.fill_betweenx([y - ww / 2, y + ww / 2], XL[1], XL[2],
                             color="#c4858f", ec="k", lw=1.0, hatch="//", alpha=0.9, zorder=3)
        ax.text((XL[1] + XL[2]) / 2, YL[0] - 0.9, "W1", fontproperties=jp,
                ha="center", fontsize=10, color="#7a3b3b", fontweight="bold")
        ax.text((XL[1] + XL[2]) / 2, YL[-1] + 0.85, "W1", fontproperties=jp,
                ha="center", fontsize=10, color="#7a3b3b", fontweight="bold")

    # ブレース構面のマーカ（柱列を縦長の点線枠で）
    if brace_axes:
        for x in [XL[0], XL[-1]]:
            ax.add_patch(mpatches.Rectangle((x - 0.6, YL[0] - 0.5), 1.2, H + 1.0,
                                            fc="none", ec="#c00000", lw=1.6,
                                            ls=(0, (4, 2)), zorder=2))
        ax.text(W / 2, H + 1.9, "X1・X4 通り = ブレース構面", fontproperties=jp,
                ha="center", fontsize=9, color="#c00000")

    # 梁
    for iy, y in enumerate(YL):
        for ix in range(NX - 1):
            ax.plot([XL[ix] + 0.5, XL[ix + 1] - 0.5], [y, y],
                    color="#1f4e79", lw=2.2, zorder=4)
    for ix, x in enumerate(XL):
        for iy in range(NY - 1):
            ax.plot([x, x], [YL[iy] + 0.5, YL[iy + 1] - 0.5],
                    color="#1f4e79", lw=2.2, zorder=4)

    # 柱
    cw = 0.85
    for ix, x in enumerate(XL):
        for iy, y in enumerate(YL):
            cat, color = col_category(ix, iy)
            if brace_axes and ix in [0, NX - 1]:
                color = "#ffb0b0"
            rect = mpatches.Rectangle((x - cw / 2, y - cw / 2), cw, cw,
                                      fc=color, ec="k", lw=1.2, zorder=5)
            ax.add_patch(rect)
            if letters:
                ax.text(x, y, col_letter(ix, iy), fontproperties=jp,
                        ha="center", va="center", fontsize=10,
                        fontweight="bold", zorder=6)
            elif col_syms and (ix, iy) in col_syms:
                ax.text(x, y, col_syms[(ix, iy)], fontproperties=jp,
                        ha="center", va="center", fontsize=8,
                        fontweight="bold", zorder=6)

    # 凡例 (right side)
    lx, ly = W + 2.5, H - 0.5
    items = [("#f4a261", "隅柱"), ("#ffd180", "妻柱"), ("#a8dadc", "中柱")]
    if walls:
        items.append(("#c4858f", "耐震壁"))
    if brace_axes:
        items.append(("#ffb0b0", "ブレース付き柱"))
    for i, (c, lab) in enumerate(items):
        ax.add_patch(mpatches.Rectangle((lx, ly - i * 0.9), 0.55, 0.55, fc=c, ec="k"))
        ax.text(lx + 0.75, ly - i * 0.9 + 0.28, lab, fontproperties=jp,
                fontsize=8, va="center")

    ax.set_xlim(-3.5, W + 5.5)
    ax.set_ylim(-3.5, H + 3.0)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title(title, fontproperties=jp, fontsize=12, fontweight="bold")


# ========== 軸組図（S造、3階建） ==========
def draw_elev(ax, title, *, braces=False, story_h=4.0, stories=3):
    HT = story_h * stories
    # グリッド
    for x in XL:
        ax.plot([x, x], [0, HT], color="lightgray", lw=0.5, ls=":")
    for s in range(stories + 1):
        ax.plot([0, W], [s * story_h, s * story_h], color="lightgray", lw=0.5, ls=":")
    # 階名
    for s in range(stories + 1):
        name = f"{s + 1}FL" if s < stories else "RFL"
        ax.text(-1.8, s * story_h, name, fontproperties=jp, fontsize=8,
                va="center", ha="center")
    # X 軸丸
    for i, x in enumerate(XL):
        ax.text(x, HT + 0.7, f"X{i + 1}", fontproperties=jp, ha="center", va="center",
                bbox=dict(boxstyle="circle,pad=0.18", fc="white", ec="k", lw=0.8),
                fontsize=8)
    # スパン寸法
    for i, s in enumerate(SPX):
        ax.text((XL[i] + XL[i + 1]) / 2, -0.8, f"{s * 1000:,}",
                fontproperties=jp, ha="center", fontsize=7, color="gray")
    # 階高
    for s in range(stories):
        ax.text(W + 0.5, s * story_h + story_h / 2, f"{int(story_h*1000)}",
                fontproperties=jp, fontsize=7, color="gray", va="center", rotation=90)

    # 柱
    cw = 0.4
    for x in XL:
        ax.add_patch(mpatches.Rectangle((x - cw / 2, 0), cw, HT,
                                        fc="#bdbdbd", ec="k", lw=1.0, zorder=3))
    # 梁
    bh = 0.28
    for s in range(1, stories + 1):
        for ix in range(NX - 1):
            ax.add_patch(mpatches.Rectangle((XL[ix] + cw / 2, s * story_h - bh / 2),
                                            XL[ix + 1] - XL[ix] - cw, bh,
                                            fc="#666", ec="k", lw=0.6, zorder=3))
    # ブレース（X1-X2 と X3-X4 スパンに各階 X 形）
    if braces:
        for s in range(stories):
            for ix in [0, 2]:
                xl = XL[ix] + cw / 2
                xr = XL[ix + 1] - cw / 2
                yb = s * story_h + bh / 2
                yt = (s + 1) * story_h - bh / 2
                ax.plot([xl, xr], [yb, yt], color="#c00000", lw=2.2, zorder=5)
                ax.plot([xl, xr], [yt, yb], color="#c00000", lw=2.2, zorder=5)

    ax.set_xlim(-3, W + 2)
    ax.set_ylim(-1.5, HT + 1.8)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title(title, fontproperties=jp, fontsize=12, fontweight="bold")


# ========== 図 1: 基本 RC ラーメン平面（letters a〜p） ==========
def fig_rc_plan():
    fig, ax = plt.subplots(figsize=(8.5, 7.5))
    draw_plan(ax, "図 9-1  RC ラーメン架構  基準階平面図", letters=True)
    return save(fig, "fig9-1_rc_plan.png")


# ========== 図 2: RC 耐震壁併用 平面 ==========
def fig_rc_wall_plan():
    fig, ax = plt.subplots(figsize=(8.5, 7.5))
    draw_plan(ax, "図 9-2  RC 耐震壁併用ラーメン  基準階平面図",
              walls=True, letters=True)
    return save(fig, "fig9-2_rc_wall_plan.png")


# ========== 図 3: S 造ラーメン 平面＋軸組図 ==========
def fig_s_ramen():
    fig, axes = plt.subplots(1, 2, figsize=(13, 6),
                              gridspec_kw={"width_ratios": [1.0, 1.1]})
    draw_plan(axes[0], "(a) S 造ラーメン  基準階平面", letters=True)
    draw_elev(axes[1], "(b) S 造ラーメン  X 方向軸組図（3階建）")
    return save(fig, "fig9-3_s_ramen.png")


# ========== 図 4: S 造ブレース併用 平面＋軸組図 ==========
def fig_s_brace():
    fig, axes = plt.subplots(1, 2, figsize=(13, 6),
                              gridspec_kw={"width_ratios": [1.0, 1.1]})
    draw_plan(axes[0], "(a) S 造ブレース併用  基準階平面",
              brace_axes=True, letters=True)
    draw_elev(axes[1], "(b) X 方向軸組図  X1・X4 通りにブレース",
              braces=True)
    return save(fig, "fig9-4_s_brace.png")


# ========== 図 5: RC 柱梁接合部（主筋干渉） ==========
def fig_rc_joint():
    fig, (a1, a2) = plt.subplots(1, 2, figsize=(11.5, 5.2))
    # ----- 平面 (joint 上から見た断面) -----
    sz = 8.0  # 柱 800×800 を 8 単位として図化
    a1.add_patch(mpatches.Rectangle((-sz / 2, -sz / 2), sz, sz,
                                     fc="#eeeeee", ec="k", lw=1.5))
    # 柱主筋（外周 12 本）
    cov = 0.7
    n = 4
    pos = np.linspace(-sz / 2 + cov, sz / 2 - cov, n)
    for px in pos:
        for py in pos:
            if abs(px) > pos[-1] - 0.01 or abs(py) > pos[-1] - 0.01 or \
               (px == pos[0] or px == pos[-1]) or (py == pos[0] or py == pos[-1]):
                a1.add_patch(plt.Circle((px, py), 0.28,
                                         fc="black", ec="black", zorder=5))
    # 梁 X 方向（左右に貫通）：上端筋 4 本（赤）
    bx_y_top = sz / 2 - 1.4
    bx_y_bot = -sz / 2 + 1.4
    bxs = np.linspace(-sz / 2 + cov + 0.2, sz / 2 - cov - 0.2, 4)
    for bx in bxs:
        a1.plot([-sz / 2 - 1, sz / 2 + 1], [bx_y_top, bx_y_top],
                color="#c00000", lw=0.6, zorder=4)
        a1.plot([-sz / 2 - 1, sz / 2 + 1], [bx_y_bot, bx_y_bot],
                color="#c00000", lw=0.6, zorder=4)
        a1.add_patch(plt.Circle((bx, bx_y_top), 0.30, fc="#c00000",
                                 ec="#c00000", zorder=6))
        a1.add_patch(plt.Circle((bx, bx_y_bot), 0.30, fc="#c00000",
                                 ec="#c00000", zorder=6))
    # 梁 Y 方向（上下に貫通）：上端筋 4 本（青）。X 方向梁筋と同じ高さに通ろうとして干渉
    by_x_top = sz / 2 - 1.4
    by_x_bot = -sz / 2 + 1.4
    bys = np.linspace(-sz / 2 + cov + 0.2, sz / 2 - cov - 0.2, 4)
    for by in bys:
        a1.plot([by_x_top, by_x_top], [-sz / 2 - 1, sz / 2 + 1],
                color="#1f4e79", lw=0.6, zorder=4)
        a1.plot([by_x_bot, by_x_bot], [-sz / 2 - 1, sz / 2 + 1],
                color="#1f4e79", lw=0.6, zorder=4)
        a1.add_patch(plt.Circle((by_x_top, by), 0.30, fc="#1f4e79",
                                 ec="#1f4e79", zorder=6))
        a1.add_patch(plt.Circle((by_x_bot, by), 0.30, fc="#1f4e79",
                                 ec="#1f4e79", zorder=6))
    # 干渉ハイライト
    for cx in [bx_y_top, bx_y_bot]:
        for cy_pos in [by_x_top, by_x_bot]:
            a1.add_patch(plt.Circle((cy_pos, cx), 0.85, fc="none",
                                     ec="orange", lw=2, ls="--", zorder=8))
    a1.text(0, -sz / 2 - 1.8, "黄丸 = X・Y 両方向の梁主筋が交差する干渉点",
            fontproperties=jp, ha="center", color="#d68000", fontsize=9)
    # 凡例
    a1.text(-sz / 2 - 1.2, sz / 2 + 1.3, "●柱主筋(黒)　●X方向梁主筋(赤)　●Y方向梁主筋(青)",
            fontproperties=jp, fontsize=9)
    a1.set_xlim(-sz / 2 - 2.5, sz / 2 + 2.5)
    a1.set_ylim(-sz / 2 - 2.8, sz / 2 + 2.2)
    a1.set_aspect("equal")
    a1.axis("off")
    a1.set_title("(a) RC 柱梁接合部 平面  ── 直交2方向梁主筋の干渉",
                 fontproperties=jp, fontsize=10)
    # ----- 立面 (柱を切った縦断面 X 方向梁が貫通) -----
    a2.add_patch(mpatches.Rectangle((-sz / 2, -sz / 2), sz, sz,
                                     fc="#eeeeee", ec="k", lw=1.5))
    # 梁高 600 mm 相当を 6 単位として、柱の中央高さに大梁が取り付く
    bh = 6
    bw = 14
    a2.add_patch(mpatches.Rectangle((-bw / 2, -bh / 2), bw, bh,
                                     fc="none", ec="#1f4e79", lw=1.2, ls="--"))
    # 梁上端筋・下端筋（折曲げ定着）
    a2.plot([-bw / 2, sz / 2 - cov], [bh / 2 - 0.5, bh / 2 - 0.5],
            color="#c00000", lw=2)
    a2.plot([sz / 2 - cov, sz / 2 - cov], [bh / 2 - 0.5, -sz / 2 + cov],
            color="#c00000", lw=2)
    a2.plot([-bw / 2, sz / 2 - cov], [-bh / 2 + 0.5, -bh / 2 + 0.5],
            color="#c00000", lw=2)
    a2.plot([sz / 2 - cov, sz / 2 - cov], [-bh / 2 + 0.5, sz / 2 - cov],
            color="#c00000", lw=2)
    a2.annotate("梁上端筋\n折曲げ定着", xy=(sz / 2 - cov, 0.5),
                xytext=(sz / 2 + 1.8, 1.5), fontproperties=jp, fontsize=8,
                arrowprops=dict(arrowstyle="->", color="#c00000"))
    a2.annotate("接合部内\n帯筋（補強筋）", xy=(0, 0), xytext=(-sz / 2 - 3, 0),
                fontproperties=jp, fontsize=8,
                arrowprops=dict(arrowstyle="->", color="#777"))
    a2.set_xlim(-sz - 1, sz / 2 + 4.5)
    a2.set_ylim(-sz / 2 - 1, sz / 2 + 1.5)
    a2.set_aspect("equal")
    a2.axis("off")
    a2.set_title("(b) RC 柱梁接合部 立面  ── 主筋の定着・接合部帯筋",
                 fontproperties=jp, fontsize=10)
    return save(fig, "fig9-5_rc_joint.png")


# ========== 図 6: S造 通しダイアフラム接合部 ==========
def fig_s_joint():
    fig, (a1, a2) = plt.subplots(1, 2, figsize=(11.5, 5.0))
    # ----- 立面：柱貫通＋通しダイアフラム -----
    cw, ch = 1.2, 10
    a1.add_patch(mpatches.Rectangle((-cw / 2, 0), cw, ch,
                                     fc="#bdbdbd", ec="k", lw=1.2))
    # 通しダイアフラム（梁フランジレベル）
    bh = 2.6
    dy1, dy2 = ch / 2 - bh / 2, ch / 2 + bh / 2
    dt = 0.32  # ダイアフラム厚
    for dy in [dy1 - dt / 2, dy2 - dt / 2]:
        a1.add_patch(mpatches.Rectangle((-cw / 2 - 0.4, dy), cw + 0.8, dt,
                                         fc="#3a3a3a", ec="k", lw=0.8))
    # 梁ウェブ
    a1.add_patch(mpatches.Rectangle((cw / 2, dy1 + dt / 2 + 0.1),
                                     5.5, bh - dt - 0.2,
                                     fc="#cccccc", ec="k", lw=0.8))
    # 梁フランジ
    a1.add_patch(mpatches.Rectangle((cw / 2, dy1 - 0.1), 5.5, 0.5,
                                     fc="#888", ec="k", lw=0.6))
    a1.add_patch(mpatches.Rectangle((cw / 2, dy2 - 0.4), 5.5, 0.5,
                                     fc="#888", ec="k", lw=0.6))
    a1.annotate("通しダイアフラム\n(t=22 など)", xy=(cw / 2 + 0.4, dy2 + dt / 2),
                xytext=(cw / 2 + 2.0, ch + 0.5), fontproperties=jp, fontsize=8,
                arrowprops=dict(arrowstyle="->", color="#333"))
    a1.annotate("梁フランジ\n(t=19 など)", xy=(cw / 2 + 1.5, dy1 + 0.15),
                xytext=(cw / 2 + 3.2, dy1 - 1.5), fontproperties=jp, fontsize=8,
                arrowprops=dict(arrowstyle="->", color="#333"))
    a1.annotate("柱 □-500×500×22", xy=(-cw / 2, ch / 2),
                xytext=(-cw / 2 - 3.5, ch / 2 - 0.5), fontproperties=jp, fontsize=8,
                arrowprops=dict(arrowstyle="->", color="#333"))
    a1.text(0, -0.8, "通しダイアフラム厚 >= 梁フランジ厚 が原則",
            fontproperties=jp, ha="center", color="#c00000", fontsize=9)
    a1.set_xlim(-5.5, cw / 2 + 6.5)
    a1.set_ylim(-1.5, ch + 2.0)
    a1.set_aspect("equal")
    a1.axis("off")
    a1.set_title("(a) S 造 通しダイアフラム形式 接合部 立面",
                 fontproperties=jp, fontsize=10)
    # ----- 立面 (継手位置の例) -----
    a2.add_patch(mpatches.Rectangle((-cw / 2, 0), cw, ch * 1.3,
                                     fc="#bdbdbd", ec="k", lw=1.2))
    # 梁
    a2.add_patch(mpatches.Rectangle((cw / 2, 4), 9, 0.8,
                                     fc="#cccccc", ec="k"))
    a2.add_patch(mpatches.Rectangle((cw / 2, 8), 9, 0.8,
                                     fc="#cccccc", ec="k"))
    # 現場継手（添え板）
    a2.add_patch(mpatches.Rectangle((cw / 2 + 2.2, 3.7), 0.5, 1.4,
                                     fc="#c00000", ec="#c00000"))
    a2.add_patch(mpatches.Rectangle((cw / 2 + 2.2, 7.7), 0.5, 1.4,
                                     fc="#c00000", ec="#c00000"))
    a2.annotate("梁の現場継手\n(梁端から L/4 程度)", xy=(cw / 2 + 2.45, 4.5),
                xytext=(cw / 2 + 4, 2.5), fontproperties=jp, fontsize=8,
                arrowprops=dict(arrowstyle="->", color="#c00000"))
    # 柱の継手位置
    a2.add_patch(mpatches.Rectangle((-cw / 2 - 0.15, 5.4), cw + 0.3, 0.5,
                                     fc="#c00000", ec="#c00000"))
    a2.annotate("柱の現場継手\n(2階床上 1m 程度)", xy=(-cw / 2, 5.65),
                xytext=(-cw / 2 - 4.5, 7.0), fontproperties=jp, fontsize=8,
                arrowprops=dict(arrowstyle="->", color="#c00000"))
    a2.text(0, -0.8, "継手位置の違い = 同断面でも別符号で扱うことが多い",
            fontproperties=jp, ha="center", color="#c00000", fontsize=9)
    a2.set_xlim(-6, cw / 2 + 11)
    a2.set_ylim(-1.5, ch * 1.3 + 1.5)
    a2.set_aspect("equal")
    a2.axis("off")
    a2.set_title("(b) S 造 柱・梁の現場継手 位置の例",
                 fontproperties=jp, fontsize=10)
    return save(fig, "fig9-5_s_joint.png")


# ========== 図 7: 解答用 RC ラーメン (符号入り) ==========
def fig_rc_answer():
    syms = {}
    for ix in range(NX):
        for iy in range(NY):
            cat, _ = col_category(ix, iy)
            syms[(ix, iy)] = {"隅": "C1", "妻X": "C2", "妻Y": "C3", "中": "C4"}[cat]
    fig, ax = plt.subplots(figsize=(8.5, 7.5))
    draw_plan(ax, "解答 9-1  RC ラーメン 柱符号", col_syms=syms)
    return save(fig, "fig9-1_answer.png")


def fig_wall_answer():
    syms = {}
    for ix in range(NX):
        for iy in range(NY):
            cat, _ = col_category(ix, iy)
            base = {"隅": "C1", "妻X": "C2", "妻Y": "C3", "中": "C4"}[cat]
            # 壁付き柱 (X2Y1, X3Y1, X2Y4, X3Y4)
            if (ix in [1, 2]) and (iy in [0, NY - 1]):
                base = "WC1"
            syms[(ix, iy)] = base
    fig, ax = plt.subplots(figsize=(8.5, 7.5))
    draw_plan(ax, "解答 9-2  RC 耐震壁併用 柱符号 (WC=壁付き柱)",
              walls=True, col_syms=syms)
    return save(fig, "fig9-2_answer.png")


figs = {
    "f1": fig_rc_plan(),
    "f2": fig_rc_wall_plan(),
    "f3": fig_s_ramen(),
    "f4": fig_s_brace(),
    "f5": fig_rc_joint(),
    "f6": fig_s_joint(),
    "f1a": fig_rc_answer(),
    "f2a": fig_wall_answer(),
}
print("figures:", list(figs.keys()))

# ===========================================================================
# Excel ビルド
# ===========================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

wb = Workbook()
C_TITLE = "1F4E79"; C_HEAD = "2E75B6"; C_ANS = "E2EFDA"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

f_title = Font(name="MS PGothic", size=15, bold=True, color="FFFFFF")
f_head = Font(name="MS PGothic", size=11, bold=True, color="FFFFFF")
f_body = Font(name="MS PGothic", size=10)
f_ans = Font(name="MS PGothic", size=10, color="375623")
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)


def setup(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False


def title_row(ws, row, text, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = f_title
    c.fill = PatternFill("solid", fgColor=C_TITLE)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 30


def head(ws, row, text, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = f_head
    c.fill = PatternFill("solid", fgColor=C_HEAD)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 22


def body(ws, row, text, span=8, ans=False, h=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = f_ans if ans else f_body
    c.alignment = wrap
    if ans:
        c.fill = PatternFill("solid", fgColor=C_ANS)
    if h:
        ws.row_dimensions[row].height = h


def table(ws, start_row, headers, rows, col1=1):
    r = start_row
    for j, htxt in enumerate(headers):
        c = ws.cell(r, col1 + j, htxt)
        c.font = Font(name="MS PGothic", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=C_HEAD)
        c.alignment = center
        c.border = border
    for data in rows:
        r += 1
        for j, v in enumerate(data):
            c = ws.cell(r, col1 + j, v)
            c.font = f_body
            c.alignment = center if j > 0 else wrap
            c.border = border
            if r % 2 == 0:
                c.fill = PatternFill("solid", fgColor="F2F7FC")
    return r


def put_img(ws, path, anchor, w=None):
    img = XLImage(path)
    if w:
        ratio = w / img.width
        img.width = w
        img.height = int(img.height * ratio)
    ws.add_image(img, anchor)


# ---- 列分類 (a〜p) ----
def all_columns_table():
    rows = []
    for iy in range(NY):
        for ix in range(NX):
            ltr = col_letter(ix, iy)
            cat, _ = col_category(ix, iy)
            rows.append([ltr.upper(), f"X{ix+1}Y{iy+1}", cat, ""])
    return rows


# ===========================================================================
# シート構築
# ===========================================================================
ws = wb.active
ws.title = "目次"
setup(ws, [4, 18, 60, 14])
title_row(ws, 1, "符号割 問題集（ゼネコン構造設計部・新人向け）", span=4)
body(ws, 2, "テーマ No.9：RC／S 造の柱・梁符号割。共通の 3×3 スパン平面（X=6m × Y=7m）"
            "を題材に、ラーメン／耐震壁／ブレース／配筋・接合部納まりへ展開する。"
            "各シートに 平面図・軸組図 等を埋め込み、空欄テーブルに符号 (C1, G1, W1, V1 …) を記入する形式。",
     span=4, h=58)
r = table(ws, 4,
          ["No.", "シート", "到達目標", "図"],
          [["9-1", "No.9-1 RC ラーメン", "RC ラーメン架構の柱・梁符号割ができる", "平面図"],
           ["9-2", "No.9-2 RC 耐震壁併用", "耐震壁併用ラーメン架構の柱・梁符号割ができる", "壁配置平面"],
           ["9-3", "No.9-3 S 造ラーメン", "S 造ラーメン架構の柱・梁符号割ができる", "平面＋軸組"],
           ["9-4", "No.9-4 S 造ブレース併用", "ブレース併用ラーメン架構の柱・梁符号割ができる", "平面＋軸組"],
           ["9-5", "No.9-5 配筋・接合部納まり", "配筋・接合部を踏まえた符号配置ができる", "接合部詳細"]])
body(ws, r + 2,
     "凡例：柱記号 = C（耐震壁付 WC、ブレース付 BC）／大梁 = G（外周 G_o、内部 G_i、壁付 WG、"
     "ブレース付 BG）／小梁 = B／耐震壁 = W／ブレース = V。色は 隅柱(橙)・妻柱(黄)・中柱(水色)・"
     "壁付柱/ブレース付柱(赤系)。",
     span=4, h=58)

# ---- No.9-1 ----
ws = wb.create_sheet("No.9-1 RC ラーメン")
setup(ws, [8, 14, 14, 14, 12, 12, 12, 12])
title_row(ws, 1, "No.9-1  RC 造ラーメン架構の柱・梁符号割")
head(ws, 3, "■ 平面図（基準階）  柱は a〜p の 16 本")
put_img(ws, figs["f1"], "A4", w=620)
head(ws, 30, "■ 問題 1　各柱に適切な符号 C1〜C? を書き入れよ")
body(ws, 31, "同じ条件（軸力・曲げ・配筋）の柱は同一符号で集約してよい。"
             "ヒント：隅柱・妻柱（長辺側／短辺側）・中柱の 4 区分で分けると整理しやすい。",
     h=40)
r = table(ws, 33,
          ["柱記号", "位置 (Xi Yj)", "区分", "符号 (記入)"],
          all_columns_table())
head(ws, r + 2, "■ 問題 2　大梁の符号割")
body(ws, r + 3, "下表に大梁の本数と用途を示す。外周梁(G_o)と内部梁(G_i)、X方向と Y方向で集約案を埋めよ。",
     h=32)
r = table(ws, r + 5,
          ["項目", "本数", "想定区分", "符号 (記入)"],
          [["Y1, Y4 通りの X 方向梁 (外周)", "6", "外周大梁", ""],
           ["Y2, Y3 通りの X 方向梁 (内部)", "6", "内部大梁", ""],
           ["X1, X4 通りの Y 方向梁 (外周)", "6", "外周大梁", ""],
           ["X2, X3 通りの Y 方向梁 (内部)", "6", "内部大梁", ""]])
head(ws, r + 2, "■ 問題 3　集約の判断")
body(ws, r + 3,
     "(1) 隅柱と中柱を同符号にしてよいか。理由とともに答えよ。", h=22)
body(ws, r + 4,
     "(2) 1〜3階の柱を全階同符号にしたい。同符号にできる条件を3つ挙げよ。", h=22)
body(ws, r + 5,
     "(3) 「符号を分けすぎる」「符号を集約しすぎる」それぞれの実務上のデメリットを各1行で。",
     h=32)

# ---- No.9-2 ----
ws = wb.create_sheet("No.9-2 RC 耐震壁併用")
setup(ws, [8, 14, 14, 14, 12, 12, 12, 12])
title_row(ws, 1, "No.9-2  RC 耐震壁併用ラーメン架構の符号割")
head(ws, 3, "■ 平面図（基準階）  Y1・Y4 通りの中央スパンに耐震壁 W1 を配置")
put_img(ws, figs["f2"], "A4", w=620)
head(ws, 30, "■ 問題 1　壁配置の意図")
body(ws, 31,
     "(1) なぜ耐震壁を妻側（Y1・Y4）に配置するのが望ましいか。剛心と重心の関係から述べよ。",
     h=32)
body(ws, 32,
     "(2) この壁配置では、X 方向地震 と Y 方向地震 のうち、どちらが壁で抵抗されるか。",
     h=22)
head(ws, 34, "■ 問題 2　柱符号（壁付き柱 WC を分離）")
body(ws, 35,
     "壁の両端柱（X2Y1, X3Y1, X2Y4, X3Y4）は壁の境界柱として軸力・曲げが大きい。"
     "通常柱 C と分けて WC1 とする。下表を埋めよ。",
     h=42)
r = table(ws, 38,
          ["柱記号", "位置", "区分", "符号 (記入)"],
          all_columns_table())
head(ws, r + 2, "■ 問題 3　壁付き梁 WG と通常梁 G")
body(ws, r + 3,
     "Y1 通り中央スパン（X2-X3 間）の梁は耐震壁の上に乗る『壁付き梁 (WG)』。"
     "通常の大梁 G と何が異なるか、2 点挙げよ（剛性／応力／配筋の観点）。",
     h=44)
head(ws, r + 5, "■ 問題 4　連層壁の符号")
body(ws, r + 6,
     "1〜10F の耐震壁を全階 W1 で通せる条件を3つ挙げよ"
     "（壁厚／縦筋径・ピッチ／横筋径・ピッチ／開口の有無 から選んで）。",
     h=44)

# ---- No.9-3 ----
ws = wb.create_sheet("No.9-3 S 造ラーメン")
setup(ws, [8, 14, 14, 14, 12, 12, 12, 12])
title_row(ws, 1, "No.9-3  S 造ラーメン架構の柱・梁符号割")
head(ws, 3, "■ 平面図(a) と X 方向軸組図(b)")
put_img(ws, figs["f3"], "A4", w=720)
head(ws, 32, "■ 問題 1　柱種別の選定")
body(ws, 33,
     "(1) 4 方向すべてラーメン抵抗する S 造の柱には BOX 柱（□形鋼管）と H 形鋼柱のどちらが適するか。"
     "理由とともに答えよ。", h=44)
body(ws, 34,
     "(2) 中柱と隅柱で同じ BOX サイズを採用する場合、符号を分けるべきか／同じでよいか。", h=22)
head(ws, 36, "■ 問題 2　柱符号の集約")
body(ws, 37,
     "全柱を BOX □-500×500×22 で統一する案と、隅柱だけ□-500×500×19 に薄くする案がある。"
     "それぞれのメリット・デメリットを述べよ（工場製作・建方・コスト）。", h=58)
head(ws, 40, "■ 問題 3　梁符号の集約")
body(ws, 41,
     "下表は各通りの梁の応力比（最大／長期＝1.0 が応力許容ぎりぎり）。"
     "応力比から H 形鋼サイズを 3 段階（A=H-600×200、B=H-500×200、C=H-400×200）で割り付け、符号を入れよ。",
     h=50)
r = table(ws, 44,
          ["通り", "区分", "応力比", "サイズ", "符号(記入)"],
          [["Y1 (外周)", "X 方向大梁", "0.95", "", ""],
           ["Y2 (内部)", "X 方向大梁", "0.62", "", ""],
           ["Y3 (内部)", "X 方向大梁", "0.55", "", ""],
           ["Y4 (外周)", "X 方向大梁", "0.92", "", ""],
           ["X1 (外周)", "Y 方向大梁", "0.85", "", ""],
           ["X2 (内部)", "Y 方向大梁", "0.45", "", ""],
           ["X3 (内部)", "Y 方向大梁", "0.42", "", ""],
           ["X4 (外周)", "Y 方向大梁", "0.83", "", ""]])
head(ws, r + 2, "■ 問題 4　継手位置")
body(ws, r + 3,
     "12m を超える梁は運搬上の制約で工場製作を分割する必要がある。"
     "Y1 通り（外周梁、長さ 18m = 3 スパン×6m）の梁を 2 分割するとき、現場継手の位置は"
     "どこに設定すべきか（応力分布の観点で）。", h=58)

# ---- No.9-4 ----
ws = wb.create_sheet("No.9-4 S 造ブレース併用")
setup(ws, [8, 14, 14, 14, 12, 12, 12, 12])
title_row(ws, 1, "No.9-4  S 造ブレース併用ラーメン架構の符号割")
head(ws, 3, "■ 平面図(a)：X1・X4 通り＝ブレース構面　／  軸組図(b)：X1-X2・X3-X4 スパンに X 形ブレース")
put_img(ws, figs["f4"], "A4", w=720)
head(ws, 32, "■ 問題 1　ブレース構面の特定")
body(ws, 33,
     "(1) この計画では X 方向と Y 方向のどちらに対してブレースで抵抗するか。",
     h=22)
body(ws, 34,
     "(2) Y 方向の水平力は何で抵抗するか。", h=22)
head(ws, 36, "■ 問題 2　柱符号（ブレース付き柱 BC を分離）")
body(ws, 37,
     "X1・X4 通りの柱は、地震時に大きな軸力（ブレースの押し引き）を受ける。"
     "通常柱 SC と分け、BC1 とする。各柱の符号を記入せよ。", h=44)
r = table(ws, 40,
          ["柱記号", "位置", "ブレース付き?", "符号 (記入)"],
          [[col_letter(ix, iy).upper(), f"X{ix+1}Y{iy+1}",
            "○" if ix in [0, NX-1] else "—", ""]
           for iy in range(NY) for ix in range(NX)])
head(ws, r + 2, "■ 問題 3　ブレースの符号")
body(ws, r + 3,
     "全 6 構面×3 階 = 18 本のブレースに符号を割り付ける。"
     "(1) 全 1 種類 (V1) で集約する条件は？"
     "(2) 1階のブレースだけ部材を太くする場合、符号は何種類になるか？", h=58)
head(ws, r + 5, "■ 問題 4　ブレース付き梁 (BG)")
body(ws, r + 6,
     "ブレース構面の梁は、ブレースの軸力の水平成分が梁の軸方向力として作用する"
     "(通常の鉛直荷重梁と異なる)。通常梁 G と別符号にする理由を述べよ。", h=44)

# ---- No.9-5 ----
ws = wb.create_sheet("No.9-5 配筋・接合部納まり")
setup(ws, [8, 14, 14, 14, 12, 12, 12, 12])
title_row(ws, 1, "No.9-5  配筋・接合部納まりを考慮した符号配置")
head(ws, 3, "■ 図 9-5-1　RC 柱梁接合部：直交2方向の梁主筋の干渉")
put_img(ws, figs["f5"], "A4", w=720)
head(ws, 28, "■ 問題 1　RC 接合部の主筋干渉")
body(ws, 29,
     "(1) X 方向梁の上端筋（赤）と Y 方向梁の上端筋（青）が同じ高さで交差する。"
     "実際の納まりでは何を上にして配筋するか（端梁ルール／梁せい・主筋径による判断）。", h=58)
body(ws, 30,
     "(2) 干渉を回避する方法を3つ挙げよ（主筋本数・段数・梁せい・接合部詳細から）。",
     h=44)
body(ws, 31,
     "(3) 同じ柱断面でも、両方向の梁主筋本数が多い接合部は『接合部だけ別符号』にして"
     "詳細図を別に描くことがある。その目的を述べよ。", h=58)
head(ws, 33, "■ 図 9-5-2　S 造 通しダイアフラム形式の接合部 と 現場継手の位置")
put_img(ws, figs["f6"], "A4", w=720)
head(ws, 56, "■ 問題 2　S 造ダイアフラム厚と符号")
body(ws, 57,
     "梁フランジ厚を 19mm から 22mm に上げると、通しダイアフラム厚もそれに合わせて変える"
     "必要がある。これにより同断面の柱でも『接合部仕様が違う』ため別符号となる場合がある。"
     "なぜダイアフラム厚 ≥ 梁フランジ厚 でなければならないか、力の伝達経路から説明せよ。",
     h=72)
head(ws, 60, "■ 問題 3　継手位置と符号")
body(ws, 61,
     "S 造柱の現場継手位置は 2 階床上 1m 程度が一般的。"
     "1 階の柱（基礎〜継手位置）と 2 階の柱（継手位置〜3 階床下）を同断面で計画した場合、"
     "符号を 1 つにすべきか、別にすべきか。建方手順と部品管理の観点で答えよ。", h=72)
head(ws, 64, "■ 問題 4　集約と分離の総合判断")
body(ws, 65,
     "ある建物の柱 5 本を比較する。下表の条件で、どこまで同符号にできるか判定せよ。",
     h=32)
r = table(ws, 67,
          ["柱", "断面", "配筋", "かぶり", "継手位置", "判定(同符号OK→○)"],
          [["A", "800×800", "16-D29", "40", "2F上 1m", ""],
           ["B", "800×800", "16-D29", "40", "2F上 1m", ""],
           ["C", "800×800", "20-D29", "40", "2F上 1m", ""],
           ["D", "800×800", "16-D29", "50 (浴室)", "2F上 1m", ""],
           ["E", "800×800", "16-D29", "40", "3F上 1m", ""]])
body(ws, r + 2, "判定の指針：断面・配筋・かぶり・継手位置すべて同じ→同符号OK。"
                "1つでも違えば原則 別符号（または同符号でサブ区分明記）。", h=32)

# ---- 解答 ----
ws = wb.create_sheet("解答")
setup(ws, [4, 22, 60, 14])
title_row(ws, 1, "解答・解説", span=4)
row = 2


def ah(t):
    global row
    head(ws, row, t, span=4); row += 1


def an(t, h=None):
    global row
    body(ws, row, t, span=4, ans=True, h=h); row += 1


def aimg(path, w=520):
    global row
    put_img(ws, path, f"A{row}", w=w); row += 18


ah("No.9-1  RC ラーメン")
an("問1 柱符号：a/d/m/p=隅柱→C1、b/c/n/o=妻柱X辺→C2、e/i/h/l=妻柱Y辺→C3、"
   "f/g/j/k=中柱→C4。長辺6m・短辺7m のため妻X と 妻Y で曲げモーメント分布が異なり、"
   "厳密には別符号が望ましい。応力比に余裕があれば C2=C3 に集約も可。", h=72)
aimg(figs["f1a"], w=520)
an("問2 大梁：Y1/Y4(外周)X方向→G1、Y2/Y3(内部)X方向→G2、"
   "X1/X4(外周)Y方向→G3、X2/X3(内部)Y方向→G4。"
   "床荷重分担が異なる(外周は片側のみ／内部は両側)ため外周と内部は別符号。", h=58)
an("問3 (1)隅柱と中柱は軸力差が大きく(中柱>>隅柱)、必要主筋本数が異なる→別符号。"
   "(2)同符号にできる条件：①断面同じ ②主筋径・本数同じ ③帯筋径・ピッチ同じ。"
   "(3)分けすぎ→図面・部材表が肥大化、施工ミス・積算ミスの元／集約しすぎ→過剰設計でコスト増、"
   "応力不足部材が混入するリスク。", h=86)

ah("No.9-2  RC 耐震壁併用")
an("問1 (1)壁を建物外周(妻側)に配置すると剛心が外側に来て、重心との偏心が小さくなり"
   "ねじれ振動が抑制される。建物中央に壁を集中させると平面剛心と重心が一致しやすいが、"
   "意匠上 妻側に配する方が室内空間を活かせる。(2)Y1・Y4 通りの壁は Y 方向に長く"
   "面内せん断・曲げで抵抗する→Y 方向地震に有効。X 方向地震に対しては壁は弱軸向きで"
   "ラーメンが負担。", h=86)
an("問2 壁付き柱の符号：X2Y1, X3Y1, X2Y4, X3Y4 → WC1。それ以外は問1と同じ C1〜C4。",
   h=44)
aimg(figs["f2a"], w=520)
an("問3 WG は ①壁による拘束で剛性が梁単体の数倍 ②壁の負担せん断が梁端に集中→せん断補強"
   "(あばら筋密)が必要 ③曲げモーメントも壁基部で最大化。通常梁 G とは断面・配筋・符号いずれも別。",
   h=58)
an("問4 連層壁を同符号にできる条件：壁厚同じ／縦筋(主筋)径・ピッチ同じ／"
   "横筋径・ピッチ同じ／開口の有無が同じ。1つでも変われば階毎に W1, W2, W3…と分ける。",
   h=58)

ah("No.9-3  S 造ラーメン")
an("問1 (1)BOX 柱。両軸方向にラーメンを組む場合、強軸・弱軸の区別がない BOX が合理的。"
   "H 形鋼は強軸方向のラーメン・弱軸方向のブレースという使い分けに向く。"
   "(2)隅柱は中柱より軸力が小さく、同サイズだと符号集約は可能だが応力比が大きく異なる場合は"
   "板厚を落として別符号にする方が経済的。", h=86)
an("問2 統一案：工場製作・建方が単純、部材管理ミス減。デメリット：隅柱で過剰設計→鋼材ロス。"
   "別符号案：軽量化で材料費低減。デメリット：部品数増加、現場での取り違えリスク、"
   "塗装・耐火被覆区分が増える。一般に同サイズで集約→鉄骨数量で経済性を確認しつつ判断。",
   h=72)
an("問3 応力比 0.9 以上=A(H-600)、0.6〜0.9=B(H-500)、0.6 未満=C(H-400)。"
   "Y1/Y4/X1/X4(外周)→A(G1)、Y2/Y3/X2/X3(内部)→B/C(G2)。"
   "X 方向と Y 方向で長さが違うため厳密にはサイズが変わるが、応力比からは集約可能。",
   h=58)
an("問4 梁端の曲げモーメントは大きく、現場継手は応力の小さい位置（梁中央 1/4 点付近、"
   "L/4〜3L/8 程度）に配する。Y1 梁 18m なら、両端から 4.5m の位置に継手 2 箇所を設け、"
   "中央 9m 区間と両側 4.5m 区間の 3 ピースに分割。", h=58)

ah("No.9-4  S 造ブレース併用")
an("問1 (1)X 方向（ブレースが X1・X4 通りに沿って X 方向に並ぶ）。"
   "(2)Y 方向は柱・梁のラーメンで抵抗。", h=44)
an("問2 X1・X4 通りの 8 本（a, d, e, h, i, l, m, p）→ BC1。"
   "それ以外（b, c, f, g, j, k, n, o）→ SC1（通常柱）。", h=44)
an("問3 (1)全階・全構面でブレース部材サイズ・接合部仕様が同じなら V1 一種類で集約。"
   "(2)1 階のみ太いブレースなら V1(2-3階)と V2(1階)の 2 種類。"
   "ブレースは部位ごとに長さが違うが、断面が同じなら同符号で OK（長さは部品表で管理）。",
   h=72)
an("問4 BG は曲げ・せん断に加え、ブレースから伝わる『軸方向力』が常時作用する。"
   "通常梁 G は軸力を無視するが、BG は軸力＋曲げの組合せで設計→断面・接合部・符号いずれも別。",
   h=58)

ah("No.9-5  配筋・接合部納まり")
an("問1 (1)梁せい(=主筋から梁下端までの距離)を確保する観点で、せいの大きい梁の主筋を上に通すのが原則。"
   "両方向同じせいなら、社内ルール(短辺方向を上、または長辺方向を上)に従う。"
   "(2)①主筋本数を減らし径を上げる ②2 段配筋にする ③梁せいを上げて主筋間隔を確保する"
   "④主筋を千鳥配置にする ⑤接合部内で機械式定着を使う(折曲げを減らす)。"
   "(3)接合部別符号化の目的：①接合部内の主筋本数・段数・帯筋を別途明示できる"
   "②柱配筋図とは別の接合部詳細図で監督・職人に確実に伝達できる ③現場での配筋順序を明確化。",
   h=130)
an("問2 ダイアフラム厚 < フランジ厚 だと、フランジから入ってきた引張力が"
   "ダイアフラム板で受けきれず、柱面（角溶接部）で応力集中・破断する恐れがある。"
   "ダイアフラム厚 ≥ フランジ厚 で力の伝達面積を確保し、柱面溶接の負担を軽減する。"
   "実務上は『フランジ厚＋3mm』程度を標準とすることが多い。", h=86)
an("問3 1 階柱と 2 階以上の柱は、断面が同じでも『継手の有無／加工内容』が異なる"
   "(1 階柱頂部に継手用エンドプレート、2 階柱底部にも継手部、ボルト孔等)。"
   "工場製作上は別部材なので、別符号(SC1a, SC1b など)とするか、"
   "同符号でサブ区分(上柱/下柱)を部品表に明記。", h=86)
an("問4 A=B（全項目一致）→同符号 ○。C は配筋本数違う→別符号。"
   "D はかぶり違う(浴室は耐久性で増す)→別符号(かぶりが異なる＝有効断面が異なる)。"
   "E は継手位置違う→別符号(製作・建方上の部品が違う)。"
   "結論：A=B=C1、C=C2、D=C3、E=C4 となる。", h=86)

XLSX = os.path.join(OUT, "符号割問題集.xlsx")
wb.save(XLSX)
print("saved:", XLSX)
