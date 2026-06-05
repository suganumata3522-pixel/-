# -*- coding: utf-8 -*-
"""土圧 問題集（図つき）Excel 生成スクリプト。
出力: docs/earth_pressure/土圧問題集.xlsx
図:   matplotlib(IPAGothic) で生成し各シートに埋め込む。
"""
import os
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib import font_manager as fm

FONT_PATH = "/usr/share/fonts/truetype/fonts-japanese-gothic.ttf"
jp = fm.FontProperties(fname=FONT_PATH)
plt.rcParams["font.family"] = jp.get_name()
fm.fontManager.addfont(FONT_PATH)
plt.rcParams["axes.unicode_minus"] = False

OUT_DIR = "docs/earth_pressure"
FIG_DIR = os.path.join(OUT_DIR, "figures")
os.makedirs(FIG_DIR, exist_ok=True)


def savefig(fig, name):
    p = os.path.join(FIG_DIR, name)
    fig.savefig(p, dpi=130, bbox_inches="tight")
    plt.close(fig)
    return p


# ---------------------------------------------------------------------------
# 図1: 壁の変位と土圧（静止・主動・受動）
# ---------------------------------------------------------------------------
def fig_3types():
    fig, ax = plt.subplots(figsize=(6.2, 4.2))
    d = np.linspace(-1.2, 1.2, 400)  # +:主動側(離れる)  -:受動側(押込む)
    K0, Ka, Kp = 0.5, 0.33, 3.0
    K = np.where(
        d >= 0,
        Ka + (K0 - Ka) * np.exp(-d / 0.05),          # 主動側：わずかな変位で急減
        K0 + (Kp - K0) * (1 - np.exp(d / 0.45)),      # 受動側：大変位で緩増
    )
    ax.plot(d, K, lw=2.2, color="#1f4e79")
    ax.axvline(0, color="gray", ls="--", lw=1)
    for y, t, c in [(K0, "静止土圧 K0", "#444"),
                    (Ka, "主動土圧 Ka", "#c00000"),
                    (Kp, "受動土圧 Kp", "#1f7a1f")]:
        ax.axhline(y, color=c, ls=":", lw=1)
        ax.text(1.22, y, f" {t}={y}", va="center", fontproperties=jp, color=c, fontsize=10)
    ax.annotate("壁が土から離れる\n（主動）", xy=(0.55, 0.7), fontproperties=jp,
                fontsize=9, ha="center", color="#c00000")
    ax.annotate("壁が土を押し込む\n（受動）", xy=(-0.65, 2.0), fontproperties=jp,
                fontsize=9, ha="center", color="#1f7a1f")
    ax.set_xlabel("壁の変位 Δ/H  （+主動 ／ −受動）", fontproperties=jp)
    ax.set_ylabel("土圧係数 K", fontproperties=jp)
    ax.set_title("壁の変位と土圧係数の関係（φ=30°の例）", fontproperties=jp)
    ax.set_xlim(-1.2, 1.7)
    ax.set_ylim(0, 3.3)
    ax.grid(alpha=0.25)
    return savefig(fig, "fig1_3types.png")


# ---------------------------------------------------------------------------
# 図2: クーロン式の擁壁断面（角度の定義）
# ---------------------------------------------------------------------------
def fig_coulomb():
    fig, ax = plt.subplots(figsize=(6.2, 4.6))
    # 壁（背面が傾斜 θ）
    H = 4.0
    theta = np.deg2rad(10)
    top_x = H * np.tan(theta)
    ax.plot([0, 0], [0, 0], lw=0)  # placeholder
    # 壁躯体（簡易）
    wall = plt.Polygon([(-0.6, 0), (0, 0), (top_x, H), (top_x - 0.6, H)],
                       closed=True, fc="#bdbdbd", ec="k")
    ax.add_patch(wall)
    # 背面地盤（傾斜 β）
    beta = np.deg2rad(15)
    gx = np.array([top_x, 4.2])
    gy = H + (gx - top_x) * np.tan(beta)
    ax.fill_between(gx, gy, H - 0.0, color="#e8d8b0", zorder=0)
    ax.plot([top_x, 4.2], [H, H + (4.2 - top_x) * np.tan(beta)], color="#8a6d3b", lw=1.5)
    # 背面線（壁背面）
    ax.plot([0, top_x], [0, H], color="k", lw=2)
    # 鉛直基準線
    ax.plot([0, 0], [0, H], color="gray", ls="--", lw=1)
    # 角度注記
    ax.annotate("θ\n(壁傾斜)", xy=(0.15, H - 0.6), fontproperties=jp, fontsize=9)
    ax.annotate("β\n(地表傾斜)", xy=(3.0, H + 0.15), fontproperties=jp, fontsize=9)
    ax.annotate("φ (内部摩擦角)\nc (粘着力)", xy=(2.2, 1.4), fontproperties=jp, fontsize=9,
                color="#8a6d3b")
    # 合力 Pa（壁背面の法線から δ 傾く）
    mid = (top_x / 2, H / 2)
    ax.annotate("", xy=(mid[0] - 1.1, mid[1] - 0.15), xytext=mid,
                arrowprops=dict(arrowstyle="-|>", color="#c00000", lw=2))
    ax.text(mid[0] - 1.25, mid[1] - 0.45, "Pa\n(δ:壁面摩擦角)", fontproperties=jp,
            color="#c00000", fontsize=9)
    ax.set_xlim(-1.0, 4.3)
    ax.set_ylim(-0.3, H + 1.4)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("クーロン式の各パラメータ（θ・β・φ・δ）", fontproperties=jp)
    return savefig(fig, "fig2_coulomb.png")


# ---------------------------------------------------------------------------
# 図3: 上載荷重 q がある場合の土圧分布
# ---------------------------------------------------------------------------
def fig_surcharge():
    fig, ax = plt.subplots(figsize=(6.2, 4.4))
    H = 5.0
    Ka, gamma, q = 0.33, 18, 10
    p_self = Ka * gamma * H   # 下端の自重土圧
    p_sur = Ka * q            # 上載土圧(一様)
    # 壁
    ax.add_patch(plt.Rectangle((-0.5, 0), 0.5, H, fc="#bdbdbd", ec="k"))
    # 地表 + 上載
    ax.plot([0, 3.5], [H, H], color="#8a6d3b", lw=1.5)
    for x in np.linspace(0.2, 3.3, 9):
        ax.annotate("", xy=(x, H), xytext=(x, H + 0.45),
                    arrowprops=dict(arrowstyle="-|>", color="#1f7a1f", lw=1.2))
    ax.text(1.7, H + 0.55, f"上載荷重 q={q} kN/m²", fontproperties=jp,
            color="#1f7a1f", ha="center", fontsize=9)
    # 上載土圧（矩形）
    sx = 0.0
    rect_w = p_sur * 0.06
    ax.fill_betweenx([0, H], sx, sx + rect_w, color="#9ec6e8", alpha=0.8)
    # 自重土圧（三角形、矩形の外側に積む）
    tri_x = [sx + rect_w, sx + rect_w + p_self * 0.06, sx + rect_w]
    tri_y = [0, 0, H]
    ax.fill(tri_x, tri_y, color="#1f4e79", alpha=0.75)
    ax.text(sx + rect_w + p_self * 0.06 + 0.05, 0.3,
            f"自重 KaγH={p_self:.0f}", fontproperties=jp, fontsize=9, color="#1f4e79")
    ax.text(sx + 0.05, H * 0.5, f"上載\nKaq={p_sur:.1f}", fontproperties=jp,
            fontsize=8, color="#1f4e79")
    # 作用点
    ax.annotate("自重→H/3", xy=(0.1, H / 3), fontproperties=jp, fontsize=8, color="#c00000")
    ax.annotate("上載→H/2", xy=(0.1, H / 2), fontproperties=jp, fontsize=8, color="#c00000")
    ax.plot([-0.5, -0.5], [0, H], color="k")
    ax.text(-0.75, H / 2, f"H={H:.0f}m", fontproperties=jp, rotation=90, va="center")
    ax.set_xlim(-1.1, 3.6)
    ax.set_ylim(-0.3, H + 1.2)
    ax.axis("off")
    ax.set_title("上載荷重による土圧（矩形）＋自重土圧（三角形）", fontproperties=jp)
    return savefig(fig, "fig3_surcharge.png")


# ---------------------------------------------------------------------------
# 図4: 土圧の算出（合力と作用点）
# ---------------------------------------------------------------------------
def fig_earth():
    fig, ax = plt.subplots(figsize=(6.0, 4.4))
    H, Ka, gamma = 5.0, 0.33, 18
    pmax = Ka * gamma * H
    ax.add_patch(plt.Rectangle((-0.5, 0), 0.5, H, fc="#bdbdbd", ec="k"))
    ax.plot([0, 3.5], [H, H], color="#8a6d3b", lw=1.5)
    tri_x = [0, pmax * 0.07, 0]
    ax.fill(tri_x, [0, 0, H], color="#1f4e79", alpha=0.7)
    ax.text(pmax * 0.07 + 0.05, 0.2, f"pa=KaγH={pmax:.0f} kN/m²",
            fontproperties=jp, fontsize=9, color="#1f4e79")
    Pa = 0.5 * Ka * gamma * H ** 2
    ax.annotate(f"合力 Pa=½KaγH²={Pa:.0f} kN/m",
                xy=(0.05, H / 3), xytext=(1.3, H / 3),
                fontproperties=jp, fontsize=9, color="#c00000",
                arrowprops=dict(arrowstyle="-|>", color="#c00000", lw=2))
    ax.plot([0, 1.1], [H / 3, H / 3], color="#c00000", ls=":", lw=1)
    ax.text(1.15, H / 3 - 0.35, "作用点 H/3", fontproperties=jp, fontsize=8, color="#c00000")
    ax.plot([-0.5, -0.5], [0, H], color="k")
    ax.text(-0.78, H / 2, f"H={H:.0f}m", fontproperties=jp, rotation=90, va="center")
    ax.set_xlim(-1.1, 3.6)
    ax.set_ylim(-0.3, H + 0.9)
    ax.axis("off")
    ax.set_title("主動土圧の三角形分布・合力・作用点", fontproperties=jp)
    return savefig(fig, "fig4_earth.png")


# ---------------------------------------------------------------------------
# 図5: 地下水位がある場合（土圧＋水圧）
# ---------------------------------------------------------------------------
def fig_water():
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(7.6, 4.4))
    H, hw = 5.0, 2.0  # 水位は GL-2m
    Ka = 0.33
    gt, gsat, gw = 18, 20, 9.8
    gp = gsat - gw
    # --- 左: 有効土圧 ---
    p_wt = Ka * gt * hw                    # 水位での土圧
    sig_b = gt * hw + gp * (H - hw)        # 下端有効鉛直応力
    p_b = Ka * sig_b
    ax1.add_patch(plt.Rectangle((-0.4, 0), 0.4, H, fc="#bdbdbd", ec="k"))
    ax1.plot([0, 2.6], [H, H], color="#8a6d3b", lw=1.5)
    ax1.axhline(H - hw, color="#2b7fc0", ls="--", lw=1.2)
    ax1.text(1.3, H - hw + 0.08, "地下水位 GL-2m", fontproperties=jp,
             color="#2b7fc0", fontsize=8, ha="center")
    xs = [0, p_wt * 0.05, p_b * 0.05, 0]
    ys = [H, H - hw, 0, 0]
    ax1.fill(xs, ys, color="#1f4e79", alpha=0.7)
    ax1.text(p_wt * 0.05 + 0.05, H - hw, f"{p_wt:.0f}", fontproperties=jp, fontsize=8)
    ax1.text(p_b * 0.05 + 0.05, 0.1, f"{p_b:.0f} kN/m²", fontproperties=jp, fontsize=8)
    ax1.set_title("① 有効土圧（水位下はγ'）", fontproperties=jp, fontsize=10)
    ax1.set_xlim(-0.6, 2.0)
    ax1.set_ylim(-0.3, H + 0.8)
    ax1.axis("off")
    # --- 右: 水圧 ---
    pw_b = gw * (H - hw)
    ax2.add_patch(plt.Rectangle((-0.4, 0), 0.4, H, fc="#bdbdbd", ec="k"))
    ax2.plot([0, 2.6], [H, H], color="#8a6d3b", lw=1.5)
    ax2.axhline(H - hw, color="#2b7fc0", ls="--", lw=1.2)
    ax2.fill([0, pw_b * 0.05, 0], [H - hw, 0, 0], color="#2b7fc0", alpha=0.6)
    ax2.text(pw_b * 0.05 + 0.05, 0.1, f"γw·z'={pw_b:.0f} kN/m²",
             fontproperties=jp, fontsize=8, color="#2b7fc0")
    ax2.set_title("② 水圧（静水圧・別途加算）", fontproperties=jp, fontsize=10)
    ax2.set_xlim(-0.6, 2.2)
    ax2.set_ylim(-0.3, H + 0.8)
    ax2.axis("off")
    fig.suptitle("地下水位がある場合：土圧と水圧を分けて評価", fontproperties=jp)
    return savefig(fig, "fig5_water.png")


figs = {
    "f1": fig_3types(),
    "f2": fig_coulomb(),
    "f3": fig_surcharge(),
    "f4": fig_earth(),
    "f5": fig_water(),
}
print("figures done:", figs)

# ===========================================================================
# Excel 構築
# ===========================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

wb = Workbook()

C_TITLE = "1F4E79"
C_HEAD = "2E75B6"
C_SUB = "DEEBF7"
C_ANS = "E2EFDA"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

f_title = Font(name="MS PGothic", size=15, bold=True, color="FFFFFF")
f_head = Font(name="MS PGothic", size=11, bold=True, color="FFFFFF")
f_sub = Font(name="MS PGothic", size=11, bold=True, color="1F4E79")
f_body = Font(name="MS PGothic", size=10)
f_ans = Font(name="MS PGothic", size=10, color="375623")
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)


def setup(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False


def title_row(ws, row, text, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = f_title
    c.fill = PatternFill("solid", fgColor=C_TITLE)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 30


def head(ws, row, text, span=8, color=C_HEAD):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = f_head
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 22


def body(ws, row, text, span=8, ans=False, h=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = f_ans if ans else f_body
    c.alignment = wrap
    if ans:
        c.fill = PatternFill("solid", fgColor=C_ANS)
    if h:
        ws.row_dimensions[row].height = h


def table(ws, start_row, headers, rows, col1=1, widthset=None):
    r = start_row
    for j, htxt in enumerate(headers):
        c = ws.cell(r, col1 + j, htxt)
        c.font = Font(name="MS PGothic", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=C_HEAD)
        c.alignment = center
        c.border = border
    for data in rows:
        r += 1
        for j, v in enumerate(data):
            c = ws.cell(r, col1 + j, v)
            c.font = f_body
            c.alignment = center if j > 0 else wrap
            c.border = border
            if r % 2 == 0:
                c.fill = PatternFill("solid", fgColor="F2F7FC")
    return r


def put_img(ws, path, anchor, w=None):
    img = XLImage(path)
    if w:
        ratio = w / img.width
        img.width = w
        img.height = int(img.height * ratio)
    ws.add_image(img, anchor)


# ---------- 目次 ----------
ws = wb.active
ws.title = "目次"
setup(ws, [4, 18, 60, 14])
title_row(ws, 1, "土圧 問題集（ゼネコン構造設計部・新人向け）", span=4)
body(ws, 2, "テーマ No.7：擁壁・地下外壁・山留め壁に作用する土圧。各シートに『問題』、最後に『解答』シート。"
            "図を見ながら、空欄の表を電卓で埋めていく形式。", span=4, h=42)
r = table(ws, 4,
          ["No.", "シート", "到達目標", "主な図"],
          [["7-1", "No.7-1 土圧の種類", "静止・主動・受動土圧の違いを説明できる", "変位-土圧図"],
           ["7-2", "No.7-2 クーロン式", "クーロン式で主動土圧係数 Ka を算出できる", "擁壁断面(角度)"],
           ["7-3", "No.7-3 上載荷重", "上載荷重の説明・設定ができる", "土圧分布図"],
           ["7-4", "No.7-4 土圧の算出", "土圧（合力・作用点）が算出できる", "三角形分布図"],
           ["7-5", "No.7-5 地下水位", "地下水位がある場合の土圧を算出できる", "土圧+水圧図"]],
          )
body(ws, r + 2, "凡例：土の定数（γ・φ・c・地下水位）は本来ボーリング報告書から引く。本問では代表値を与える。"
                "有効数字は3桁、γw=9.8 kN/m² とする。", span=4, h=42)

# ---------- No.7-1 ----------
ws = wb.create_sheet("No.7-1 土圧の種類")
setup(ws, [10, 14, 14, 14, 12, 12, 12, 12])
title_row(ws, 1, "No.7-1　静止土圧・主動土圧・受動土圧の違い")
head(ws, 3, "■ 解説図：壁の変位と土圧係数の関係")
put_img(ws, figs["f1"], "A4", w=470)
head(ws, 22, "■ 問題 1　概念の穴埋め")
body(ws, 23, "次の文の空欄を埋めよ。", h=18)
body(ws, 24, "・壁が全く動かないときに作用する土圧を【　　①　　】土圧といい、係数は K₀=1−sinφ で表す。", h=18)
body(ws, 25, "・壁が背面土から離れる側にわずかに変位すると土圧は【減少／増加】し、最小値の【　　②　　】土圧になる。", h=18)
body(ws, 26, "・壁が背面土を押し込む側に大きく変位すると土圧は増加し、最大値の【　　③　　】土圧になる。", h=18)
body(ws, 27, "・係数の大小関係は　Ka【　＜／＞　】K₀【　＜／＞　】Kp　である。", h=18)
body(ws, 28, "・主動状態に必要な変位は受動状態より【大きい／小さい】（＝主動はわずかな変位で発生する）。", h=18)
head(ws, 30, "■ 問題 2　設計対象はどの土圧か（K0 / Ka / Kp を記入）")
r = table(ws, 31,
          ["No.", "設計・検討の対象", "使う土圧（記入）"],
          [["(1)", "地下1階RC外壁（剛で変位しない）の曲げ設計", ""],
           ["(2)", "重力式擁壁 H=4m の転倒・滑動の検討", ""],
           ["(3)", "山留め鋼矢板の根入れ前面の抵抗", ""],
           ["(4)", "グラウンドアンカー定着部（アンカーブロック）の抜け抵抗", ""],
           ["(5)", "ボックスカルバート側壁（拘束され変位小）", ""],
           ["(6)", "もたれ式擁壁 H=6m の背面主働側", ""]],
          )
body(ws, r + 2, "■ 問題 3　記述：『静止・主動・受動』の違いを、次の語をすべて使って3行で説明せよ。"
                "　〔 壁の変位／最小・最大／動員される摩擦角 〕", h=40)

# ---------- No.7-2 ----------
ws = wb.create_sheet("No.7-2 クーロン式")
setup(ws, [10, 14, 14, 14, 12, 12, 12, 12])
title_row(ws, 1, "No.7-2　クーロン式で主動土圧係数 Ka を算出する")
head(ws, 3, "■ 解説図：クーロン式の各パラメータ")
put_img(ws, figs["f2"], "A4", w=440)
head(ws, 24, "■ 基本式")
body(ws, 25, "ランキン式（壁面摩擦・傾斜なし）：　Ka = tan²(45° − φ/2)", h=18)
body(ws, 26, "クーロン式：　Ka = cos²(φ−θ) ÷ [ cos²θ·cos(δ+θ)·(1+√R)² ]，"
             "　R = sin(φ+δ)·sin(φ−β) ÷ [cos(δ+θ)·cos(θ−β)]", h=32)
head(ws, 28, "■ 問題 1　ランキン式で Ka を計算（壁面摩擦・傾斜なし、有効数字3桁）")
r = table(ws, 29,
          ["No.", "φ", "45°−φ/2", "Ka=tan²(45−φ/2)"],
          [["(1)", "25°", "", ""], ["(2)", "30°", "", ""],
           ["(3)", "35°", "", ""], ["(4)", "40°", "", ""]])
head(ws, r + 2, "■ 問題 2　クーロン式で Ka を計算")
body(ws, r + 3, "もたれ式擁壁：φ=30°、壁傾斜 θ=10°、壁面摩擦 δ=2φ/3=20°、地表傾斜 β=0°。"
                "上式に代入して Ka を求めよ（電卓可、degree モードに注意）。", span=8, h=40)
body(ws, r + 4, "計算欄：R=______、(1+√R)²=______、分子 cos²(φ−θ)=______、Ka=______", h=18)
head(ws, r + 6, "■ 問題 3　考察")
body(ws, r + 7, "(1) φ=30°のとき、ランキンの Ka とクーロン(δ=20°)の Ka はどちらが小さいか。理由とともに答えよ。", h=30)
body(ws, r + 8, "(2) θ=δ=β=0 のときクーロン式がランキン式に一致することを式で確認せよ。", h=18)
body(ws, r + 9, "(3) 角度の取り違え（45+φ/2 と 45−φ/2）をすると何が起こるか。Ka と Kp の関係から述べよ。", h=30)

# ---------- No.7-3 ----------
ws = wb.create_sheet("No.7-3 上載荷重")
setup(ws, [10, 16, 16, 14, 12, 12, 12, 12])
title_row(ws, 1, "No.7-3　上載荷重の説明・設定")
head(ws, 3, "■ 解説図：上載荷重による土圧（矩形）と自重土圧（三角形）")
put_img(ws, figs["f3"], "A4", w=460)
head(ws, 24, "■ 問題 1　上載荷重の標準値を設定")
r = table(ws, 25,
          ["No.", "背面の用途", "上載荷重 q (kN/m²)を記入"],
          [["(1)", "一般地表面・宅地", ""],
           ["(2)", "駐車場・通路", ""],
           ["(3)", "消防車進入路（重車両）", ""],
           ["(4)", "工事中（建機・資材仮置き）", ""]])
head(ws, r + 2, "■ 問題 2　土被り換算")
body(ws, r + 3, "γ=18 kN/m³ の地盤で、上載荷重 q=10 kN/m² は土層厚 何 m に相当するか（h=q/γ）。", h=18)
head(ws, r + 5, "■ 問題 3　作用点の理解")
body(ws, r + 6, "高さ H の壁で、(a)自重土圧の合力と (b)上載土圧の合力は、それぞれ壁下端から何の高さに作用するか。"
                "また、なぜ作用点が異なるのかを分布形状（三角形／矩形）から1行で説明せよ。", span=8, h=46)
r2 = table(ws, r + 8,
           ["成分", "圧力分布", "合力(式)", "作用点(下端から)"],
           [["自重土圧", "三角形", "½KaγH²", ""],
            ["上載土圧", "矩形", "KaqH", ""]])
head(ws, r2 + 2, "■ 問題 4　施工時 vs 供用時")
body(ws, r2 + 3, "工事中の重機荷重 q=20 kN/m²（短期）と、供用時の駐車場荷重 q=10 kN/m²。"
                 "擁壁設計でこの2つをどう扱い分けるか（荷重の組合せ・許容応力度の長期/短期）を述べよ。", span=8, h=44)

# ---------- No.7-4 ----------
ws = wb.create_sheet("No.7-4 土圧の算出")
setup(ws, [10, 16, 16, 14, 14, 12, 12, 12])
title_row(ws, 1, "No.7-4　土圧の算出（合力・作用点・転倒モーメント）")
head(ws, 3, "■ 解説図：主動土圧の三角形分布")
put_img(ws, figs["f4"], "A4", w=420)
head(ws, 23, "■ 問題（もたれ式擁壁）")
body(ws, 24, "条件：壁高 H=5.0m、γ=18 kN/m³、φ=30°（→ Ka=0.33）、壁面摩擦なし（ランキン）、"
             "上載荷重 q=10 kN/m²、地下水位なし。", span=8, h=32)
head(ws, 26, "(1) 土圧分布の値（kN/m²）")
r = table(ws, 27,
          ["項目", "式", "値(記入)"],
          [["下端の自重土圧 paγ", "Ka·γ·H", ""],
           ["上載土圧 paq（一様）", "Ka·q", ""]])
head(ws, r + 2, "(2) 合力と作用点")
r = table(ws, r + 3,
          ["成分", "合力(kN/m)記入", "作用点 z(m)記入", "Pa·z(kN·m/m)記入"],
          [["自重土圧 ½KaγH²", "", "H/3=", ""],
           ["上載土圧 KaqH", "", "H/2=", ""],
           ["合計", "", "←加重平均", ""]])
head(ws, r + 2, "(3) 設問")
body(ws, r + 3, "① 全主動土圧合力 Pa（自重＋上載）を求めよ。", h=18)
body(ws, r + 4, "② 合力の作用点（壁下端からの高さ）をモーメント加重平均で求めよ。", h=18)
body(ws, r + 5, "③ 壁下端まわりの転倒モーメント M=ΣPa·z を求めよ。", h=18)
body(ws, r + 6, "④ 上載 q を 0 にすると Pa は何 % 減るか。上載の寄与の大きさを述べよ。", h=18)

# ---------- No.7-5 ----------
ws = wb.create_sheet("No.7-5 地下水位")
setup(ws, [10, 16, 16, 14, 14, 12, 12, 12])
title_row(ws, 1, "No.7-5　地下水位がある場合の土圧")
head(ws, 3, "■ 解説図：土圧（左）と水圧（右）を分けて評価")
put_img(ws, figs["f5"], "A4", w=540)
head(ws, 24, "■ 問題")
body(ws, 25, "条件：壁高 H=5.0m、地下水位 GL−2.0m、上載なし、φ=30°（Ka=0.33）。"
             "γt=18（水位上）、γsat=20（水位下）、γw=9.8 kN/m³。γ'=γsat−γw=10.2。", span=8, h=34)
head(ws, 27, "(1) 有効土圧（水位下は γ' を使う）")
r = table(ws, 28,
          ["深さ", "有効鉛直応力σ'v", "土圧 pa=Ka·σ'v(記入)"],
          [["GL（z=0）", "0", ""],
           ["水位 GL-2m", "γt·2", ""],
           ["下端 GL-5m", "γt·2 + γ'·3", ""]])
head(ws, r + 2, "(2) 水圧（静水圧、Ka は掛けない）")
r = table(ws, r + 3,
          ["深さ", "水圧 pw=γw·z'(記入)"],
          [["水位 GL-2m（z'=0）", ""],
           ["下端 GL-5m（z'=3m）", ""]])
head(ws, r + 2, "(3) 設問")
body(ws, r + 3, "① 有効土圧の合力 Pa（土圧図の面積）と、その作用点を求めよ。", h=18)
body(ws, r + 4, "② 水圧の合力 Pw（三角形）と作用点を求めよ。", h=18)
body(ws, r + 5, "③ 全側圧 Pa+Pw を求めよ。", h=18)
body(ws, r + 6, "④ 同じ壁を『地下水位なし・全層 γt=18』とした場合の主動土圧合力（=½KaγH²）と比較し、"
                "水位がある方が何倍になるか述べよ。", span=8, h=32)
body(ws, r + 7, "⑤ 背面排水（裏込め砕石＋水抜き）で水圧を0にできる。その設計的・経済的意義を述べよ。", h=30)

# ---------- 解答 ----------
ws = wb.create_sheet("解答")
setup(ws, [4, 22, 60, 14])
title_row(ws, 1, "解答・解説", span=4)
row = 2


def ans_head(t):
    global row
    head(ws, row, t, span=4)
    row += 1


def ans(t, h=None):
    global row
    body(ws, row, t, span=4, ans=True, h=h)
    row += 1


ans_head("No.7-1　土圧の種類")
ans("問1：①静止　②主動　③受動／土圧は離れる側で『減少』して最小=主動／"
    "大小は Ka＜K₀＜Kp／主動に必要な変位は受動より『小さい』。", h=46)
ans("問2：(1)K0 剛で変位しない地下壁 (2)Ka 擁壁は前傾でわずかに動く (3)Kp 前面抵抗 "
    "(4)Kp アンカー抵抗 (5)K0 拘束され変位小 (6)Ka 背面主働。", h=46)
ans("問3例：壁の変位の向き・量で土圧は変わる。離れる側に微小変位すると土の強度(摩擦角)が"
    "動員されて土圧は最小=主動Kaに、押し込むと最大=受動Kpになる。変位ゼロは中間の静止K0。", h=46)

ans_head("No.7-2　クーロン式")
ans("問1：(1)φ25°→45−12.5=32.5°→Ka=tan²32.5=0.406　(2)φ30°→30°→0.333　"
    "(3)φ35°→27.5°→0.271　(4)φ40°→25°→0.217。", h=46)
ans("問2：φ=30,θ=10,δ=20,β=0。R=sin50·sin30/(cos30·cos10)=0.766·0.5/(0.866·0.985)=0.449，"
    "√R=0.670，(1+√R)²=2.789。分子cos²(20)=0.883。分母cos²10·cos30·2.789="
    "0.970·0.866·2.789=2.343。Ka=0.883/2.343≒0.377。", h=72)
ans("問3：(1)壁面摩擦δを見込むクーロンの方が小さくなりそうだが、本例は壁傾斜θ=10°の影響で"
    "ランキン0.333よりやや大きい0.377。δ単独なら土圧は減る方向、θ(前傾でなく後傾)は増やす方向。"
    "(2)θ=δ=β=0でKa=cos²φ/(1+sinφ)²=(1−sinφ)/(1+sinφ)=tan²(45−φ/2)。"
    "(3)45+φ/2 を使うと Kp(受動)になり、主動の3〜10倍の過大な土圧。係数を取り違えると設計が成立しない。", h=86)

ans_head("No.7-3　上載荷重")
ans("問1：(1)10 (2)10〜12 (3)12〜20 (4)施工計画で別途設定（例20、短期扱い）。"
    "※数値は指針・条例で確認。", h=32)
ans("問2：h=q/γ=10/18=0.56m 相当。", h=18)
ans("問3：自重土圧=三角形分布→合力は下端からH/3。上載土圧=深さに依らず一様の矩形分布→合力はH/2。"
    "分布形が違うため作用点(図心)が異なる。", h=46)
ans("問4：施工時の重機20kN/m²は『短期(暴風・施工時)』荷重として短期許容応力度で照査、"
    "供用時の駐車場10kN/m²は常時=長期で照査。両者は同時に作用しない想定なら別ケースで包絡。", h=46)

ans_head("No.7-4　土圧の算出（H=5, γ=18, Ka=0.33, q=10）")
ans("(1) 下端自重土圧 paγ=Ka·γ·H=0.33·18·5=29.7≒30 kN/m²。上載土圧 paq=Ka·q=0.33·10=3.3 kN/m²。", h=32)
ans("(2) 自重合力=½·0.33·18·5²=74.3 kN/m（z=H/3=1.67m）。上載合力=0.33·10·5=16.5 kN/m（z=H/2=2.5m）。", h=32)
ans("(3) ①Pa=74.3+16.5=90.8 kN/m。②作用点 z=(74.3·1.67+16.5·2.5)/90.8=(124.1+41.3)/90.8=1.82m。"
    "③M=ΣPa·z=124.1+41.3=165.4 kN·m/m。④q=0でPa=74.3、減少分16.5/90.8=18%減。"
    "上載は全土圧の約2割を占め無視できない。", h=72)

ans_head("No.7-5　地下水位（H=5, 水位GL-2, Ka=0.33, γt=18, γ'=10.2, γw=9.8）")
ans("(1) σ'v：GL=0／GL-2=18·2=36→pa=0.33·36=11.9 kN/m²／"
    "GL-5=36+10.2·3=66.6→pa=0.33·66.6=22.0 kN/m²。", h=46)
ans("(2) 水圧：GL-2(z'=0)=0／GL-5(z'=3)=9.8·3=29.4 kN/m²。", h=18)
ans("(3) ①有効土圧合力Pa＝上段三角形(0→11.9,厚2m)=11.9 ＋ 下段台形(11.9→22.0,厚3m)。"
    "台形=矩形11.9·3=35.7＋三角½·(22.0−11.9)·3=15.2 → Pa=11.9+35.7+15.2=62.8 kN/m。"
    "作用点(下端から)≒1.78m。②水圧Pw=½·29.4·3=44.1 kN/m、作用点=3m層の1/3=下端から1.0m。"
    "③全側圧=62.8+44.1=106.9 kN/m。", h=86)
ans("④水位なし全層γt：Pa=½·0.33·18·5²=74.3 kN/m。水位ありは106.9/74.3≒1.44倍。"
    "（土圧は減るが水圧が加わり、合計は約1.4倍に増える）", h=46)
ans("⑤背面排水で水圧0にすれば全側圧は土圧のみ＝大幅低減。壁厚・配筋・基礎を縮小でき経済的。"
    "水抜き穴・裏込め砕石・不織布フィルターは擁壁の必須ディテール。", h=46)

XLSX = os.path.join(OUT_DIR, "土圧問題集.xlsx")
wb.save(XLSX)
print("saved:", XLSX)
