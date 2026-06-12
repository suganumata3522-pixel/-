# -*- coding: utf-8 -*-
"""鉄筋比 問題集（図つき）Excel 生成スクリプト。
出力: docs/steel_ratio/鉄筋比問題集.xlsx
No.5-1 柱pg・梁pt  /  5-2 柱梁pw  /  5-3 壁ps  /  5-4 釣合鉄筋比 pb
5-5 断面に対する適切な鉄筋量の判断
RC造マンションの設計担当を想定。
"""
import os
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib import font_manager as fm

FONT_PATH = "/usr/share/fonts/truetype/fonts-japanese-gothic.ttf"
jp = fm.FontProperties(fname=FONT_PATH)
fm.fontManager.addfont(FONT_PATH)
plt.rcParams["font.family"] = jp.get_name()
plt.rcParams["axes.unicode_minus"] = False

OUT = "docs/steel_ratio"
FIG = os.path.join(OUT, "figures")
os.makedirs(FIG, exist_ok=True)


def save(fig, name):
    p = os.path.join(FIG, name)
    fig.savefig(p, dpi=130, bbox_inches="tight")
    plt.close(fig)
    return p


# ===========================================================================
# 図 5-1: 柱 pg と 梁 pt の定義
# ===========================================================================
def fig_5_1():
    fig, axes = plt.subplots(1, 2, figsize=(13, 5.4))
    # --- (a) 柱断面 ---
    ax = axes[0]
    B, D = 4.0, 4.0
    ax.add_patch(mpatches.Rectangle((0, 0), B, D, fc="#d9d9d9", ec="k", lw=1.2))
    # 主筋（12-D25）
    nb = 4
    cov = 0.35
    pts = np.linspace(cov, B - cov, nb)
    for px in pts:
        for py in [cov, D - cov]:
            ax.add_patch(plt.Circle((px, py), 0.18, fc="k", zorder=4))
    for py in pts[1:-1]:
        for px in [cov, B - cov]:
            ax.add_patch(plt.Circle((px, py), 0.18, fc="k", zorder=4))
    # 帯筋（外周）
    ax.add_patch(mpatches.Rectangle((cov - 0.18, cov - 0.18),
                                     B - 2 * cov + 0.36,
                                     D - 2 * cov + 0.36,
                                     fc="none", ec="#c00000", lw=1.2,
                                     ls="-"))
    # 寸法
    ax.annotate("", xy=(0, -0.35), xytext=(B, -0.35),
                arrowprops=dict(arrowstyle="<|-|>", color="k", lw=1))
    ax.text(B / 2, -0.7, "b", fontproperties=jp, ha="center", fontsize=11)
    ax.annotate("", xy=(-0.35, 0), xytext=(-0.35, D),
                arrowprops=dict(arrowstyle="<|-|>", color="k", lw=1))
    ax.text(-0.7, D / 2, "D", fontproperties=jp, va="center", ha="center",
            fontsize=11, rotation=90)
    ax.text(B + 0.3, D, "12-D25", fontproperties=jp, fontsize=9,
            color="#1f7a1f", va="top")
    # 式
    ax.text(B / 2, -1.6, "pg = ΣAg / (b·D)\n= 主筋全断面積 / 柱の全断面積",
            fontproperties=jp, ha="center", fontsize=11, color="#1f4e79",
            fontweight="bold")
    ax.set_xlim(-1.2, B + 2.0)
    ax.set_ylim(-2.4, D + 0.8)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(a) 柱 主筋比 pg", fontproperties=jp,
                 fontsize=12, fontweight="bold")

    # --- (b) 梁断面 ---
    ax = axes[1]
    Bb, Db = 2.5, 4.4
    dd = 0.5  # 上下かぶり位置
    d = Db - dd
    ax.add_patch(mpatches.Rectangle((0, 0), Bb, Db, fc="#d9d9d9",
                                     ec="k", lw=1.2))
    # 上端筋（引張側=正曲げで下、負曲げで上）3本
    for px in np.linspace(0.4, Bb - 0.4, 3):
        ax.add_patch(plt.Circle((px, Db - dd), 0.17, fc="k", zorder=4))
    for px in np.linspace(0.4, Bb - 0.4, 3):
        ax.add_patch(plt.Circle((px, dd), 0.17, fc="#1f7a1f", zorder=4))
    # あばら筋 外周
    ax.add_patch(mpatches.Rectangle((0.18, 0.18), Bb - 0.36, Db - 0.36,
                                     fc="none", ec="#c00000", lw=1.2))
    # 寸法
    ax.annotate("", xy=(0, -0.35), xytext=(Bb, -0.35),
                arrowprops=dict(arrowstyle="<|-|>", color="k", lw=1))
    ax.text(Bb / 2, -0.7, "b", fontproperties=jp, ha="center", fontsize=11)
    ax.annotate("", xy=(-0.35, 0), xytext=(-0.35, Db),
                arrowprops=dict(arrowstyle="<|-|>", color="k", lw=1))
    ax.text(-0.7, Db / 2, "D", fontproperties=jp, va="center", ha="center",
            fontsize=11, rotation=90)
    ax.annotate("", xy=(Bb + 0.35, dd), xytext=(Bb + 0.35, Db),
                arrowprops=dict(arrowstyle="<|-|>", color="#c00000", lw=1))
    ax.text(Bb + 0.65, d / 2 + dd / 2, "d\n(有効せい)", fontproperties=jp,
            va="center", fontsize=9, color="#c00000")
    ax.text(Bb / 2, Db + 0.35, "上端筋 Ast", fontproperties=jp,
            ha="center", fontsize=9, color="k")
    ax.text(Bb / 2, -1.6, "pt = Ast / (b·d)\n= 引張側主筋断面積 / 有効断面積",
            fontproperties=jp, ha="center", fontsize=11, color="#1f4e79",
            fontweight="bold")
    ax.set_xlim(-1.4, Bb + 2.6)
    ax.set_ylim(-2.4, Db + 0.8)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(b) 梁 引張鉄筋比 pt", fontproperties=jp,
                 fontsize=12, fontweight="bold")
    fig.suptitle("図 5-1  柱主筋比 pg と 梁引張鉄筋比 pt の定義",
                 fontproperties=jp, fontsize=13, fontweight="bold")
    return save(fig, "fig5-1_pg_pt.png")


# ===========================================================================
# 図 5-2: 柱・梁のせん断補強筋比 pw
# ===========================================================================
def fig_5_2():
    fig, axes = plt.subplots(1, 2, figsize=(13.5, 5.6))
    # --- (a) 柱の帯筋（立面） ---
    ax = axes[0]
    B, H = 1.8, 5.4
    s = 1.0  # ピッチ
    ax.add_patch(mpatches.Rectangle((0, 0), B, H, fc="#d9d9d9", ec="k", lw=1.2))
    # 主筋（縦線）
    for x in [0.25, B - 0.25]:
        ax.plot([x, x], [0, H], color="k", lw=2)
    # 帯筋（水平線）
    for i in range(int(H / s) + 1):
        y = i * s + 0.3
        if y > H - 0.1:
            break
        ax.plot([0.25, B - 0.25], [y, y], color="#c00000", lw=2)
    # ピッチ寸法
    ax.annotate("", xy=(B + 0.3, 0.3), xytext=(B + 0.3, 0.3 + s),
                arrowprops=dict(arrowstyle="<|-|>", color="k", lw=1))
    ax.text(B + 0.5, 0.3 + s / 2, "s", fontproperties=jp, va="center",
            color="k", fontsize=11)
    # 帯筋1組面積
    ax.annotate("帯筋1組\n断面積 aw\n(脚数×Aw)",
                xy=(B / 2, 2.3), xytext=(B + 1.5, 2.8),
                fontproperties=jp, fontsize=9, color="#c00000",
                arrowprops=dict(arrowstyle="->", color="#c00000"))
    ax.text(B / 2, -1.0, "pw = aw / (b·s)\n柱・梁ともに同じ定義",
            fontproperties=jp, ha="center", fontsize=11, color="#1f4e79",
            fontweight="bold")
    ax.set_xlim(-0.5, B + 3.5)
    ax.set_ylim(-1.8, H + 0.5)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(a) 柱の帯筋（立面）", fontproperties=jp,
                 fontsize=12, fontweight="bold")

    # --- (b) 柱断面の帯筋（2脚 + サブタイ） ---
    ax = axes[1]
    B2, D2 = 4.0, 4.0
    ax.add_patch(mpatches.Rectangle((0, 0), B2, D2, fc="#d9d9d9",
                                     ec="k", lw=1.2))
    nb = 4
    cov = 0.35
    pts = np.linspace(cov, B2 - cov, nb)
    for px in pts:
        for py in [cov, D2 - cov]:
            ax.add_patch(plt.Circle((px, py), 0.18, fc="k", zorder=4))
    for py in pts[1:-1]:
        for px in [cov, B2 - cov]:
            ax.add_patch(plt.Circle((px, py), 0.18, fc="k", zorder=4))
    # 外周帯筋
    ax.add_patch(mpatches.Rectangle((cov - 0.18, cov - 0.18),
                                     B2 - 2 * cov + 0.36,
                                     D2 - 2 * cov + 0.36,
                                     fc="none", ec="#c00000", lw=1.5))
    # 中子筋（サブタイ）
    ax.plot([pts[1], pts[1]], [cov - 0.18, D2 - cov + 0.18],
            color="#c00000", lw=1.5)
    ax.plot([pts[2], pts[2]], [cov - 0.18, D2 - cov + 0.18],
            color="#c00000", lw=1.5)
    # b 方向の脚数を矢印で示す
    ax.annotate("", xy=(0.15, D2 + 0.4), xytext=(B2 - 0.15, D2 + 0.4),
                arrowprops=dict(arrowstyle="<|-|>", color="#c00000", lw=1.2))
    ax.text(B2 / 2, D2 + 0.7,
            "b 方向にかかる脚数 = 4 (外周2＋中子2)",
            fontproperties=jp, ha="center", fontsize=9, color="#c00000")
    ax.text(B2 / 2, -1.0,
            "脚数 n、1 本断面積 Aw のとき\n aw = n × Aw",
            fontproperties=jp, ha="center", fontsize=10, color="#1f4e79")
    ax.annotate("", xy=(0, -0.35), xytext=(B2, -0.35),
                arrowprops=dict(arrowstyle="<|-|>", color="k", lw=1))
    ax.text(B2 / 2, -0.7, "b", fontproperties=jp, ha="center", fontsize=11)
    ax.set_xlim(-0.8, B2 + 1.0)
    ax.set_ylim(-2.0, D2 + 1.4)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(b) 柱断面：帯筋＋中子筋（サブタイ）",
                 fontproperties=jp, fontsize=12, fontweight="bold")
    fig.suptitle("図 5-2  せん断補強筋比 pw の定義（柱の帯筋・梁のあばら筋）",
                 fontproperties=jp, fontsize=13, fontweight="bold")
    return save(fig, "fig5-2_pw.png")


# ===========================================================================
# 図 5-3: 壁の壁筋比 ps（縦筋・横筋）
# ===========================================================================
def fig_5_3():
    fig, axes = plt.subplots(1, 2, figsize=(13.5, 5.4))
    # --- (a) 壁立面 ---
    ax = axes[0]
    L, H = 5.4, 3.0
    ax.add_patch(mpatches.Rectangle((0, 0), L, H, fc="#e8c9ce", ec="k",
                                     lw=1.2, alpha=0.5))
    # 縦筋（s_v ピッチ）
    sv = 0.6
    for x in np.arange(0.3, L, sv):
        ax.plot([x, x], [0, H], color="#1f4e79", lw=1.4)
    # 横筋（s_h ピッチ）
    sh = 0.5
    for y in np.arange(0.25, H, sh):
        ax.plot([0, L], [y, y], color="#c00000", lw=1.4)
    # ピッチ表示
    ax.annotate("", xy=(0.3, H + 0.3), xytext=(0.3 + sv, H + 0.3),
                arrowprops=dict(arrowstyle="<|-|>", color="#1f4e79", lw=1))
    ax.text(0.3 + sv / 2, H + 0.55, "s（縦筋ピッチ）",
            fontproperties=jp, ha="center", fontsize=8, color="#1f4e79")
    ax.annotate("", xy=(L + 0.3, 0.25), xytext=(L + 0.3, 0.25 + sh),
                arrowprops=dict(arrowstyle="<|-|>", color="#c00000", lw=1))
    ax.text(L + 0.55, 0.25 + sh / 2, "s（横筋\nピッチ）",
            fontproperties=jp, va="center", fontsize=8, color="#c00000")
    ax.text(L / 2, -0.7, "縦筋（青）と横筋（赤）を格子状に配置（複配筋＝2 層）",
            fontproperties=jp, ha="center", fontsize=9, color="#444")
    ax.set_xlim(-0.5, L + 1.8)
    ax.set_ylim(-1.2, H + 1.2)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(a) 耐震壁の壁筋配置（立面）",
                 fontproperties=jp, fontsize=11, fontweight="bold")

    # --- (b) 壁断面と式 ---
    ax = axes[1]
    t = 1.2
    Lc = 4.5
    ax.add_patch(mpatches.Rectangle((0, 0), Lc, t, fc="#e8c9ce",
                                     ec="k", lw=1.2, alpha=0.5))
    # 縦筋（断面に対しては点）2層
    for x in np.linspace(0.3, Lc - 0.3, 8):
        ax.add_patch(plt.Circle((x, 0.18), 0.08, fc="#1f4e79"))
        ax.add_patch(plt.Circle((x, t - 0.18), 0.08, fc="#1f4e79"))
    ax.annotate("", xy=(-0.25, 0), xytext=(-0.25, t),
                arrowprops=dict(arrowstyle="<|-|>", color="k", lw=1))
    ax.text(-0.5, t / 2, "t (壁厚)", fontproperties=jp, va="center",
            ha="center", fontsize=10, rotation=90)
    ax.text(Lc / 2, -1.4,
            "ps = aw / (t·s)\naw＝1 ピッチに含まれる鉄筋の断面積（複配筋なら 2 本分）",
            fontproperties=jp, ha="center", fontsize=11, color="#1f4e79",
            fontweight="bold")
    ax.text(Lc / 2, 1.9,
            "縦筋ps と 横筋ps は別々に算出する\n（縦も横も基準値以上が必要）",
            fontproperties=jp, ha="center", fontsize=9.5, color="#c00000")
    ax.set_xlim(-1.3, Lc + 0.4)
    ax.set_ylim(-2.4, 3.2)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(b) 壁断面と壁筋比の式",
                 fontproperties=jp, fontsize=11, fontweight="bold")
    fig.suptitle("図 5-3  壁筋比 ps の定義（耐震壁・縦筋／横筋）",
                 fontproperties=jp, fontsize=13, fontweight="bold")
    return save(fig, "fig5-3_ps.png")


# ===========================================================================
# 図 5-4: 釣合鉄筋比 pb（ひずみ・応力／3 領域）
# ===========================================================================
def fig_5_4():
    fig, axes = plt.subplots(1, 3, figsize=(15, 5.0),
                              gridspec_kw={"width_ratios": [1, 1, 1.1]})
    # (a) ひずみ分布
    ax = axes[0]
    D = 3.6; d = 3.2
    eps_cu, ey = 0.003, 0.00175
    xn_b = eps_cu / (eps_cu + ey) * d
    ax.plot([0, 0], [0, D], color="k", lw=2)
    ax.fill([0, -1.5, 0], [D, D, D - xn_b], color="#bcd2ea", alpha=0.7)
    ax.fill([0, 1.5, 0], [D - d, D - d, D - xn_b],
            color="#cfead4", alpha=0.7)
    ax.text(-1.6, D - 0.05, "εc=0.003", fontproperties=jp, fontsize=9,
            color="#1f4e79", ha="right", va="top")
    ax.text(1.6, D - d + 0.05, "εs=εy", fontproperties=jp, fontsize=9,
            color="#1f7a1f", ha="left")
    ax.axhline(D - xn_b, color="#c00000", lw=1, ls="--")
    ax.text(2.0, D - xn_b + 0.08, "中立軸 xn", fontproperties=jp,
            color="#c00000", fontsize=9)
    ax.text(0, -0.8, "圧縮縁εcu と引張鉄筋εy が\n同時に成立する状態 = 釣合い",
            fontproperties=jp, ha="center", fontsize=8.5, color="#c00000")
    ax.set_xlim(-2.5, 2.5)
    ax.set_ylim(-1.6, D + 0.5)
    ax.axis("off")
    ax.set_title("(a) 釣合い時のひずみ分布", fontproperties=jp,
                 fontsize=11, fontweight="bold")

    # (b) 応力分布
    ax = axes[1]
    beta = 0.85
    a = beta * xn_b
    ax.plot([0, 0], [0, D], color="k", lw=2)
    ax.add_patch(mpatches.Rectangle((-1.4, D - a), 1.4, a,
                                     fc="#1f4e79", alpha=0.7, ec="k"))
    ax.text(-1.5, D - a / 2, "0.85·fc'", fontproperties=jp,
            color="white", ha="right", va="center", fontsize=9)
    ax.annotate("", xy=(0.2, D - d), xytext=(2.0, D - d),
                arrowprops=dict(arrowstyle="<|-", color="#1f7a1f", lw=2.0))
    ax.text(1.4, D - d - 0.45, "T = Ast·σy", fontproperties=jp,
            ha="center", fontsize=10, color="#1f7a1f")
    ax.annotate("", xy=(-0.2, D - a / 2), xytext=(-2.0, D - a / 2),
                arrowprops=dict(arrowstyle="-|>", color="#c00000", lw=2.0))
    ax.text(-1.3, D - a / 2 + 0.4, "C = 0.85·fc'·b·a",
            fontproperties=jp, ha="center", fontsize=10, color="#c00000")
    ax.text(0, -0.8, "釣合い条件： C = T → pb が決まる",
            fontproperties=jp, ha="center", fontsize=9, color="#c00000")
    ax.set_xlim(-2.6, 2.5)
    ax.set_ylim(-1.6, D + 0.7)
    ax.axis("off")
    ax.set_title("(b) 応力分布と釣合い", fontproperties=jp,
                 fontsize=11, fontweight="bold")

    # (c) M-φ 曲線（過小・釣合い・過大）
    ax = axes[2]
    phi = np.linspace(0, 1.0, 200)
    M_under = np.where(phi < 0.18, phi / 0.18 * 0.6,
                       0.6 + (phi - 0.18) * 0.1)
    M_under = np.where(phi < 0.85, M_under,
                       0.6 + (0.85 - 0.18) * 0.1 - (phi - 0.85) * 0.5)
    M_bal = np.where(phi < 0.22, phi / 0.22 * 0.78, 0.78)
    M_bal = np.where(phi < 0.32, M_bal, 0.78 - (phi - 0.32) * 3)
    M_bal = np.maximum(M_bal, 0)
    M_over = np.where(phi < 0.18, phi / 0.18 * 0.92, 0.92)
    M_over = np.where(phi < 0.22, M_over, 0.92 - (phi - 0.22) * 6)
    M_over = np.maximum(M_over, 0)
    ax.plot(phi, M_under, color="#1f7a1f", lw=2.4, label="過小鉄筋 p<pb")
    ax.plot(phi, M_bal, color="#c00000", lw=2.4, ls="--",
            label="釣合い p=pb")
    ax.plot(phi, M_over, color="#444", lw=2.4, ls=":", label="過大鉄筋 p>pb")
    ax.set_xlabel("曲率 φ", fontproperties=jp)
    ax.set_ylabel("曲げモーメント M", fontproperties=jp)
    ax.set_xlim(0, 1.0); ax.set_ylim(0, 1.1)
    ax.grid(alpha=0.3)
    ax.legend(prop=jp, fontsize=9, loc="lower right")
    ax.set_title("(c) 鉄筋比による M-φ の違い",
                 fontproperties=jp, fontsize=11, fontweight="bold")
    fig.suptitle("図 5-4  釣合鉄筋比 pb と 3 領域（過小／釣合い／過大）",
                 fontproperties=jp, fontsize=13, fontweight="bold")
    return save(fig, "fig5-4_pb.png")


# ===========================================================================
# 図 5-5: 鉄筋量の判断フロー
# ===========================================================================
def fig_5_5():
    fig, ax = plt.subplots(figsize=(13, 5.2))

    def box(x, y, w, h, text, color):
        ax.add_patch(mpatches.FancyBboxPatch(
            (x, y), w, h, boxstyle="round,pad=0.05",
            fc=color, ec="k", lw=1))
        ax.text(x + w / 2, y + h / 2, text, fontproperties=jp,
                ha="center", va="center", fontsize=9.5)

    def arrow(x1, y1, x2, y2):
        ax.annotate("", xy=(x2, y2), xytext=(x1, y1),
                    arrowprops=dict(arrowstyle="-|>", color="#444", lw=1.5))

    box(0.2, 4.0, 2.0, 1.0, "断面 b・D・d\n材料 fc'・σy 確定", "#cfe0f0")
    arrow(2.2, 4.5, 2.8, 4.5)
    box(2.8, 4.0, 2.0, 1.0, "鉄筋本数・径\nAst を仮定", "#cfe0f0")
    arrow(4.8, 4.5, 5.4, 4.5)
    box(5.4, 4.0, 2.0, 1.0, "鉄筋比 p を算出\n（pg / pt / pw / ps）",
        "#fde2c4")
    arrow(7.4, 4.5, 8.0, 4.5)
    box(8.0, 4.0, 2.6, 1.0, "上下限・pb と比較", "#fde2c4")
    # 下方フロー
    arrow(9.3, 4.0, 9.3, 3.0)
    box(7.5, 2.0, 3.5, 1.0,
        "下限未満／pb 超過？\n→ 断面増 or 配筋見直し", "#f8c0c0")
    box(2.5, 2.0, 3.5, 1.0,
        "下限以上・pb／pmax 以下？\n→ OK（応力計算へ）", "#cfead4")
    arrow(9.3, 2.0, 6.0, 1.6)
    arrow(6.0, 1.6, 4.5, 2.0)
    arrow(4.0, 2.0, 4.0, 1.2)
    box(2.5, 0.2, 3.5, 1.0,
        "曲げ・せん断応力\n度の照査へ", "#cfead4")
    ax.set_xlim(-0.2, 11.0)
    ax.set_ylim(-0.4, 5.4)
    ax.axis("off")
    ax.set_title("図 5-5  鉄筋量の判断フロー（pg・pt・pw・ps・pb の確認）",
                 fontproperties=jp, fontsize=12, fontweight="bold")
    return save(fig, "fig5-5_flow.png")


figs = {
    "f1": fig_5_1(),
    "f2": fig_5_2(),
    "f3": fig_5_3(),
    "f4": fig_5_4(),
    "f5": fig_5_5(),
}
print("figures:", list(figs.keys()))

# ===========================================================================
# Excel 構築
# ===========================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

wb = Workbook()
C_TITLE = "1F4E79"; C_HEAD = "2E75B6"; C_ANS = "E2EFDA"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
f_title = Font(name="MS PGothic", size=15, bold=True, color="FFFFFF")
f_head = Font(name="MS PGothic", size=11, bold=True, color="FFFFFF")
f_body = Font(name="MS PGothic", size=10)
f_ans = Font(name="MS PGothic", size=10, color="375623")
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)


def setup(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False


def title_row(ws, row, text, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = f_title; c.fill = PatternFill("solid", fgColor=C_TITLE)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 30


def head(ws, row, text, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = f_head; c.fill = PatternFill("solid", fgColor=C_HEAD)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 22


def body(ws, row, text, span=8, ans=False, h=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = f_ans if ans else f_body
    c.alignment = wrap
    if ans:
        c.fill = PatternFill("solid", fgColor=C_ANS)
    if h:
        ws.row_dimensions[row].height = h


def table(ws, start_row, headers, rows, col1=1):
    r = start_row
    for j, htxt in enumerate(headers):
        c = ws.cell(r, col1 + j, htxt)
        c.font = Font(name="MS PGothic", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=C_HEAD)
        c.alignment = center; c.border = border
    for data in rows:
        r += 1
        for j, v in enumerate(data):
            c = ws.cell(r, col1 + j, v)
            c.font = f_body
            c.alignment = center if j > 0 else wrap
            c.border = border
            if r % 2 == 0:
                c.fill = PatternFill("solid", fgColor="F2F7FC")
    return r


def put_img(ws, path, anchor, w=None):
    img = XLImage(path)
    if w:
        ratio = w / img.width
        img.width = w; img.height = int(img.height * ratio)
    ws.add_image(img, anchor)


# ---- 目次 ----
ws = wb.active
ws.title = "目次"
setup(ws, [4, 22, 60, 14])
title_row(ws, 1, "鉄筋比 問題集（RC マンション設計担当・新人向け）", span=4)
body(ws, 2,
     "テーマ No.5：RC 部材の鉄筋比 pg／pt／pw／ps と釣合鉄筋比 pb を、定義式・代表値・"
     "手計算・実物大の判断まで通す。図と表で覚え、Excel セルで計算する形式。",
     span=4, h=44)
r = table(ws, 4,
          ["No.", "シート", "到達目標", "図"],
          [["5-1", "5-1 柱pg・梁pt",
            "柱pg・梁pt の算出、pg 下限・pt 上下限の暗記", "断面定義"],
           ["5-2", "5-2 柱梁pw",
            "柱梁pw の算出、pw 上下限の暗記", "帯筋・あばら筋"],
           ["5-3", "5-3 壁ps",
            "壁ps の算出、耐震壁ps 上下限の暗記", "壁筋格子"],
           ["5-4", "5-4 釣合鉄筋比",
            "釣合鉄筋比 pb の説明・算出ができる", "ひずみ・M-φ"],
           ["5-5", "5-5 適切な鉄筋量",
            "断面に対する適切な鉄筋量の判断ができる", "判断フロー"],
           ["—", "上限下限値早見表",
            "暗記用の一覧表", "—"]])
body(ws, r + 2,
     "凡例：pg=柱主筋比、pt=梁引張鉄筋比、pw=せん断補強筋比（柱帯筋／梁あばら筋）、"
     "ps=壁筋比、pb=釣合鉄筋比。数値は『RC 規準』『建築基準法施行令』に基づく"
     "一般的な目安。最新版規準・自社基準も併用すること。",
     span=4, h=58)

# ---- 上限下限値 早見表 ----
ws = wb.create_sheet("上限下限値早見表")
setup(ws, [4, 16, 12, 14, 16, 30])
title_row(ws, 1, "上限下限値 早見表（暗記用）", span=6)
body(ws, 2, "下表の数値は『RC 規準 2018』『建築基準法施行令』をベースにした"
            "目安値。これを覚えてから細部規準を当たる。", span=6, h=32)
r = table(ws, 4,
          ["記号", "意味", "下限", "上限", "根拠・備考"],
          [["pg", "柱主筋比 = ΣAg/(b·D)", "0.8 %",
            "4 %（一般）／6 %（短柱・最下層）",
            "0.8% 未満は鉄筋無視・無筋扱い。4%超は配筋過密・施工性低下"],
           ["pt", "梁引張鉄筋比 = Ast/(b·d)",
            "0.4 %（小梁は 0.2 % 程度可）",
            "下端筋 1.8 %、上端筋 0.8 % 程度（pmax = 0.75·pb 目安）",
            "上限は釣合鉄筋比 pb の 0.75 倍程度（靭性確保）"],
           ["pw（柱帯筋）", "せん断補強筋比 = aw/(b·s)",
            "0.2 %（基準法令）", "0.6〜1.2 %（柱中間部）", "脚部・梁端は密に"],
           ["pw（梁あばら筋）", "せん断補強筋比 = aw/(b·s)",
            "0.2 %（基準法令）", "0.8〜1.2 %（梁端部）",
            "梁端は密、中央は粗のピッチ変化が一般"],
           ["ps（壁筋）", "壁筋比 = aw/(t·s)",
            "0.25 %（縦・横ともに）",
            "1.2 % 程度", "複配筋（2 層）が原則、シングル配筋は薄壁のみ"],
           ["pb", "釣合鉄筋比 = 0.85·β1·(fc'/σy)·(εcu/(εcu+εy))",
            "—", "—（計算で求める）",
            "fc'=24, σy=345 で pb ≒ 3.2 %、pmax = 0.75·pb ≒ 2.4 %"]])
body(ws, r + 2,
     "覚え方のコツ：pg は『最低 0.8、最大 4、短柱なら 6』／"
     "pt は『下限 0.4、上限はだいたい 0.75·pb』／"
     "pw（柱・梁）は『最低 0.2、上限 1.2』／"
     "ps（壁）は『縦横とも最低 0.25、上限 1.2』。"
     "『0.2／0.25／0.4／0.8』が下限値、『1.2／4／6』が上限値。",
     span=6, h=72)

# ---- 5-1 ----
ws = wb.create_sheet("5-1 柱pg・梁pt")
setup(ws, [8, 16, 18, 16, 14, 12, 12, 12])
title_row(ws, 1, "5-1  柱 pg・梁 pt の算出と上下限値")
head(ws, 3, "■ 図 5-1  pg と pt の定義")
put_img(ws, figs["f1"], "A4", w=820)
head(ws, 30, "■ 問題 1  定義の穴埋め")
body(ws, 31, "(1) 柱主筋比 pg は、主筋全断面積 ΣAg を【 ① 】(=b·D) で割った比である。"
             "下限値は通常【 ② 】%、上限値は柱中間部で【 ③ 】%、短柱・最下層で【 ④ 】%。",
     h=44)
body(ws, 32, "(2) 梁引張鉄筋比 pt は、引張側主筋断面積 Ast を【 ⑤ 】(=b·d) で割った比である。"
             "下限値は通常【 ⑥ 】%、上限値は釣合鉄筋比 pb の【 ⑦ 】倍を目安にする。", h=44)
head(ws, 34, "■ 問題 2  柱 pg の計算（マンション中柱）")
body(ws, 35, "柱断面 b×D=800×800 mm、主筋 12-D25（1 本 As=507 mm²）。",
     h=22)
r = table(ws, 37,
          ["項目", "値（記入）"],
          [["ΣAg = 12 × 507 (mm²)", ""],
           ["柱全断面積 b·D (mm²)", ""],
           ["pg (%)", ""],
           ["下限 0.8% との比較（OK/NG）", ""],
           ["上限 4% との比較（OK/NG）", ""]])
head(ws, r + 2, "■ 問題 3  梁 pt の計算（マンション大梁・上端筋）")
body(ws, r + 3, "梁断面 b×D=400×700、d=640 mm、上端筋 5-D25（As=507）、"
                "fc'=24、σy=345 とする。", h=22)
r = table(ws, r + 5,
          ["項目", "値（記入）"],
          [["Ast = 5×507 (mm²)", ""],
           ["有効断面 b·d (mm²)", ""],
           ["pt (%)", ""],
           ["下限 0.4% との比較", ""],
           ["参考：pb≒3.22%（次表参照）、pmax=0.75pb≒2.42% との比較", ""]])
head(ws, r + 2, "■ 問題 4  pg 上限を超えると何が起こるか")
body(ws, r + 3, "pg を上限以上に上げると、断面性能上の問題が 2 つ生じる。"
                "(a) 施工性／配筋（あき寸法・コンクリート充填）／"
                "(b) 構造性能（軸力負担・かぶり）── それぞれ 1 行で述べよ。",
     h=58)
head(ws, r + 5, "■ 問題 5  暗記確認")
body(ws, r + 6, "pg・pt の下限・上限を諳んじよ（記述）。間違えやすい数値は"
                "『0.8％を 0.4％（梁の下限）と取り違える』こと。", h=44)

# ---- 5-2 ----
ws = wb.create_sheet("5-2 柱梁pw")
setup(ws, [8, 16, 18, 16, 14, 12, 12, 12])
title_row(ws, 1, "5-2  柱・梁のせん断補強筋比 pw の算出と上下限値")
head(ws, 3, "■ 図 5-2  pw の定義（柱帯筋・梁あばら筋）")
put_img(ws, figs["f2"], "A4", w=820)
head(ws, 30, "■ 問題 1  定義の穴埋め")
body(ws, 31, "せん断補強筋比 pw は、ピッチ s に含まれる帯筋（あばら筋）の断面積 aw を、"
             "【 ① 】(=b·s) で割った比である。下限値は【 ② 】%（建築基準法令）、"
             "上限値は概ね【 ③ 】%（過大にすると主筋座屈は防げるがコストと施工性が悪化）。",
     h=44)
body(ws, 32, "脚数 n、1 本の断面積 Aw のとき、aw = 【 ④ 】 で表す。", h=22)
head(ws, 34, "■ 問題 2  柱 pw の計算")
body(ws, 35, "柱 b=800、帯筋 D13（Aw=127 mm²）、4 脚（外周2＋中子2）、ピッチ s=100 mm。",
     h=22)
r = table(ws, 37,
          ["項目", "値（記入）"],
          [["aw = 4×127 (mm²)", ""],
           ["b·s (mm²)", ""],
           ["pw (%)", ""],
           ["下限 0.2% との比較", ""],
           ["上限 1.2% との比較", ""]])
head(ws, r + 2, "■ 問題 3  梁 pw の計算（端部）")
body(ws, r + 3, "梁 b=400、あばら筋 D10（Aw=71）、2 脚、ピッチ s=100。",
     h=22)
r = table(ws, r + 5,
          ["項目", "値（記入）"],
          [["aw = 2×71 (mm²)", ""],
           ["b·s (mm²)", ""],
           ["pw (%)", ""],
           ["下限 0.2% との比較", ""]])
head(ws, r + 2, "■ 問題 4  ピッチ変化")
body(ws, r + 3, "梁の端部 100mm、中央 200mm でピッチを変える理由を 2 つ"
                "（端部のせん断力大、靭性確保）述べよ。"
                "中央のピッチを 200 にすると pw はどう変わるか計算で確認せよ。", h=58)
head(ws, r + 5, "■ 問題 5  pw を上げすぎてはいけない理由")
body(ws, r + 6, "pw を 1.2% を超えて配置すると何が問題か。"
                "(a) 経済性（鉄筋コスト・施工手間）／"
                "(b) 構造的（せん断耐力が頭打ち、コンクリート圧壊先行）── 1 行ずつ。",
     h=58)

# ---- 5-3 ----
ws = wb.create_sheet("5-3 壁ps")
setup(ws, [8, 16, 18, 16, 14, 12, 12, 12])
title_row(ws, 1, "5-3  壁筋比 ps の算出と耐震壁の上下限値")
head(ws, 3, "■ 図 5-3  ps の定義（縦筋・横筋の格子）")
put_img(ws, figs["f3"], "A4", w=820)
head(ws, 30, "■ 問題 1  定義の穴埋め")
body(ws, 31, "壁筋比 ps は、ピッチ s に含まれる壁筋断面積 aw を【 ① 】(=t·s) で割った比。"
             "耐震壁では縦筋・横筋とも【 ② 】% 以上、上限は概ね【 ③ 】% を目安とする。"
             "複配筋（2 層）の場合、aw は【 ④ 】本分で計算する。", h=58)
head(ws, 33, "■ 問題 2  壁 ps の計算（縦筋）")
body(ws, 34, "壁厚 t=180 mm、縦筋 D13（Aw=127 mm²）、複配筋（2 層）、ピッチ s=200 mm。",
     h=22)
r = table(ws, 36,
          ["項目", "値（記入）"],
          [["aw = 2×127 (mm²)（複配筋）", ""],
           ["t·s (mm²)", ""],
           ["ps (%)", ""],
           ["下限 0.25% との比較", ""],
           ["上限 1.2% との比較", ""]])
head(ws, r + 2, "■ 問題 3  壁 ps の計算（横筋）")
body(ws, r + 3, "壁厚 t=180、横筋 D10（Aw=71）、複配筋、ピッチ s=200。", h=22)
r = table(ws, r + 5,
          ["項目", "値（記入）"],
          [["aw = 2×71 (mm²)", ""],
           ["t·s (mm²)", ""],
           ["ps (%)", ""],
           ["下限 0.25% との比較", ""]])
head(ws, r + 2, "■ 問題 4  シングル配筋と複配筋")
body(ws, r + 3, "壁厚 180mm 以下ではシングル配筋、180mm 超では複配筋（2 層）が原則。"
                "(1) シングル配筋を耐震壁に使ってよい厚さの目安は何 mm 以下か。"
                "(2) 複配筋にする構造上のメリットを 2 つ述べよ"
                "（曲げ耐力／面外座屈／靭性 から選んで）。", h=58)
head(ws, r + 5, "■ 問題 5  開口補強と ps")
body(ws, r + 6, "開口周囲には縦・横の補強筋に加え、隅角部に斜め筋を配する。"
                "壁全体の ps が下限を満たしていても、開口周囲は別途補強が必要な理由を 1 行で。",
     h=40)

# ---- 5-4 ----
ws = wb.create_sheet("5-4 釣合鉄筋比")
setup(ws, [8, 16, 18, 16, 14, 12, 12, 12])
title_row(ws, 1, "5-4  釣合鉄筋比 pb の説明と算出")
head(ws, 3, "■ 図 5-4  ひずみ／応力／M-φ")
put_img(ws, figs["f4"], "A4", w=820)
head(ws, 30, "■ 問題 1  概念の穴埋め")
body(ws, 31, "釣合鉄筋比 pb とは、圧縮縁ひずみが【 ① 】(=0.003)、引張鉄筋ひずみが"
             "【 ② 】(=σy/Es) に同時に達するときの引張鉄筋比 p = Ast/(b·d) のことである。"
             "p < pb の断面を【 ③ 】鉄筋（靭性的）、p > pb の断面を【 ④ 】鉄筋（脆性的）と呼ぶ。",
     h=58)
head(ws, 33, "■ 問題 2  公式の導出（穴埋め）")
body(ws, 34, "(1) C = T より：0.85·fc'·b·(β1·xn) = Ast·σy", h=22)
body(ws, 35, "(2) ひずみ適合：xn/d = εcu / (εcu + εy)", h=22)
body(ws, 36, "(3) p = Ast/(b·d) を代入すると", h=22)
body(ws, 37, "    pb = 0.85·β1·(fc'/σy)·(εcu / (εcu + εy))     ← この式を覚える",
     h=32)
head(ws, 39, "■ 問題 3  pb の計算（手計算）")
body(ws, 40, "εcu=0.003、Es=2.05×10⁵、β1=0.85（fc'≤28）として、次の組合せの pb と "
             "pmax=0.75·pb を求めよ（%、有効数字3桁）。", h=32)
r = table(ws, 43,
          ["No.", "fc' (N/mm²)", "σy (N/mm²)", "εy (記入)",
           "xn/d (記入)", "pb (%) (記入)", "pmax (%) (記入)"],
          [["(1)", "21", "295", "", "", "", ""],
           ["(2)", "24", "345", "", "", "", ""],
           ["(3)", "27", "345", "", "", "", ""],
           ["(4)", "30", "390", "", "", "", ""]])
head(ws, r + 2, "■ 問題 4  pmax = 0.75·pb の意味")
body(ws, r + 3, "なぜ pb そのものでなく 0.75·pb を上限とするのか。"
                "M-φ 曲線の変形能力と靭性の観点で述べよ。", h=40)

# ---- 5-5 ----
ws = wb.create_sheet("5-5 適切な鉄筋量")
setup(ws, [8, 16, 18, 16, 14, 12, 12, 12])
title_row(ws, 1, "5-5  断面に対する適切な鉄筋量の判断")
head(ws, 3, "■ 図 5-5  鉄筋量の判断フロー")
put_img(ws, figs["f5"], "A4", w=820)
head(ws, 28, "■ 問題 1  総合判定（柱・梁・壁）")
body(ws, 29, "下表 5 ケース（マンション設計）について、上限下限値・pb と比較して"
             "OK / NG を判定し、NG なら改善案を 1 つ書け。", h=40)
r = table(ws, 32,
          ["No.", "部材", "断面・配筋", "鉄筋比 (記入)", "判定 (記入)",
           "改善案 (記入)"],
          [["(1)", "中柱", "800×800、主筋 12-D25 (As=507)、fc=24",
            "pg=", "", ""],
           ["(2)", "中柱", "600×600、主筋 8-D22 (As=387)",
            "pg=", "", ""],
           ["(3)", "大梁", "400×700, d=640, 上端筋 5-D25 (As=507)、fc=24, σy=345",
            "pt=", "", ""],
           ["(4)", "大梁", "400×700, d=640, 上端筋 10-D25 (過密)",
            "pt=", "", ""],
           ["(5)", "耐震壁", "t=180、縦筋 D13@200 複配筋",
            "ps=", "", ""]])
head(ws, r + 2, "■ 問題 2  判断のフロー（穴埋め）")
body(ws, r + 3, "鉄筋量を決めるとき、まず【 ① 】（応力比から必要鉄筋量を仮算定）、"
                "次に【 ② 】（下限・上限・pb との比較）、最後に【 ③ 】"
                "（配筋のあき・かぶり・施工性）の 3 段階で確認する。", h=58)
head(ws, r + 5, "■ 問題 3  典型 NG パターンの理解")
body(ws, r + 6, "次の NG ケースで何が起こるか、それぞれ 1 行で述べよ。", h=22)
r = table(ws, r + 8,
          ["No.", "状況", "起こる問題 (記入)"],
          [["(1)", "柱 pg = 0.5% （下限割れ）", ""],
           ["(2)", "柱 pg = 5%（上限超）", ""],
           ["(3)", "梁 pt = 3%（pb 近傍）", ""],
           ["(4)", "柱 pw = 0.1%（下限割れ）", ""],
           ["(5)", "壁 ps = 0.2%（下限割れ）", ""]])
head(ws, r + 2, "■ 問題 4  マンション実務まとめ")
body(ws, r + 3, "新人として RC 梁・柱・壁の配筋をチェックする際、必ず確認する 3 つの数値"
                "（最低限）を挙げよ：① 鉄筋比（pg/pt/pw/ps）の下限・上限、② 釣合鉄筋比"
                "との比、③ ピッチ・本数・あき寸法の物理的妥当性。", h=58)

# ---- 解答 ----
ws = wb.create_sheet("解答")
setup(ws, [4, 22, 60, 14])
title_row(ws, 1, "解答・解説", span=4)
row = 2


def ah(t):
    global row
    head(ws, row, t, span=4); row += 1


def an(t, h=None):
    global row
    body(ws, row, t, span=4, ans=True, h=h); row += 1


ah("5-1  柱 pg・梁 pt")
an("問1：①柱全断面積 b·D ②0.8 ③4 ④6 ⑤有効断面積 b·d ⑥0.4 ⑦0.75。",
   h=32)
an("問2：ΣAg = 12×507 = 6,084 mm²。b·D = 800×800 = 640,000 mm²。"
   "pg = 6,084/640,000 = 0.0095 → 0.95%。"
   "下限 0.8% 以上 → OK。上限 4% 以下 → OK。", h=58)
an("問3：Ast = 5×507 = 2,535 mm²。b·d = 400×640 = 256,000 mm²。"
   "pt = 2,535/256,000 = 0.0099 → 0.99%。"
   "下限 0.4% 以上 → OK。pmax = 0.75×3.22% = 2.42% 以下 → OK。靭性的に余裕。",
   h=58)
an("問4：(a) 主筋本数増 → あき寸法不足、コンクリートが充填されず豆板。"
   "(b) かぶり厚不足、鉄筋座屈拘束のための帯筋増、軸力負担は限界（柱断面増の方が経済的）。",
   h=58)
an("問5：pg：下限 0.8%、上限 4%（短柱・最下層 6%）。"
   "pt：下限 0.4%、上限 0.75·pb（fc24/σy345 で約 2.4%）。"
   "数字の取り違え注意：『0.8％は柱・0.4％は梁の下限』。", h=58)

ah("5-2  柱梁 pw")
an("問1：①脚数 1 ピッチ分の断面 b·s ②0.2 ③1.2 ④n × Aw。", h=32)
an("問2：aw = 4×127 = 508 mm²。b·s = 800×100 = 80,000 mm²。"
   "pw = 508/80,000 = 0.00635 → 0.64%。下限 0.2% 以上 → OK。"
   "上限 1.2% 以下 → OK。", h=58)
an("問3：aw = 2×71 = 142 mm²。b·s = 400×100 = 40,000 mm²。"
   "pw = 142/40,000 = 0.00355 → 0.36%。下限 0.2% 以上 → OK。", h=44)
an("問4：梁端ではせん断力が最大、かつ塑性ヒンジが形成されるため靭性確保のために密に。"
   "中央 200mm では pw = 142/(400×200) = 0.00178 → 0.18%。0.2% 下限に届かない"
   "ためピッチを 150mm 程度に詰めるか、断面径を上げる必要がある（要注意例）。",
   h=72)
an("問5：(a) 帯筋・あばら筋を密にすると鉄筋コスト増、配筋作業のあき確保が困難。"
   "(b) せん断耐力はある量で頭打ち（コンクリート圧縮強度で律速）。"
   "それ以上の pw は無効で、靭性確保には主筋座屈拘束（中子筋）の方が有効。",
   h=58)

ah("5-3  壁 ps")
an("問1：①壁の単位幅断面 t·s ②0.25 ③1.2 ④2。", h=32)
an("問2：aw = 2×127 = 254 mm²。t·s = 180×200 = 36,000 mm²。"
   "ps = 254/36,000 = 0.00706 → 0.71%。"
   "下限 0.25% 以上 → OK。上限 1.2% 以下 → OK。", h=58)
an("問3：aw = 2×71 = 142 mm²。t·s = 180×200 = 36,000。"
   "ps = 142/36,000 = 0.00394 → 0.39%。下限 0.25% 以上 → OK。"
   "縦横とも 0.25% 以上が必要なことを確認。", h=58)
an("問4：(1) 壁厚 t ≤ 180 mm（一般に 200mm 未満）まで。"
   "(2) ①面外座屈に強い（鉄筋が両表面にあり面外曲げに有効）"
   "②壁の曲げ・せん断耐力向上 ③ひび割れ抑制と靭性向上。",
   h=58)
an("問5：開口は応力の流れを乱し、隅角部に斜めひび割れ・応力集中が生じる。"
   "壁全体の ps だけでは補えないため、開口周囲に縦横補強筋・隅角部斜め筋を"
   "別途配置する必要がある。", h=58)

ah("5-4  釣合鉄筋比")
an("問1：①εcu ②εy ③過小（under-reinforced）④過大（over-reinforced）。",
   h=22)
an("問2：式の展開は本文どおり。"
   "pb = 0.85·β1·(fc'/σy)·(εcu/(εcu+εy))。"
   "β1=0.85（fc'≤28、強度に応じ低減）、εcu=0.003、εy=σy/Es。",
   h=58)
an("問3：(1)fc=21, σy=295：εy=1.44×10⁻³, xn/d=0.676, "
   "pb=0.7225·21/295·0.676=0.0348→3.48%, pmax=2.61%。"
   "(2)fc=24, σy=345：εy=1.68×10⁻³, xn/d=0.641, "
   "pb=0.7225·24/345·0.641=0.0322→3.22%, pmax=2.42%。"
   "(3)fc=27, σy=345：xn/d=0.641, "
   "pb=0.7225·27/345·0.641=0.0363→3.63%, pmax=2.72%。"
   "(4)fc=30, σy=390：εy=1.90×10⁻³, xn/d=0.612, "
   "pb=0.7225·30/390·0.612=0.0340→3.40%, pmax=2.55%。",
   h=170)
an("問4：pb は鉄筋が降伏すると同時にコンクリートが圧壊する状態で変形余裕が少ない。"
   "0.75·pb 程度に抑えると、降伏後に塑性回転が確保でき、M-φ 曲線にプラトーが現れて"
   "靭性的破壊（粘り強い）になるため。", h=58)

ah("5-5  適切な鉄筋量")
an("問1：(1)pg=6,084/640,000=0.95% → 下限0.8%以上、上限4%以下 → OK。"
   "(2)ΣAg=8×387=3,096、b·D=600×600=360,000、pg=0.86%。下限 0.8% ぎりぎり OK。"
   "(3)Ast=5×507=2,535、b·d=400·640=256,000、pt=0.99%。OK（前出と同じ）。"
   "(4)Ast=10×507=5,070、pt=5,070/256,000=1.98%。pmax=2.42% 以下だが pb=3.22% に"
   "近く靭性が心配。改善：梁せい D を上げて d を増やす、または鉄筋径を細くする。"
   "(5)ps=2·127/(180·200)=0.71%。下限 0.25% 以上、上限 1.2% 以下 → OK。",
   h=170)
an("問2：①応力（必要鉄筋量の算定） ②鉄筋比の下限・上限・pb との比較"
   "③配筋詳細（あき・かぶり・施工性）。", h=44)
an("問3：(1)pg<0.8%：軸力負担不足、コンクリート単体扱い（無筋）と見なされ危険。"
   "(2)pg>4%：あき・かぶり不足で施工性悪化、コンクリート充填不良で耐久性低下。"
   "(3)pt≒pb：脆性破壊先行で靭性不足、地震時に粘らず急激な破壊。"
   "(4)pw<0.2%：せん断耐力不足、せん断破壊先行で脆性的。"
   "(5)ps<0.25%：壁筋不足、曲げ・せん断耐力不足、ひび割れ拡大。", h=130)
an("問4：①下限・上限：pg(0.8〜4)、pt(0.4〜0.75pb)、pw(0.2〜1.2)、ps(0.25〜1.2)。"
   "②pb との比：pt/pb、適切な余裕（実用 p < 0.5·pb 程度）。"
   "③物理的妥当性：あき寸法（鉄筋径×1.5 以上）、かぶり厚、ピッチ変化の整合。",
   h=86)

XLSX = os.path.join(OUT, "鉄筋比問題集.xlsx")
wb.save(XLSX)
print("saved:", XLSX)
