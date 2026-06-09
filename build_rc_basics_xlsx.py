# -*- coding: utf-8 -*-
"""RC造の基礎理解 問題集（図つき）Excel 生成スクリプト。
出力: docs/rc_basics/RC造基礎問題集.xlsx
No.1-1 材料の役割 / 1-2 長所短所 / 1-3 スパン・片持ち / 1-4 耐震スリット / 1-5 要求性能
"""
import os
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib import font_manager as fm

FONT_PATH = "/usr/share/fonts/truetype/fonts-japanese-gothic.ttf"
jp = fm.FontProperties(fname=FONT_PATH)
fm.fontManager.addfont(FONT_PATH)
plt.rcParams["font.family"] = jp.get_name()
plt.rcParams["axes.unicode_minus"] = False

OUT = "docs/rc_basics"
FIG = os.path.join(OUT, "figures")
os.makedirs(FIG, exist_ok=True)


def save(fig, name):
    p = os.path.join(FIG, name)
    fig.savefig(p, dpi=130, bbox_inches="tight")
    plt.close(fig)
    return p


# ===========================================================================
# 図 1-1: コンクリートと鉄筋の役割（単純梁の曲げ）
# ===========================================================================
def fig_roles():
    fig, axes = plt.subplots(1, 2, figsize=(13, 4.8))
    # --- (a) 梁の曲げと応力分担 ---
    ax = axes[0]
    L = 10
    D = 2.4
    # 梁本体
    ax.add_patch(mpatches.Rectangle((0, 0), L, D, fc="#e8e8e8", ec="k", lw=1.2))
    # 支点
    ax.plot(0.6, -0.25, "^", color="k", ms=14)
    ax.plot(L - 0.6, -0.25, "o", color="k", ms=12)
    # 等分布荷重
    for x in np.linspace(0.6, L - 0.6, 11):
        ax.annotate("", xy=(x, D), xytext=(x, D + 0.8),
                    arrowprops=dict(arrowstyle="-|>", color="#1f4e79", lw=1.3))
    ax.text(L / 2, D + 1.0, "荷重 w", fontproperties=jp, ha="center",
            color="#1f4e79", fontsize=10)
    # 中立軸
    ax.plot([0, L], [D * 0.62, D * 0.62], color="gray", ls="--", lw=1)
    ax.text(L + 0.2, D * 0.62, "中立軸", fontproperties=jp, fontsize=8, va="center")
    # 圧縮側（上）
    ax.annotate("", xy=(2.5, D - 0.15), xytext=(4.0, D - 0.15),
                arrowprops=dict(arrowstyle="-|>", color="#c00000", lw=2))
    ax.annotate("", xy=(7.5, D - 0.15), xytext=(6.0, D - 0.15),
                arrowprops=dict(arrowstyle="-|>", color="#c00000", lw=2))
    ax.text(L / 2, D - 0.45, "上側：圧縮 → コンクリートが負担",
            fontproperties=jp, ha="center", color="#c00000", fontsize=9)
    # 引張側（下）＋鉄筋
    ax.annotate("", xy=(4.0, 0.3), xytext=(2.5, 0.3),
                arrowprops=dict(arrowstyle="-|>", color="#1f7a1f", lw=2))
    ax.annotate("", xy=(6.0, 0.3), xytext=(7.5, 0.3),
                arrowprops=dict(arrowstyle="-|>", color="#1f7a1f", lw=2))
    # 主筋
    for fx in np.linspace(0.5, L - 0.5, 4):
        ax.add_patch(plt.Circle((fx, 0.35), 0.12, fc="k", zorder=5))
    ax.text(L / 2, -0.7, "下側：引張 → 鉄筋（主筋）が負担",
            fontproperties=jp, ha="center", color="#1f7a1f", fontsize=9)
    ax.set_xlim(-0.5, L + 2.0)
    ax.set_ylim(-1.3, D + 1.5)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(a) 単純梁：圧縮はコンクリート、引張は鉄筋",
                 fontproperties=jp, fontsize=11, fontweight="bold")

    # --- (b) 役割の相補性 表 ---
    ax = axes[1]
    ax.axis("off")
    ax.set_title("(b) 両者の相補的な役割", fontproperties=jp,
                 fontsize=11, fontweight="bold")
    data = [
        ["性質", "コンクリート", "鉄筋"],
        ["圧縮", "強い ○", "（座屈）"],
        ["引張", "弱い ×", "強い ○"],
        ["耐火", "強い ○", "弱い ×→被覆"],
        ["錆", "錆びない", "錆びる→かぶり"],
        ["線膨張係数", "≒ 1.0×10^-5", "≒ 1.0×10^-5 (一致)"],
    ]
    tb = ax.table(cellText=data, loc="center", cellLoc="center",
                  bbox=[0.0, 0.05, 1.0, 0.85])
    tb.auto_set_font_size(False)
    tb.set_fontsize(10)
    for (r, c), cell in tb.get_celld().items():
        cell.set_text_props(fontproperties=jp)
        cell.set_edgecolor("#999")
        if r == 0:
            cell.set_facecolor("#2E75B6")
            cell.set_text_props(fontproperties=jp, color="white")
        elif r % 2 == 0:
            cell.set_facecolor("#F2F7FC")
    ax.text(0.5, -0.02, "アルカリ性のコンクリートが鉄筋の錆を防ぎ、"
                        "線膨張係数がほぼ等しいので温度変化で剥離しない",
            transform=ax.transAxes, fontproperties=jp, ha="center",
            fontsize=8, color="#444")
    return save(fig, "fig1-1_roles.png")


# ===========================================================================
# 図 1-3: 適切なスパンと片持ち出寸法
# ===========================================================================
def fig_span():
    fig, axes = plt.subplots(1, 2, figsize=(13, 5.0))
    # --- (a) 梁せいの目安 D = L/10〜L/12 ---
    ax = axes[0]
    L = 10
    D = 1.0
    ax.add_patch(mpatches.Rectangle((0, 2), L, D, fc="#cfe0f0", ec="k", lw=1.2))
    for x in [0.4, L - 0.4]:
        ax.add_patch(mpatches.Rectangle((x - 0.25, 0), 0.5, 3.2,
                                         fc="#bdbdbd", ec="k", lw=1))
    # スパン寸法線
    ax.annotate("", xy=(0.4, 1.4), xytext=(L - 0.4, 1.4),
                arrowprops=dict(arrowstyle="<|-|>", color="k"))
    ax.text(L / 2, 1.1, "スパン L", fontproperties=jp, ha="center", fontsize=10)
    # 梁せい
    ax.annotate("", xy=(L + 0.5, 2), xytext=(L + 0.5, 3),
                arrowprops=dict(arrowstyle="<|-|>", color="#c00000"))
    ax.text(L + 0.7, 2.5, "梁せい D", fontproperties=jp, fontsize=10,
            color="#c00000", va="center")
    ax.text(L / 2, 3.6, "RC 大梁：D ≒ L/10 〜 L/12 が目安",
            fontproperties=jp, ha="center", fontsize=11, color="#c00000",
            fontweight="bold")
    ax.text(L / 2, -0.5, "例) L=8m → D≒650〜800mm　　L=10m → D≒850〜1000mm",
            fontproperties=jp, ha="center", fontsize=9, color="#444")
    ax.set_xlim(-1, L + 3)
    ax.set_ylim(-1.2, 4.2)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(a) 梁せいとスパンの関係", fontproperties=jp,
                 fontsize=11, fontweight="bold")

    # --- (b) 片持ち梁の出寸法とたわみ ---
    ax = axes[1]
    Lc = 6
    Dc = 1.0
    # 固定端の壁
    ax.add_patch(mpatches.Rectangle((-0.6, -1), 0.6, 4, fc="#bdbdbd", ec="k"))
    # 片持ち梁（根元厚い→先端薄いハンチ）
    ax.add_patch(mpatches.Polygon([(0, 2), (Lc, 2.35), (Lc, 1.9), (0, 1.0)],
                                   closed=True, fc="#cfe0f0", ec="k", lw=1.2))
    # たわみ曲線
    xs = np.linspace(0, Lc, 50)
    ys = 1.5 - 0.9 * (xs / Lc) ** 2
    ax.plot(xs, ys, color="#c00000", ls="--", lw=1.5)
    ax.annotate("先端たわみ δ", xy=(Lc, ys[-1]), xytext=(Lc - 2.5, -0.3),
                fontproperties=jp, fontsize=9, color="#c00000",
                arrowprops=dict(arrowstyle="->", color="#c00000"))
    # 出寸法
    ax.annotate("", xy=(0, 3.0), xytext=(Lc, 3.0),
                arrowprops=dict(arrowstyle="<|-|>", color="k"))
    ax.text(Lc / 2, 3.2, "片持ち出 Lc", fontproperties=jp, ha="center", fontsize=10)
    # 上端筋（引張は上側）
    ax.plot([0, Lc], [1.95, 2.25], color="#1f7a1f", lw=2)
    ax.text(Lc / 2, 2.5, "上端が引張 → 主筋は上側",
            fontproperties=jp, ha="center", fontsize=8, color="#1f7a1f")
    ax.text(Lc / 2, -1.1, "片持ち：Lc ≒ 2m 程度が目安（バルコニー等）\n"
                          "出が大きいと たわみ・ひび割れ大 → 先端梁・ハンチで対応",
            fontproperties=jp, ha="center", fontsize=9, color="#444")
    ax.set_xlim(-1.5, Lc + 1)
    ax.set_ylim(-2.0, 3.6)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(b) 片持ち梁の出寸法とたわみ", fontproperties=jp,
                 fontsize=11, fontweight="bold")
    return save(fig, "fig1-3_span.png")


# ===========================================================================
# 図 1-4: 耐震スリット
# ===========================================================================
def fig_slit():
    fig, axes = plt.subplots(1, 2, figsize=(13, 5.2))
    # --- (a) 腰壁・垂れ壁付き柱（短柱化） ---
    for idx, (ax, with_slit) in enumerate(zip(axes, [False, True])):
        # 柱
        ax.add_patch(mpatches.Rectangle((2.3, 0), 0.8, 6, fc="#bdbdbd",
                                         ec="k", lw=1.2, zorder=3))
        ax.add_patch(mpatches.Rectangle((6.1, 0), 0.8, 6, fc="#bdbdbd",
                                         ec="k", lw=1.2, zorder=3))
        # 梁
        ax.add_patch(mpatches.Rectangle((2.3, 5.4), 4.6, 0.6, fc="#888",
                                         ec="k", zorder=2))
        # 腰壁（下）と垂れ壁（上）
        ax.add_patch(mpatches.Rectangle((3.1, 0), 3.0, 1.6, fc="#cfe0f0",
                                         ec="k", lw=1, zorder=1))
        ax.add_patch(mpatches.Rectangle((3.1, 4.2), 3.0, 1.2, fc="#cfe0f0",
                                         ec="k", lw=1, zorder=1))
        ax.text(4.6, 0.8, "腰壁", fontproperties=jp, ha="center", fontsize=8)
        ax.text(4.6, 4.8, "垂れ壁", fontproperties=jp, ha="center", fontsize=8)
        # 窓
        ax.text(4.6, 3.0, "窓", fontproperties=jp, ha="center", fontsize=9,
                color="gray")
        if not with_slit:
            # 短柱の有効高さ（腰壁〜垂れ壁の間）を強調
            ax.annotate("", xy=(3.0, 1.6), xytext=(3.0, 4.2),
                        arrowprops=dict(arrowstyle="<|-|>", color="#c00000", lw=2))
            ax.text(1.5, 2.9, "短柱化\nh'小", fontproperties=jp, ha="center",
                    fontsize=9, color="#c00000")
            ax.text(4.6, -0.9, "壁が柱に拘束 → 柱の有効長さが短い\n"
                               "→ せん断破壊（脆性破壊）の危険",
                    fontproperties=jp, ha="center", fontsize=9, color="#c00000")
            ax.set_title("(a) スリット無し：腰壁・垂れ壁で短柱化",
                         fontproperties=jp, fontsize=11, fontweight="bold")
        else:
            # スリット（柱と壁の間に隙間）
            for sx in [3.1, 6.1]:
                ax.plot([sx, sx], [0, 1.6], color="#c00000", lw=3, zorder=5)
                ax.plot([sx, sx], [4.2, 5.4], color="#c00000", lw=3, zorder=5)
            ax.annotate("耐震スリット\n(柱と壁を絶縁)", xy=(3.1, 0.8),
                        xytext=(0.2, 2.2), fontproperties=jp, fontsize=9,
                        color="#c00000",
                        arrowprops=dict(arrowstyle="->", color="#c00000"))
            ax.annotate("", xy=(3.0, 0), xytext=(3.0, 5.4),
                        arrowprops=dict(arrowstyle="<|-|>", color="#1f7a1f", lw=2))
            ax.text(1.5, 2.7, "柱は\n全長有効", fontproperties=jp, ha="center",
                    fontsize=9, color="#1f7a1f")
            ax.text(4.6, -0.9, "スリットで壁を切り離す → 柱は本来の長さで変形\n"
                               "→ 靭性的な曲げ降伏に誘導",
                    fontproperties=jp, ha="center", fontsize=9, color="#1f7a1f")
            ax.set_title("(b) スリット有り：柱を壁から絶縁",
                         fontproperties=jp, fontsize=11, fontweight="bold")
        ax.set_xlim(0, 8)
        ax.set_ylim(-1.6, 6.6)
        ax.set_aspect("equal")
        ax.axis("off")
    fig.suptitle("図 1-4  耐震スリットの目的 ── 短柱せん断破壊の防止",
                 fontproperties=jp, fontsize=13, fontweight="bold")
    return save(fig, "fig1-4_slit.png")


# ===========================================================================
# 図 1-5: 要求性能（3 レベル）
# ===========================================================================
def fig_performance():
    fig, ax = plt.subplots(figsize=(11, 6))
    levels = [
        ("使用性", "中小地震 (まれ)", "無損傷・継続使用", "#a8dadc",
         "弾性範囲。ひび割れ・たわみを使用上の許容内に"),
        ("損傷制御性", "中地震 (時々)", "軽微な損傷・補修可", "#ffd180",
         "部分的な降伏は許すが補修して使い続けられる"),
        ("安全性", "大地震 (極めてまれ)", "倒壊しない・人命保護", "#f4a261",
         "大きく損傷しても靭性で倒壊を防ぎ人命を守る"),
    ]
    y = 5
    for name, quake, target, color, desc in levels:
        ax.add_patch(mpatches.FancyBboxPatch((0.5, y), 9, 1.2,
                                              boxstyle="round,pad=0.05",
                                              fc=color, ec="k", lw=1))
        ax.text(1.0, y + 0.85, name, fontproperties=jp, fontsize=13,
                fontweight="bold", va="center")
        ax.text(1.0, y + 0.35, f"対象：{quake}", fontproperties=jp,
                fontsize=9, va="center", color="#333")
        ax.text(4.3, y + 0.85, f"目標：{target}", fontproperties=jp,
                fontsize=11, va="center", color="#7a3b3b")
        ax.text(4.3, y + 0.35, desc, fontproperties=jp, fontsize=8,
                va="center", color="#444")
        y -= 1.7
    # 地震規模の矢印
    ax.annotate("", xy=(0.1, 5.0), xytext=(0.1, 7.4),
                arrowprops=dict(arrowstyle="-|>", color="#c00000", lw=2.5))
    ax.text(-0.15, 6.2, "地震が大きくなるほど許容する損傷も大きく",
            fontproperties=jp, rotation=90, fontsize=9, color="#c00000",
            va="center", ha="center")
    ax.text(5, 7.7, "RC 造の 3 つの要求性能（性能設計の考え方）",
            fontproperties=jp, ha="center", fontsize=13, fontweight="bold")
    ax.text(5, 0.3, "建築基準法：一次設計(中小地震=損傷させない) / "
                    "二次設計(大地震=倒壊させない) の 2 段階に対応",
            fontproperties=jp, ha="center", fontsize=9, color="#1f4e79")
    ax.set_xlim(-0.6, 10)
    ax.set_ylim(0, 8.2)
    ax.axis("off")
    return save(fig, "fig1-5_performance.png")


figs = {
    "roles": fig_roles(),
    "span": fig_span(),
    "slit": fig_slit(),
    "perf": fig_performance(),
}
print("figures:", list(figs.keys()))

# ===========================================================================
# Excel 構築
# ===========================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

wb = Workbook()
C_TITLE = "1F4E79"; C_HEAD = "2E75B6"; C_ANS = "E2EFDA"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
f_title = Font(name="MS PGothic", size=15, bold=True, color="FFFFFF")
f_head = Font(name="MS PGothic", size=11, bold=True, color="FFFFFF")
f_body = Font(name="MS PGothic", size=10)
f_ans = Font(name="MS PGothic", size=10, color="375623")
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)


def setup(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False


def title_row(ws, row, text, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = f_title; c.fill = PatternFill("solid", fgColor=C_TITLE)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 30


def head(ws, row, text, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = f_head; c.fill = PatternFill("solid", fgColor=C_HEAD)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 22


def body(ws, row, text, span=8, ans=False, h=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = f_ans if ans else f_body
    c.alignment = wrap
    if ans:
        c.fill = PatternFill("solid", fgColor=C_ANS)
    if h:
        ws.row_dimensions[row].height = h


def table(ws, start_row, headers, rows, col1=1):
    r = start_row
    for j, htxt in enumerate(headers):
        c = ws.cell(r, col1 + j, htxt)
        c.font = Font(name="MS PGothic", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=C_HEAD)
        c.alignment = center; c.border = border
    for data in rows:
        r += 1
        for j, v in enumerate(data):
            c = ws.cell(r, col1 + j, v)
            c.font = f_body
            c.alignment = center if j > 0 else wrap
            c.border = border
            if r % 2 == 0:
                c.fill = PatternFill("solid", fgColor="F2F7FC")
    return r


def put_img(ws, path, anchor, w=None):
    img = XLImage(path)
    if w:
        ratio = w / img.width
        img.width = w; img.height = int(img.height * ratio)
    ws.add_image(img, anchor)


# ---- 目次 ----
ws = wb.active
ws.title = "目次"
setup(ws, [4, 22, 60, 14])
title_row(ws, 1, "RC 造の基礎理解 問題集（構造設計部・新人向け）", span=4)
body(ws, 2, "テーマ No.1：鉄筋コンクリート造を学ぶ最初の一歩。材料の役割、長所短所、"
            "スパン感覚、耐震スリット、要求性能の 5 項目。各シートに図、最後に解答。",
     span=4, h=44)
r = table(ws, 4,
          ["No.", "シート", "到達目標", "図"],
          [["1-1", "No.1-1 材料の役割", "コンクリートと鉄筋の役割を理解している", "梁の曲げ応力"],
           ["1-2", "No.1-2 長所・短所", "RC 造の長所・短所を理解している", "比較表"],
           ["1-3", "No.1-3 スパン・片持ち", "適切なスパン・片持ち出寸法を理解している", "梁せい/片持ち"],
           ["1-4", "No.1-4 耐震スリット", "耐震スリットの目的・要求性能を理解している", "短柱化の防止"],
           ["1-5", "No.1-5 要求性能", "使用性・損傷制御性・安全性を理解している", "3レベル図"]])
body(ws, r + 2, "※ ご提示の項番で No.1-3 が 2 つあったため、耐震スリットを 1-4、"
                "要求性能を 1-5 に整理しています。", span=4, h=32)

# ---- 1-1 ----
ws = wb.create_sheet("No.1-1 材料の役割")
setup(ws, [8, 16, 18, 16, 12, 12, 12, 12])
title_row(ws, 1, "No.1-1  コンクリートと鉄筋の役割")
head(ws, 3, "■ 図 1-1  単純梁における応力分担")
put_img(ws, figs["roles"], "A4", w=760)
head(ws, 26, "■ 問題 1  役割の穴埋め")
body(ws, 27, "(1) コンクリートは【 ① 】に強く【 ② 】に弱い。鉄筋は【 ③ 】に強い。", h=20)
body(ws, 28, "(2) 単純梁に鉛直荷重が載るとき、上側は【 ④ 】・下側は【 ⑤ 】になる。"
             "よって主筋は梁の【 ⑥ 】側に多く配置する。", h=32)
body(ws, 29, "(3) 片持ち梁では引張が【 ⑦ 】側に出るので、主筋は【 ⑧ 】側に配置する。", h=20)
head(ws, 31, "■ 問題 2  なぜ両者を組み合わせるか")
body(ws, 32, "コンクリートと鉄筋が一体で働ける理由を 3 つ挙げよ"
             "（付着／線膨張係数／アルカリ性 の語を使う）。", h=40)
head(ws, 34, "■ 問題 3  かぶり厚さの意味")
body(ws, 35, "鉄筋を覆うコンクリート（かぶり厚さ）の役割を 2 つ挙げよ"
             "（耐久性・耐火 の観点）。また、かぶり不足で起こる代表的な劣化現象を答えよ。",
     h=44)
head(ws, 37, "■ 問題 4  記号の対応")
r = table(ws, 38, ["部材", "主な役割", "対応(記入)"],
          [["主筋（軸方向鉄筋）", "曲げによる引張力を負担", ""],
           ["あばら筋・帯筋（せん断補強筋）", "せん断力負担・主筋座屈拘束・じん性", ""],
           ["コンクリート", "圧縮力負担・鉄筋保護", ""]])

# ---- 1-2 ----
ws = wb.create_sheet("No.1-2 長所・短所")
setup(ws, [8, 20, 26, 16, 12, 12, 12])
title_row(ws, 1, "No.1-2  RC 造の長所・短所")
head(ws, 3, "■ 問題 1  長所・短所の分類")
body(ws, 4, "次の特徴を RC 造の【長所】か【短所】に分類し、表に○を付けよ。", h=20)
r = table(ws, 6,
          ["No.", "特徴", "長所/短所(記入)"],
          [["(1)", "耐火性に優れる", ""],
           ["(2)", "自重が重い（地震力が大きくなる）", ""],
           ["(3)", "遮音性・遮熱性が高い", ""],
           ["(4)", "自由な形状をつくれる（型枠次第）", ""],
           ["(5)", "工期が長い（養生期間が必要）", ""],
           ["(6)", "ひび割れが生じやすい", ""],
           ["(7)", "耐久性が高い（適切なかぶりで長寿命）", ""],
           ["(8)", "解体・改修が難しい", ""],
           ["(9)", "剛性が高く揺れにくい（居住性良）", ""],
           ["(10)", "品質が施工・養生に左右される", ""]])
head(ws, r + 2, "■ 問題 2  他構造との比較")
body(ws, r + 3, "RC 造・S 造・木造 を、次の観点で◎○△で相対評価せよ。", h=20)
r = table(ws, r + 5,
          ["観点", "RC造", "S造", "木造"],
          [["耐火性", "", "", ""],
           ["自重（軽さ）", "", "", ""],
           ["大スパン適性", "", "", ""],
           ["工期", "", "", ""],
           ["遮音性", "", "", ""],
           ["剛性（揺れにくさ）", "", "", ""]])
head(ws, r + 2, "■ 問題 3  短所への対策")
body(ws, r + 3, "RC 造の代表的短所「自重が重い」「ひび割れ」「工期が長い」に対し、"
                "実務でとられる対策を各 1 つ挙げよ"
                "（軽量コンクリート/ひび割れ誘発目地/プレキャスト 等）。", h=44)

# ---- 1-3 ----
ws = wb.create_sheet("No.1-3 スパン・片持ち")
setup(ws, [8, 18, 18, 16, 12, 12, 12])
title_row(ws, 1, "No.1-3  適切なスパン・片持ち出寸法")
head(ws, 3, "■ 図 1-3  梁せいとスパン／片持ち梁")
put_img(ws, figs["span"], "A4", w=760)
head(ws, 27, "■ 問題 1  梁せいの目安")
body(ws, 28, "RC 大梁の梁せい D は スパン L の 1/10〜1/12 が目安。次の表を埋めよ。", h=20)
r = table(ws, 30,
          ["スパン L (m)", "D=L/12 (mm)", "D=L/10 (mm)", "採用例 (記入)"],
          [["6", "", "", ""],
           ["8", "", "", ""],
           ["10", "", "", ""],
           ["12", "", "", ""]])
head(ws, r + 2, "■ 問題 2  スパンの考え方")
body(ws, r + 3, "(1) RC ラーメンの一般的な経済スパンは何 m 程度か。また大スパン（12m 超）に"
                "するとどんな問題が出るか 2 つ。", h=40)
body(ws, r + 4, "(2) スラブの厚さの目安（短辺スパンの 1/30〜1/40）から、短辺 6m のスラブ厚を求めよ。",
     h=32)
head(ws, r + 6, "■ 問題 3  片持ち部材")
body(ws, r + 7, "(1) バルコニーなど RC 片持ち梁の出寸法 Lc の一般的な上限は何 m 程度か。", h=20)
body(ws, r + 8, "(2) 片持ち梁で主筋を上側に配する理由を、引張側の観点から説明せよ。", h=20)
body(ws, r + 9, "(3) 出寸法が大きいと顕在化する問題を 2 つ挙げ、対策（先端梁・ハンチ・"
                "プレストレス 等）を 1 つ述べよ。", h=40)
head(ws, r + 11, "■ 問題 4  たわみの感覚")
body(ws, r + 12, "片持ち梁のたわみは出寸法の 4 乗に比例する（δ∝Lc⁴）。"
                 "出を 1.5m から 2.0m にすると、たわみは何倍になるか計算せよ。", h=32)

# ---- 1-4 ----
ws = wb.create_sheet("No.1-4 耐震スリット")
setup(ws, [8, 18, 18, 16, 12, 12, 12])
title_row(ws, 1, "No.1-4  耐震スリットの目的・要求性能")
head(ws, 3, "■ 図 1-4  耐震スリットによる短柱せん断破壊の防止")
put_img(ws, figs["slit"], "A4", w=760)
head(ws, 27, "■ 問題 1  短柱問題の理解")
body(ws, 28, "(1) 腰壁・垂れ壁が柱に取り付くと、柱の有効長さはどうなるか。"
             "それにより柱の剛性・負担せん断力はどう変化するか。", h=40)
body(ws, 29, "(2) 「短柱」が危険とされる理由を、せん断破壊と曲げ破壊の違い"
             "（脆性 vs 靭性）から説明せよ。", h=40)
head(ws, 31, "■ 問題 2  耐震スリットの目的")
body(ws, 32, "耐震スリットを設ける目的を 2 つ挙げよ"
             "（短柱化の防止／非構造壁の応力集中回避）。", h=32)
head(ws, 34, "■ 問題 3  完全スリットと部分スリット")
r = table(ws, 35,
          ["種類", "壁との縁切り", "特徴", "用途(記入)"],
          [["完全スリット", "全周を絶縁", "柱を完全に独立。耐震上は確実", ""],
           ["部分スリット", "一部のみ絶縁", "壁の一部を耐震要素として残す", ""]])
body(ws, r + 2, "上表をふまえ、「壁を耐震要素として積極活用する設計」と"
                "「スリットで柱を守る設計」のどちらを選ぶかは何で決まるか述べよ。", h=40)
head(ws, r + 4, "■ 問題 4  スリットの要求性能")
body(ws, r + 5, "耐震スリット（材料・納まり）に求められる性能を挙げよ。"
                "(a) 地震時の層間変形に追従  (b) 止水・耐火・遮音  (c) 経年劣化 "
                "── それぞれ具体的にどう対応するか。", h=58)

# ---- 1-5 ----
ws = wb.create_sheet("No.1-5 要求性能")
setup(ws, [8, 18, 18, 16, 12, 12, 12])
title_row(ws, 1, "No.1-5  使用性・損傷制御性・安全性の要求性能")
head(ws, 3, "■ 図 1-5  3 つの要求性能と地震レベルの対応")
put_img(ws, figs["perf"], "A4", w=680)
head(ws, 25, "■ 問題 1  3 性能の対応づけ")
body(ws, 26, "次の性能・地震レベル・目標状態を正しく結びつけよ。", h=20)
r = table(ws, 28,
          ["要求性能", "対象の地震", "目標とする状態(記入)"],
          [["使用性", "中小地震（数十年に一度）", ""],
           ["損傷制御性", "中地震（数十〜百年に一度）", ""],
           ["安全性", "大地震（数百年に一度）", ""]])
head(ws, r + 2, "■ 問題 2  一次設計・二次設計との対応")
body(ws, r + 3, "建築基準法の一次設計（許容応力度計算）と二次設計（保有水平耐力等）は、"
                "上の 3 性能のどれに対応するか整理せよ。", h=40)
head(ws, r + 5, "■ 問題 3  性能と設計指標")
r = table(ws, r + 6,
          ["性能", "主な照査指標", "代表的な限界値の例(記入)"],
          [["使用性", "層間変形角・ひび割れ幅・たわみ", ""],
           ["損傷制御性", "部材の降伏・層間変形角", ""],
           ["安全性", "保有水平耐力・靭性（崩壊機構）", ""]])
head(ws, r + 2, "■ 問題 4  なぜ大地震で損傷を許すのか")
body(ws, r + 3, "「大地震時に損傷を許容するが倒壊させない」という考え方が"
                "合理的である理由を、経済性と安全性の両面から説明せよ"
                "（弾性設計でまかなう場合との比較）。", h=58)
head(ws, r + 5, "■ 問題 5  RC 部材を靭性的にする工夫")
body(ws, r + 6, "安全性能（大地震で粘る）を確保するための RC 部材の工夫を 3 つ挙げよ"
                "（せん断補強筋を密に／曲げ降伏先行（柱梁耐力比）／靭性指針の遵守 等）。",
     h=44)

# ---- 解答 ----
ws = wb.create_sheet("解答")
setup(ws, [4, 22, 60, 14])
title_row(ws, 1, "解答・解説", span=4)
row = 2


def ah(t):
    global row
    head(ws, row, t, span=4); row += 1


def an(t, h=None):
    global row
    body(ws, row, t, span=4, ans=True, h=h); row += 1


ah("No.1-1  材料の役割")
an("問1：①圧縮 ②引張 ③引張 ④圧縮 ⑤引張 ⑥下 ⑦上 ⑧上。"
   "（曲げで引張が出る側に主筋を置くのが大原則）", h=40)
an("問2：①付着＝鉄筋表面の節とコンクリートが噛み合い一体で力を伝達。"
   "②線膨張係数がほぼ等しい(≒1.0×10⁻⁵)ため温度変化で剥離しない。"
   "③コンクリートのアルカリ性が鉄筋表面に不動態被膜を作り錆を防ぐ。", h=58)
an("問3：①耐久性＝鉄筋の錆（中性化・塩害）を防ぐ。②耐火＝火災時に鉄筋温度上昇を遅らせ"
   "強度低下を防ぐ。かぶり不足→中性化が早く進み鉄筋が錆びて膨張し"
   "『かぶりコンクリートの剥落（爆裂）』が起こる。", h=58)
an("問4：主筋＝曲げ引張、あばら筋/帯筋＝せん断・座屈拘束・靭性、コンクリート＝圧縮・鉄筋保護。",
   h=32)

ah("No.1-2  長所・短所")
an("問1：長所＝(1)(3)(4)(7)(9)、短所＝(2)(5)(6)(8)(10)。"
   "耐火・遮音・自由形状・耐久・剛性は長所、重量・工期・ひび割れ・解体困難・品質ばらつきは短所。",
   h=58)
an("問2(例)：耐火＝RC◎/S△(被覆要)/木△。軽さ＝RC△/S○/木◎。"
   "大スパン＝RC△/S◎/木△。工期＝RC△/S○/木◎。遮音＝RC◎/S△/木△。"
   "剛性＝RC◎/S○/木△。", h=58)
an("問3：自重→軽量コンクリート・中空スラブ・部材断面の最適化。"
   "ひび割れ→ひび割れ誘発目地・収縮低減剤・適切な養生。"
   "工期→プレキャスト(PCa)化・ハーフPCa・型枠の合理化。", h=58)

ah("No.1-3  スパン・片持ち")
an("問1：D=L/12 と L/10：L6m→500/600、L8m→667/800、L10m→833/1000、L12m→1000/1200(mm)。",
   h=32)
an("問2：(1)RCラーメンの経済スパンは 6〜8m 程度。大スパン化→梁せいが過大になり階高増・"
   "自重増、たわみ・ひび割れ制御が困難（PC やS梁の検討要）。"
   "(2)短辺6m → 厚さ6000/40〜6000/30 = 150〜200mm 程度。", h=58)
an("問3：(1)片持ち出 Lc は概ね 2m 程度が目安(バルコニー)。"
   "(2)片持ちは固定端上側が引張になるため主筋を上側に配する。"
   "(3)問題＝たわみ大・ひび割れ・振動。対策＝先端に小梁、根元ハンチ、PC 導入、出寸法の抑制。",
   h=58)
an("問4：δ∝Lc⁴。(2.0/1.5)⁴ = (1.333)⁴ ≒ 3.16 倍。出を1.33倍にするとたわみは約3.2倍。",
   h=32)

ah("No.1-4  耐震スリット")
an("問1：(1)腰壁・垂れ壁が付くと柱の自由に変形できる長さ(内法高さ)が短くなる(短柱化)。"
   "剛性は長さの3乗に反比例で急増し、地震力(せん断力)が短柱に集中する。"
   "(2)短柱はせん断破壊しやすい。せん断破壊は予兆なく急激に耐力を失う脆性破壊で、"
   "曲げ降伏(靭性的でねばる)より危険。1968十勝沖・1995阪神で多数被害。", h=72)
an("問2：①腰壁・垂れ壁による柱の短柱化を防ぎ、柱を本来の長さで曲げ降伏させる。"
   "②非構造の雑壁が地震時に構造体へ予期せぬ応力集中・損傷を与えるのを防ぐ。", h=44)
an("問3：完全スリット→柱を守りたい/雑壁を構造に見込まない設計。"
   "部分スリット→壁の一部を耐震要素に使いつつ短柱化は避けたい設計。"
   "選択は『その壁を構造耐力に算入するか否か(構造計画)』で決まる。", h=58)
an("問4：(a)層間変形追従＝スリット幅を想定変形以上に確保し、変形を吸収できる材料"
   "(発泡材＋シーリング)。(b)止水・耐火・遮音＝耐火目地材・バックアップ材・シーリングで確保。"
   "(c)経年劣化＝シーリングの打替えメンテ、紫外線・水を考慮した材料選定。", h=72)

ah("No.1-5  要求性能")
an("問1：使用性→中小地震→無損傷で継続使用(弾性)。損傷制御性→中地震→補修可能な軽微損傷。"
   "安全性→大地震→倒壊せず人命保護。", h=44)
an("問2：一次設計(許容応力度計算)＝中小地震で部材を許容応力度内に収める→使用性・損傷制御性。"
   "二次設計(保有水平耐力・層間変形・剛性率偏心率)＝大地震で倒壊させない→安全性。", h=44)
an("問3(例)：使用性＝層間変形角 1/200 以下、ひび割れ幅 0.3mm 程度以下。"
   "損傷制御性＝主要部材が降伏しない〜軽微、層間変形角 1/100 程度。"
   "安全性＝保有水平耐力 ≥ 必要保有水平耐力、靭性ある崩壊機構(全体降伏型)。", h=58)
an("問4：大地震を全て弾性で受けると部材が過大・非経済。"
   "発生頻度が極めて低い大地震に対しては『損傷を許容して靭性で倒壊を防ぐ』方が"
   "経済合理的かつ人命安全を確保できる。これが性能設計(2段階設計)の基本思想。", h=58)
an("問5：①せん断補強筋(帯筋・あばら筋)を密に配しせん断破壊を防ぐ。"
   "②柱梁耐力比>1で梁先行降伏(全体崩壊型)に誘導し層崩壊を防ぐ。"
   "③靭性保証型設計指針に従い、曲げ降伏がせん断耐力に先行するよう設計。", h=58)

XLSX = os.path.join(OUT, "RC造基礎問題集.xlsx")
wb.save(XLSX)
print("saved:", XLSX)
