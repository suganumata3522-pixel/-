# -*- coding: utf-8 -*-
"""バランス・耐震要素配置 問題集（図つき）Excel 生成スクリプト。
出力: docs/balance_layout/バランス配置問題集.xlsx
"""
import os
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib import font_manager as fm

FONT_PATH = "/usr/share/fonts/truetype/fonts-japanese-gothic.ttf"
jp = fm.FontProperties(fname=FONT_PATH)
fm.fontManager.addfont(FONT_PATH)
plt.rcParams["font.family"] = jp.get_name()
plt.rcParams["axes.unicode_minus"] = False

OUT = "docs/balance_layout"
FIG = os.path.join(OUT, "figures")
os.makedirs(FIG, exist_ok=True)

# 共通プラン: 4スパン × 3スパン (X=6m×4=24m, Y=7m×3=21m)
SPX = [6, 6, 6, 6]
SPY = [7, 7, 7]
XL = np.cumsum([0] + SPX)
YL = np.cumsum([0] + SPY)
W, H = sum(SPX), sum(SPY)
NX, NY = len(XL), len(YL)  # 5, 4


def save(fig, name):
    p = os.path.join(FIG, name)
    fig.savefig(p, dpi=130, bbox_inches="tight")
    plt.close(fig)
    return p


def add_wall(ax, x1, y1, x2, y2, color="#c4858f"):
    """壁を描画。y1==y2 ならX方向壁、x1==x2 ならY方向壁。"""
    ww = 0.5
    if y1 == y2:
        ax.add_patch(mpatches.Rectangle((min(x1, x2), y1 - ww / 2),
                                         abs(x2 - x1), ww,
                                         fc=color, ec="k", lw=0.7, hatch="//",
                                         alpha=0.9, zorder=4))
    else:
        ax.add_patch(mpatches.Rectangle((x1 - ww / 2, min(y1, y2)),
                                         ww, abs(y2 - y1),
                                         fc=color, ec="k", lw=0.7, hatch="//",
                                         alpha=0.9, zorder=4))


def draw_plan(ax, title, walls=None, cg=None, cs=None, *,
              spx=None, spy=None, show_axes=True, note=None):
    """汎用平面図描画関数"""
    spx = spx or SPX
    spy = spy or SPY
    xl = np.cumsum([0] + list(spx))
    yl = np.cumsum([0] + list(spy))
    w, h = sum(spx), sum(spy)
    # グリッド
    for x in xl:
        ax.plot([x, x], [-0.3, h + 0.3], color="lightgray", lw=0.5, ls=":", zorder=1)
    for y in yl:
        ax.plot([-0.3, w + 0.3], [y, y], color="lightgray", lw=0.5, ls=":", zorder=1)
    # 外形
    ax.add_patch(mpatches.Rectangle((0, 0), w, h, fc="none", ec="k", lw=1.0))
    # 軸線
    if show_axes:
        for i, x in enumerate(xl):
            ax.text(x, -1.2, f"X{i + 1}", fontproperties=jp, ha="center", va="center",
                    fontsize=7, bbox=dict(boxstyle="circle,pad=0.10", fc="white",
                                          ec="k", lw=0.5))
        for i, y in enumerate(yl):
            ax.text(-1.4, y, f"Y{i + 1}", fontproperties=jp, ha="center", va="center",
                    fontsize=7, bbox=dict(boxstyle="circle,pad=0.10", fc="white",
                                          ec="k", lw=0.5))
    # 梁
    for iy, y in enumerate(yl):
        for ix in range(len(xl) - 1):
            ax.plot([xl[ix] + 0.3, xl[ix + 1] - 0.3], [y, y],
                    color="#1f4e79", lw=1.2, zorder=2)
    for ix, x in enumerate(xl):
        for iy in range(len(yl) - 1):
            ax.plot([x, x], [yl[iy] + 0.3, yl[iy + 1] - 0.3],
                    color="#1f4e79", lw=1.2, zorder=2)
    # 壁
    if walls:
        for ww_ in walls:
            add_wall(ax, *ww_)
    # 柱
    cw = 0.55
    for x in xl:
        for y in yl:
            ax.add_patch(mpatches.Rectangle((x - cw / 2, y - cw / 2), cw, cw,
                                             fc="#bdbdbd", ec="k", lw=0.5, zorder=5))
    # G・R
    if cg:
        ax.plot(*cg, "o", color="#1f7a1f", markersize=11, zorder=10)
        ax.text(cg[0] + 0.4, cg[1] - 1.0, "G", fontproperties=jp,
                color="#1f7a1f", fontsize=10, fontweight="bold", zorder=11)
    if cs:
        ax.plot(*cs, "*", color="#c00000", markersize=18, zorder=10)
        ax.text(cs[0] + 0.4, cs[1] + 0.9, "R", fontproperties=jp,
                color="#c00000", fontsize=10, fontweight="bold", zorder=11)
    if cg and cs and (cg[0] != cs[0] or cg[1] != cs[1]):
        ax.plot([cg[0], cs[0]], [cg[1], cs[1]],
                color="#c00000", lw=1.0, ls="--", zorder=8)
    if note:
        ax.text(w / 2, -2.6, note, fontproperties=jp, ha="center",
                fontsize=9, color="#444")
    ax.set_xlim(-2.2, w + 0.8)
    ax.set_ylim(-3.2, h + 0.8)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title(title, fontproperties=jp, fontsize=10, fontweight="bold")


# ========== 図 10-1: 4 プラン比較 ==========
def fig_10_1():
    fig, axes = plt.subplots(2, 2, figsize=(13, 11))
    # (a) NG 偏在
    walls_a = [(0, 0, 0, 7), (0, 0, 6, 0)]
    draw_plan(axes[0, 0], "(a) NG  壁が左下に偏在 ── ねじれ振動大",
              walls=walls_a, cg=(12, 10.5), cs=(2, 1.8),
              note="壁が片寄り：剛心 R が重心 G から大きく離れる")
    # (b) NG 一方向偏在
    walls_b = [(0, 7, 0, 14), (24, 7, 24, 14), (12, 7, 12, 14)]
    draw_plan(axes[0, 1], "(b) NG  Y 方向壁のみ ── X 方向地震に弱い",
              walls=walls_b, cg=(12, 10.5), cs=(12, 10.5),
              note="偏心は無いが、Y 方向地震に抵抗する X 方向壁が皆無")
    # (c) GOOD 外周対称
    walls_c = [(0, 0, 0, 7), (0, 14, 0, 21),
               (24, 0, 24, 7), (24, 14, 24, 21),
               (6, 0, 12, 0), (12, 21, 18, 21)]
    draw_plan(axes[1, 0], "(c) GOOD  外周両側にバランス配置",
              walls=walls_c, cg=(12, 10.5), cs=(12, 10.5),
              note="両方向に壁を均等配置：剛心 R = 重心 G、壁量も充足")
    # (d) 中央コア
    walls_d = [(6, 7, 6, 14), (18, 7, 18, 14),
               (6, 7, 18, 7), (6, 14, 18, 14)]
    draw_plan(axes[1, 1], "(d) 注意  中央コア集中 ── 偏心無しだが捻れに弱い",
              walls=walls_d, cg=(12, 10.5), cs=(12, 10.5),
              note="弾力半径 re が小さい → 偏心率は小さく見えても捻れ剛性不足")
    fig.suptitle("図 10-1  耐震要素の配置バランス  ── 4プラン比較",
                 fontproperties=jp, fontsize=13, fontweight="bold")
    return save(fig, "fig10-1_balance.png")


# ========== 図 10-2: 偏心率改善 ==========
def fig_10_2():
    fig, axes = plt.subplots(1, 2, figsize=(14, 6))
    # 改善前
    walls_b = [(0, 0, 0, 21), (6, 0, 6, 14)]
    # ΣKy = 21+14 = 35,  剛心X = (21·0 + 14·6)/35 = 84/35 = 2.4
    draw_plan(axes[0], "(a) 改善前  ── 偏心率 Re_X ≒ 0.45 (NG)",
              walls=walls_b, cg=(12, 10.5), cs=(2.4, 10.5),
              note="重心G(12, 10.5) と 剛心R(2.4, 10.5)、偏心 ex=9.6m")
    # 改善後
    walls_a = [(0, 0, 0, 21), (6, 0, 6, 14),
               (18, 7, 18, 21), (24, 0, 24, 14)]
    # 剛心X = (21·0 + 14·6 + 14·18 + 14·24)/63 = 672/63 = 10.67
    draw_plan(axes[1], "(b) 改善後  ── 偏心率 Re_X ≒ 0.06 (OK)",
              walls=walls_a, cg=(12, 10.5), cs=(10.67, 10.5),
              note="右側に壁を追加：剛心 R が重心 G に接近、偏心 ex=1.3m")
    fig.suptitle("図 10-2  偏心率の改善  ── 壁追加による剛心の中心化",
                 fontproperties=jp, fontsize=13, fontweight="bold")
    return save(fig, "fig10-2_eccentricity.png")


# ========== 図 10-3: 剛性率 ==========
def fig_10_3():
    fig, axes = plt.subplots(1, 3, figsize=(16, 6))
    spx_e = [6, 6, 6]
    xl_e = np.cumsum([0] + spx_e)
    W_e = sum(spx_e)
    stories = 4
    story_h = 4.0
    HT = stories * story_h

    def draw_elev(ax, walls_per_floor, title):
        for x in xl_e:
            ax.plot([x, x], [0, HT], color="lightgray", lw=0.4, ls=":")
        for s in range(stories + 1):
            ax.plot([0, W_e], [s * story_h, s * story_h], color="lightgray",
                    lw=0.4, ls=":")
        cw, bh = 0.35, 0.25
        for x in xl_e:
            ax.add_patch(mpatches.Rectangle((x - cw / 2, 0), cw, HT,
                                             fc="#bdbdbd", ec="k", lw=0.6, zorder=2))
        for s in range(1, stories + 1):
            for ix in range(len(xl_e) - 1):
                ax.add_patch(mpatches.Rectangle((xl_e[ix] + cw / 2,
                                                  s * story_h - bh / 2),
                                                 xl_e[ix + 1] - xl_e[ix] - cw, bh,
                                                 fc="#666", ec="k", lw=0.4, zorder=2))
        for s, span_ids in enumerate(walls_per_floor):
            for ix in span_ids:
                ax.add_patch(mpatches.Rectangle((xl_e[ix] + cw / 2,
                                                  s * story_h + bh / 2),
                                                 xl_e[ix + 1] - xl_e[ix] - cw,
                                                 story_h - bh,
                                                 fc="#c4858f", ec="k", lw=0.4,
                                                 hatch="//", alpha=0.85, zorder=3))
        for s in range(stories + 1):
            ax.text(-0.8, s * story_h, f"{s + 1}FL" if s < stories else "RF",
                    fontproperties=jp, fontsize=8, va="center", ha="center")
        ax.set_xlim(-1.5, W_e + 1)
        ax.set_ylim(-1, HT + 1)
        ax.set_aspect("equal")
        ax.axis("off")
        ax.set_title(title, fontproperties=jp, fontsize=10, fontweight="bold")

    walls_before = [[], [0, 2], [0, 2], [0, 2]]
    walls_after = [[0, 2], [0, 2], [0, 2], [0, 2]]
    draw_elev(axes[0], walls_before, "(a) 改善前  1階ピロティ駐車場")
    draw_elev(axes[1], walls_after, "(b) 改善後  1階にも耐震壁追加")

    ax = axes[2]
    floors = ["1F", "2F", "3F", "4F"]
    R_before = [40, 100, 95, 90]
    R_after = [95, 100, 95, 90]
    rs_b = [r / np.mean(R_before) for r in R_before]
    rs_a = [r / np.mean(R_after) for r in R_after]
    x = np.arange(len(floors))
    bw = 0.35
    ax.barh(x - bw / 2, rs_b, bw, label="改善前", color="#c4858f", ec="k", lw=0.5)
    ax.barh(x + bw / 2, rs_a, bw, label="改善後", color="#a8dadc", ec="k", lw=0.5)
    ax.axvline(0.6, color="red", ls="--", lw=1.5, label="基準 Rs>=0.6")
    ax.set_yticks(x)
    ax.set_yticklabels(floors, fontproperties=jp)
    ax.invert_yaxis()
    ax.set_xlabel("剛性率 Rs", fontproperties=jp)
    ax.set_title("(c) 各階剛性率の比較", fontproperties=jp, fontsize=10,
                  fontweight="bold")
    ax.legend(prop=jp, fontsize=8, loc="lower right")
    ax.grid(axis="x", alpha=0.3)
    ax.set_xlim(0, 1.4)
    for i, v in enumerate(rs_b):
        ax.text(v + 0.02, i - bw / 2, f"{v:.2f}", fontsize=7, va="center")
    for i, v in enumerate(rs_a):
        ax.text(v + 0.02, i + bw / 2, f"{v:.2f}", fontsize=7, va="center")

    fig.suptitle("図 10-3  剛性率の改善  ── ピロティ問題と立面バランス",
                 fontproperties=jp, fontsize=13, fontweight="bold")
    return save(fig, "fig10-3_stiffness.png")


# ========== 図 10-4: RC 共同住宅 平面 ==========
def fig_10_4():
    fig, ax = plt.subplots(figsize=(11, 5.5))
    spx_m = [5.5, 5.5, 5.5, 5.5, 5.5]
    spy_m = [3, 8, 2]  # バルコニー / 住戸 / 廊下
    xl_m = np.cumsum([0] + spx_m)
    yl_m = np.cumsum([0] + spy_m)
    W_m, H_m = sum(spx_m), sum(spy_m)
    for x in xl_m:
        ax.plot([x, x], [0, H_m], color="lightgray", lw=0.4, ls=":")
    for y in yl_m:
        ax.plot([0, W_m], [y, y], color="lightgray", lw=0.4, ls=":")
    ax.add_patch(mpatches.Rectangle((0, 0), W_m, H_m, fc="none", ec="k", lw=1.2))
    # ゾーン色分け
    ax.fill_between([0, W_m], yl_m[2], yl_m[3], color="#f0f0f0", alpha=0.5, zorder=0)
    ax.text(W_m / 2, yl_m[2] + 1.0, "共用廊下", fontproperties=jp,
            ha="center", fontsize=9, color="gray")
    ax.fill_between([0, W_m], yl_m[0], yl_m[1], color="#fff5e0", alpha=0.5, zorder=0)
    ax.text(W_m / 2, yl_m[0] + 1.5, "バルコニー", fontproperties=jp,
            ha="center", fontsize=9, color="gray")
    # 戸境壁(Y方向壁、5箇所)
    for x in xl_m:
        add_wall(ax, x, yl_m[1], x, yl_m[2])
    # X方向耐震壁(妻側端部)
    add_wall(ax, 0, yl_m[2], 5.5, yl_m[2])
    add_wall(ax, W_m - 5.5, yl_m[2], W_m, yl_m[2])
    # 柱
    cw = 0.5
    for x in xl_m:
        for y in yl_m:
            ax.add_patch(mpatches.Rectangle((x - cw / 2, y - cw / 2), cw, cw,
                                             fc="#bdbdbd", ec="k", lw=0.6, zorder=5))
    # 住戸ラベル
    for i in range(5):
        cx = (xl_m[i] + xl_m[i + 1]) / 2
        ax.text(cx, (yl_m[1] + yl_m[2]) / 2, f"住戸{i + 1}",
                fontproperties=jp, ha="center", va="center", fontsize=9, color="#444")
    # 矢印
    ax.annotate("Y方向耐震壁\n(戸境壁を活用)",
                xy=(xl_m[1], (yl_m[1] + yl_m[2]) / 2),
                xytext=(xl_m[1] + 1.5, yl_m[1] - 2.5),
                fontproperties=jp, fontsize=9, color="#7a3b3b",
                arrowprops=dict(arrowstyle="->", color="#7a3b3b"))
    ax.annotate("X方向耐震壁\n(妻側両端のみ)",
                xy=(2.5, yl_m[2]), xytext=(0, H_m + 0.5),
                fontproperties=jp, fontsize=9, color="#7a3b3b",
                arrowprops=dict(arrowstyle="->", color="#7a3b3b"))
    ax.text(W_m / 2, -2.5, "マンションは Y 方向 (短手) は戸境壁が並ぶため壁量充足／"
                            "X 方向 (長手) は妻壁とコア (EV シャフト等) で確保",
            fontproperties=jp, ha="center", fontsize=8, color="#444")
    ax.set_xlim(-1, W_m + 1)
    ax.set_ylim(-3.5, H_m + 2)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("図 10-4  RC 造共同住宅 (5戸×1階)  耐震壁配置例",
                 fontproperties=jp, fontsize=11, fontweight="bold")
    return save(fig, "fig10-4_mansion.png")


# ========== 図 10-5: S 造工場 平面 + 軸組 ==========
def fig_10_5():
    fig, axes = plt.subplots(1, 2, figsize=(14, 5.5))
    spx_f = [10, 10, 10, 10, 10]
    spy_f = [6, 6, 6, 6, 6]
    xl_f = np.cumsum([0] + spx_f)
    yl_f = np.cumsum([0] + spy_f)
    W_f, H_f = sum(spx_f), sum(spy_f)

    ax = axes[0]
    for x in xl_f:
        ax.plot([x, x], [0, H_f], color="lightgray", lw=0.4, ls=":")
    for y in yl_f:
        ax.plot([0, W_f], [y, y], color="lightgray", lw=0.4, ls=":")
    ax.add_patch(mpatches.Rectangle((0, 0), W_f, H_f, fc="none", ec="k", lw=1.0))
    # シャッター (両妻 中央)
    sh_y1, sh_y2 = yl_f[1], yl_f[4]
    ax.fill_between([-0.6, 0.4], sh_y1, sh_y2, color="#ffdc8e", alpha=0.7)
    ax.text(-1.2, (sh_y1 + sh_y2) / 2, "シャッター", fontproperties=jp,
            ha="center", rotation=90, fontsize=8, color="#7a5a00", va="center")
    ax.fill_between([W_f - 0.4, W_f + 0.6], sh_y1, sh_y2, color="#ffdc8e", alpha=0.7)
    # Y方向ブレース構面 (X1, X3, X6通り) → X方向地震抵抗
    for ix in [0, 2, 5]:
        ax.add_patch(mpatches.Rectangle((xl_f[ix] - 0.4, yl_f[0] - 0.3),
                                         0.8, H_f + 0.6,
                                         fc="none", ec="#c00000", lw=1.4,
                                         ls=(0, (4, 2)), zorder=3))
    ax.text(W_f / 2, H_f + 1.5, "← 赤枠 = Y方向ブレース構面 (X方向地震に抵抗)",
            fontproperties=jp, ha="center", fontsize=9, color="#c00000")
    # X方向ブレース構面 (Y1, Y6通り) → Y方向地震抵抗
    for iy in [0, 5]:
        ax.add_patch(mpatches.Rectangle((xl_f[0] - 0.3, yl_f[iy] - 0.4),
                                         W_f + 0.6, 0.8,
                                         fc="none", ec="#1f7a1f", lw=1.4,
                                         ls=(0, (4, 2)), zorder=3))
    ax.text(W_f / 2, -2.4, "緑枠 = X方向ブレース構面 (Y方向地震に抵抗)",
            fontproperties=jp, ha="center", fontsize=9, color="#1f7a1f")
    cw = 0.6
    for x in xl_f:
        for y in yl_f:
            ax.add_patch(mpatches.Rectangle((x - cw / 2, y - cw / 2), cw, cw,
                                             fc="#bdbdbd", ec="k", lw=0.5))
    for i, s in enumerate(spx_f):
        ax.text((xl_f[i] + xl_f[i + 1]) / 2, -1.6, f"{s * 1000:,}",
                fontproperties=jp, ha="center", fontsize=7, color="gray")
    ax.set_xlim(-3, W_f + 5)
    ax.set_ylim(-3, H_f + 3)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(a) S 造工場 (50m × 30m) 平面 ── ブレース構面配置",
                 fontproperties=jp, fontsize=10, fontweight="bold")

    # 軸組
    ax2 = axes[1]
    stories = 2
    story_h = 5
    HT = stories * story_h
    cw2, bh2 = 0.5, 0.4
    for x in xl_f:
        ax2.add_patch(mpatches.Rectangle((x - cw2 / 2, 0), cw2, HT,
                                          fc="#bdbdbd", ec="k", lw=0.6))
    for s in range(1, stories + 1):
        for ix in range(len(xl_f) - 1):
            ax2.add_patch(mpatches.Rectangle((xl_f[ix] + cw2 / 2,
                                               s * story_h - bh2 / 2),
                                              xl_f[ix + 1] - xl_f[ix] - cw2, bh2,
                                              fc="#666", ec="k", lw=0.5))
    # 両端スパンに X 形ブレース
    for s in range(stories):
        for ix in [0, 4]:
            xL = xl_f[ix] + cw2 / 2
            xR = xl_f[ix + 1] - cw2 / 2
            yB = s * story_h + bh2 / 2
            yT = (s + 1) * story_h - bh2 / 2
            ax2.plot([xL, xR], [yB, yT], color="#c00000", lw=2.2)
            ax2.plot([xL, xR], [yT, yB], color="#c00000", lw=2.2)
    for s in range(stories + 1):
        ax2.text(-2, s * story_h, f"{s + 1}FL" if s < stories else "RF",
                 fontproperties=jp, fontsize=8, va="center", ha="center")
    ax2.set_xlim(-3.5, W_f + 1)
    ax2.set_ylim(-1, HT + 1.5)
    ax2.set_aspect("equal")
    ax2.axis("off")
    ax2.set_title("(b) Y1 妻面 軸組図 ── 両端スパンに X 形ブレース",
                  fontproperties=jp, fontsize=10, fontweight="bold")
    fig.suptitle("図 10-5  S 造工場 ブレース配置計画",
                 fontproperties=jp, fontsize=13, fontweight="bold")
    return save(fig, "fig10-5_factory.png")


# ========== 図 10-6: 建築要望のケース ==========
def fig_10_6():
    fig, axes = plt.subplots(1, 2, figsize=(13, 5.5))
    spx = [6, 6, 6]
    xl = np.cumsum([0] + spx)
    W_e = sum(spx)
    stories = 4
    story_h = 3.5
    HT = stories * story_h
    cw, bh = 0.3, 0.2

    def base(ax):
        for x in xl:
            ax.add_patch(mpatches.Rectangle((x - cw / 2, 0), cw, HT,
                                             fc="#bdbdbd", ec="k", lw=0.5))
        for s in range(1, stories + 1):
            for ix in range(len(xl) - 1):
                ax.add_patch(mpatches.Rectangle((xl[ix] + cw / 2,
                                                  s * story_h - bh / 2),
                                                 xl[ix + 1] - xl[ix] - cw, bh,
                                                 fc="#666", ec="k", lw=0.4))
        for s in range(stories + 1):
            ax.text(-0.8, s * story_h, f"{s + 1}FL" if s < stories else "RF",
                    fontproperties=jp, fontsize=7, va="center", ha="center")

    ax = axes[0]
    base(ax)
    for s in range(1, stories):
        for ix in [0, 2]:
            ax.add_patch(mpatches.Rectangle((xl[ix] + cw / 2, s * story_h + bh / 2),
                                             xl[ix + 1] - xl[ix] - cw, story_h - bh,
                                             fc="#c4858f", ec="k", lw=0.4, hatch="//",
                                             alpha=0.85, zorder=3))
    ax.text(W_e / 2, story_h / 2, "1階 = 駐車場\n壁なし",
            fontproperties=jp, ha="center", fontsize=9, color="#c00000",
            fontweight="bold")
    ax.annotate("剛性ジャンプ\n(Rs < 0.6 で NG)",
                xy=(W_e + 0.3, story_h * 0.9),
                xytext=(W_e + 1.2, story_h * 1.6),
                fontproperties=jp, fontsize=8, color="#c00000",
                arrowprops=dict(arrowstyle="->", color="#c00000"))
    ax.set_xlim(-1.5, W_e + 3.5)
    ax.set_ylim(-1, HT + 1)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("(a) 要望「1階を駐車場に」── ピロティ問題",
                 fontproperties=jp, fontsize=10, fontweight="bold")

    ax2 = axes[1]
    base(ax2)
    for s in range(1, stories):
        for ix in [0, 2]:
            ax2.add_patch(mpatches.Rectangle((xl[ix] + cw / 2, s * story_h + bh / 2),
                                              xl[ix + 1] - xl[ix] - cw, story_h - bh,
                                              fc="#c4858f", ec="k", lw=0.4, hatch="//",
                                              alpha=0.85, zorder=3))
    # 1階にブレース
    for ix in [0, 2]:
        xL = xl[ix] + cw / 2
        xR = xl[ix + 1] - cw / 2
        yB = bh / 2 + 0.1
        yT = story_h - bh / 2 - 0.1
        ax2.plot([xL, xR], [yB, yT], color="#c00000", lw=2)
        ax2.plot([xL, xR], [yT, yB], color="#c00000", lw=2)
    ax2.text(W_e / 2, story_h / 2, "1階 = 大開口\n+ 鉄骨ブレース",
             fontproperties=jp, ha="center", fontsize=9, color="#1f7a1f",
             fontweight="bold")
    ax2.set_xlim(-1.5, W_e + 1.5)
    ax2.set_ylim(-1, HT + 1)
    ax2.set_aspect("equal")
    ax2.axis("off")
    ax2.set_title("(b) 代替案「1階に鉄骨ブレース併用」── 大開口維持＋剛性確保",
                  fontproperties=jp, fontsize=10, fontweight="bold")
    fig.suptitle("図 10-6  建築要望と構造的対応 ── 代替案の提案",
                 fontproperties=jp, fontsize=13, fontweight="bold")
    return save(fig, "fig10-6_request.png")


figs = {
    "f1": fig_10_1(),
    "f2": fig_10_2(),
    "f3": fig_10_3(),
    "f4": fig_10_4(),
    "f5": fig_10_5(),
    "f6": fig_10_6(),
}
print("figures:", list(figs.keys()))

# ===========================================================================
# Excel 構築
# ===========================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

wb = Workbook()
C_TITLE = "1F4E79"; C_HEAD = "2E75B6"; C_ANS = "E2EFDA"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

f_title = Font(name="MS PGothic", size=15, bold=True, color="FFFFFF")
f_head = Font(name="MS PGothic", size=11, bold=True, color="FFFFFF")
f_body = Font(name="MS PGothic", size=10)
f_ans = Font(name="MS PGothic", size=10, color="375623")
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)


def setup(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False


def title_row(ws, row, text, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = f_title
    c.fill = PatternFill("solid", fgColor=C_TITLE)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 30


def head(ws, row, text, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = f_head
    c.fill = PatternFill("solid", fgColor=C_HEAD)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 22


def body(ws, row, text, span=8, ans=False, h=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = f_ans if ans else f_body
    c.alignment = wrap
    if ans:
        c.fill = PatternFill("solid", fgColor=C_ANS)
    if h:
        ws.row_dimensions[row].height = h


def table(ws, start_row, headers, rows, col1=1):
    r = start_row
    for j, htxt in enumerate(headers):
        c = ws.cell(r, col1 + j, htxt)
        c.font = Font(name="MS PGothic", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=C_HEAD)
        c.alignment = center
        c.border = border
    for data in rows:
        r += 1
        for j, v in enumerate(data):
            c = ws.cell(r, col1 + j, v)
            c.font = f_body
            c.alignment = center if j > 0 else wrap
            c.border = border
            if r % 2 == 0:
                c.fill = PatternFill("solid", fgColor="F2F7FC")
    return r


def put_img(ws, path, anchor, w=None):
    img = XLImage(path)
    if w:
        ratio = w / img.width
        img.width = w
        img.height = int(img.height * ratio)
    ws.add_image(img, anchor)


# ---- 目次 ----
ws = wb.active
ws.title = "目次"
setup(ws, [4, 22, 60, 14])
title_row(ws, 1, "バランス配置 問題集（ゼネコン構造設計部・新人向け）", span=4)
body(ws, 2, "テーマ No.10：耐震要素（壁・ブレース）の平面・立面バランス配置。"
            "偏心率 Re ≤ 0.15、剛性率 Rs ≥ 0.6 を満たす配置と、用途別の典型計画。",
     span=4, h=44)
r = table(ws, 4,
          ["No.", "シート", "到達目標", "主な図"],
          [["10-1", "No.10-1 配置バランス",
            "平面的・立面的にバランスに配慮した配置ができる", "4プラン比較"],
           ["10-2", "No.10-2 偏心率",
            "偏心率の改善に向けた平面配置検討ができる", "改善前後の重心剛心"],
           ["10-3", "No.10-3 剛性率",
            "剛性率の改善に向けた鉛直配置検討ができる", "立面＋剛性棒グラフ"],
           ["10-4", "No.10-4 RC マンション",
            "RC 造共同住宅の耐震壁配置計画ができる", "マンション平面"],
           ["10-5", "No.10-5 S 造工場",
            "S 造工場・倉庫のブレース配置計画ができる", "工場平面＋軸組"],
           ["10-6", "No.10-6 建築要望",
            "建築要望に配慮した配置計画ができる", "ピロティ＋代替案"]])
body(ws, r + 2,
     "凡例：G=重心(緑●)、R=剛心(赤★)、耐震壁=ピンク斜線、ブレース=赤線。"
     "基準法的に Re ≤ 0.15／Rs ≥ 0.6 が標準（超過するとFe・Fs 割増し）。",
     span=4, h=44)

# ---- 10-1 ----
ws = wb.create_sheet("No.10-1 配置バランス")
setup(ws, [8, 16, 18, 16, 12, 12, 12, 12])
title_row(ws, 1, "No.10-1  平面的・立面的にバランスに配慮した配置")
head(ws, 3, "■ 図 10-1  耐震要素の配置バランス（4プラン比較）")
put_img(ws, figs["f1"], "A4", w=720)
head(ws, 36, "■ 問題 1  各プランの良し悪し判定")
body(ws, 37, "下表の各プランが GOOD か NG か、その理由を 1 行で記入せよ。", h=22)
r = table(ws, 39,
          ["プラン", "概要", "G/NG", "理由 (記入)"],
          [["(a)", "壁が左下に偏在", "", ""],
           ["(b)", "Y 方向壁のみ", "", ""],
           ["(c)", "外周両側にバランス配置", "", ""],
           ["(d)", "中央コア集中", "", ""]])
head(ws, r + 2, "■ 問題 2  平面バランスのチェック項目")
body(ws, r + 3, "良い平面配置の条件を、次のキーワードを使って 5 項目で挙げよ。", h=22)
body(ws, r + 4,
     "〔 ① 両方向に耐震要素／② 重心と剛心の距離／③ 弾力半径／"
     "④ 外周・内側のバランス／⑤ コア集中 vs 分散 〕", h=44)
head(ws, r + 6, "■ 問題 3  立面バランス")
body(ws, r + 7,
     "立面的にバランスが悪い 3 つのパターンを挙げよ"
     "（ピロティ／中間階の壁抜け／逆三角形配置 から選んで、各 1 行で説明）。",
     h=58)

# ---- 10-2 ----
ws = wb.create_sheet("No.10-2 偏心率")
setup(ws, [8, 16, 18, 14, 12, 12, 12, 12])
title_row(ws, 1, "No.10-2  偏心率の改善に向けた平面配置検討")
head(ws, 3, "■ 図 10-2  偏心率の改善 (改善前 → 改善後)")
put_img(ws, figs["f2"], "A4", w=720)
head(ws, 30, "■ 問題 1  偏心率の定義（穴埋め）")
body(ws, 31, "偏心率 Re = e / re  において、e は【 ① 】と【 ② 】の距離、"
              "re は【 ③ 】である。基準法では Re ≤ 【 ④ 】を標準とし、"
              "超えると Fe 割増しを行う。", h=44)
head(ws, 33, "■ 問題 2  改善前の剛心 R を計算")
body(ws, 34,
     "改善前は Y 方向壁が 2 枚。下表より ΣKy、Σ(Ky·x)、剛心 X 座標を求めよ。"
     "K は壁の長さに比例とする（K = 長さ／7m）。",
     h=44)
r = table(ws, 37,
          ["壁", "位置 X (m)", "長さ (m)", "Ky", "Ky·x (記入)"],
          [["W1 (X1 通り)", "0", "21", "3.0", ""],
           ["W2 (X2 通り)", "6", "14", "2.0", ""],
           ["合計", "—", "—", "5.0", ""]])
body(ws, r + 2,
     "→ 剛心 X = Σ(Ky·x) / ΣKy = ______ m、重心 X = W/2 = 12 m、"
     "偏心 ex = |12 − 剛心X| = ______ m", h=32)
head(ws, r + 4, "■ 問題 3  改善後の剛心 R を計算")
body(ws, r + 5, "改善後は Y 方向壁を 2 枚追加（W3 と W4）。剛心 X を求めよ。", h=22)
r = table(ws, r + 7,
          ["壁", "X (m)", "L (m)", "Ky", "Ky·x"],
          [["W1", "0", "21", "3.0", "0"],
           ["W2", "6", "14", "2.0", "12"],
           ["W3 (追加)", "18", "14", "2.0", "36"],
           ["W4 (追加)", "24", "14", "2.0", "48"],
           ["合計", "—", "—", "9.0", "96"]])
body(ws, r + 2,
     "→ 剛心 X = 96 / 9 = ______ m、重心 X = 12 m、偏心 ex = ______ m。"
     "改善前と比べ偏心がどう変化したか述べよ。", h=32)
head(ws, r + 4, "■ 問題 4  自主提案")
body(ws, r + 5,
     "改善後の 4 枚配置の代わりに、もっと少ない枚数で同じ効果を得るには"
     "どこに壁を置けばよいか。1 枚の追加案を平面上に図示（位置と長さ）。",
     h=44)

# ---- 10-3 ----
ws = wb.create_sheet("No.10-3 剛性率")
setup(ws, [8, 16, 18, 14, 12, 12, 12, 12])
title_row(ws, 1, "No.10-3  剛性率の改善に向けた鉛直配置検討")
head(ws, 3, "■ 図 10-3  剛性率の改善（立面 + 各階剛性 棒グラフ）")
put_img(ws, figs["f3"], "A4", w=780)
head(ws, 32, "■ 問題 1  剛性率の定義（穴埋め）")
body(ws, 33,
     "剛性率 Rs = rs / rs̄。 rs は当該階の【 ① 】の逆数 (1/Ri)、"
     "rs̄ は全層平均値。基準法では Rs ≥ 【 ② 】を標準とし、超えると Fs 割増し。",
     h=44)
head(ws, 35, "■ 問題 2  改善前 Rs の確認")
body(ws, 36,
     "下表の R (層せん断剛性、相対値) から、各階の Rs を計算せよ。"
     "Rs_i = R_i / R̄ (簡略形)。",
     h=32)
r = table(ws, 39,
          ["階", "R (相対)", "R̄ (記入)", "Rs (記入)", "判定 (Rs≥0.6?)"],
          [["4F", "90", "", "", ""],
           ["3F", "95", "", "", ""],
           ["2F", "100", "", "", ""],
           ["1F", "40 (ピロティ)", "", "", ""]])
body(ws, r + 2,
     "ヒント：R̄ = (90+95+100+40)/4 = 81.25。各階の Rs を計算。"
     "1階の Rs が 0.6 を下回るか確認せよ。", h=32)
head(ws, r + 4, "■ 問題 3  改善案の検討")
body(ws, r + 5,
     "1階の R を 40 から 95 に上げるには何を追加すればよいか。"
     "(a) RC 壁を 2 枚追加、(b) 鉄骨ブレースで対応、(c) 柱断面を大きく"
     "── 3 案のメリット・デメリットを比較せよ。",
     h=72)
head(ws, r + 7, "■ 問題 4  ピロティの実例")
body(ws, r + 8,
     "1981 年新耐震以前のマンションに『1 階駐車場ピロティ』が多数あり、"
     "阪神大震災 (1995) で被害が集中した。被害が出やすい力学的理由を"
     "『剛性率』『応力集中』『塑性ヒンジ』の語を使って説明せよ。", h=72)

# ---- 10-4 ----
ws = wb.create_sheet("No.10-4 RC マンション")
setup(ws, [8, 16, 18, 14, 12, 12, 12, 12])
title_row(ws, 1, "No.10-4  RC 造共同住宅の耐震壁配置計画")
head(ws, 3, "■ 図 10-4  共同住宅 (5戸×1階) の耐震壁配置例")
put_img(ws, figs["f4"], "A4", w=720)
head(ws, 28, "■ 問題 1  方向別の壁量バランス")
body(ws, 29,
     "(1) この計画では X 方向 / Y 方向 それぞれに何枚の耐震壁があるか。表の通り列を埋めよ。",
     h=32)
r = table(ws, 32,
          ["方向", "壁の数", "壁の長さ合計 (m)", "備考"],
          [["Y 方向 (戸境壁)", "", "", "5 戸境壁 × 8m"],
           ["X 方向 (妻壁)", "", "", "妻側 2 箇所 × 5.5m"]])
body(ws, r + 2,
     "(2) 上記から X 方向と Y 方向どちらの壁量が多いか。地震時にどちらの方向が支配的に効くか。",
     h=44)
head(ws, r + 4, "■ 問題 2  X 方向 (長手) の弱さへの対策")
body(ws, r + 5,
     "X 方向は長手で開口 (廊下・バルコニー) が多く、壁量が不足しがち。"
     "次の対策のうち実務的に有効なものを 3 つ選び理由を述べよ。", h=44)
body(ws, r + 6,
     "(a) 妻壁を厚くする  (b) EV シャフト周りをコア壁化  (c) 雁行配置で X 方向壁を追加  "
     "(d) バルコニー側に独立耐震壁を立てる  (e) 鉄骨ブレース併用",
     h=44)
head(ws, r + 8, "■ 問題 3  開口位置との競合")
body(ws, r + 9,
     "戸境壁にユニットバス・キッチン排水のスリーブを通す要望がある。"
     "壁にスリーブを通す際の注意点を 3 つ挙げよ（位置・大きさ・補強筋）。",
     h=58)
head(ws, r + 11, "■ 問題 4  連層 vs 一部抜け")
body(ws, r + 12,
     "戸境壁を 1 階だけ無くしてエントランス用に使いたい意匠要望が出た。"
     "構造的に何が問題か、どう代替するか説明せよ。", h=58)

# ---- 10-5 ----
ws = wb.create_sheet("No.10-5 S 造工場")
setup(ws, [8, 16, 18, 14, 12, 12, 12, 12])
title_row(ws, 1, "No.10-5  S 造工場・倉庫のブレース配置計画")
head(ws, 3, "■ 図 10-5  S 造工場 (50m × 30m) のブレース配置")
put_img(ws, figs["f5"], "A4", w=780)
head(ws, 28, "■ 問題 1  ブレース構面の特定")
body(ws, 29,
     "(1) 図中の赤枠は何方向の地震に抵抗するブレース構面か。",
     h=22)
body(ws, 30,
     "(2) 緑枠は何方向の地震に抵抗するか。", h=22)
body(ws, 31,
     "(3) X 方向 (50m長手) と Y 方向 (30m短手) で、ブレース構面の数が違う理由は？",
     h=32)
head(ws, 33, "■ 問題 2  シャッターとの競合")
body(ws, 34,
     "両妻面 (X1・X6 通り) の中央にシャッター開口がある。"
     "ブレース構面を X1・X6 通りに置く案と、シャッターを避けて X2・X5 通りに置く案を比較せよ。"
     "(a) どちらが構造的に良いか  (b) シャッター開口への影響  (c) 中央スパンの扱い",
     h=72)
head(ws, 36, "■ 問題 3  連層ブレース vs 階別ブレース")
body(ws, 37,
     "工場 2 階建て (各階高 5m) でブレースを 1階のみ／全階／X 形でなく V 形 など"
     "選択肢がある。耐震上の優劣と建方上の留意点を 200 字以内で述べよ。",
     h=72)
head(ws, 39, "■ 問題 4  クレーン荷重との関係")
body(ws, 40,
     "天井クレーン (荷重 5t) が両妻方向に走行する。クレーン走行時の水平制動力が"
     "Y 方向ブレース構面に加わる。ブレース配置で考慮すべき点を 3 点挙げよ。",
     h=58)

# ---- 10-6 ----
ws = wb.create_sheet("No.10-6 建築要望")
setup(ws, [8, 16, 18, 14, 12, 12, 12, 12])
title_row(ws, 1, "No.10-6  建築要望に配慮した配置計画")
head(ws, 3, "■ 図 10-6  ピロティ問題と代替案")
put_img(ws, figs["f6"], "A4", w=760)
head(ws, 27, "■ 問題 1  ピロティ問題の指摘と対策")
body(ws, 28,
     "(1) 図(a) の問題点を『剛性率』『偏心率』『応力集中』の3観点で指摘せよ。",
     h=44)
body(ws, 29,
     "(2) 図(b) の代替案でなぜ問題が解消するか、簡潔に説明せよ。", h=32)
body(ws, 30,
     "(3) 他の対策案 ((c)柱断面増 (d)制振ダンパー (e)1階壁の一部復活) "
     "それぞれのメリット・デメリットを 1 行で。", h=58)
head(ws, 32, "■ 問題 2  店舗 (大開口) 要望への対応")
body(ws, 33,
     "1 階に間口 6m × 高さ 3.5m の大開口の店舗が入り、外周壁を抜きたい要望。"
     "次の代替案を組合せて提案せよ。" , h=44)
r = table(ws, 36,
          ["案", "内容", "メリット", "デメリット", "推奨度"],
          [["A", "鉄骨ブレースを 1 階に挿入", "", "", ""],
           ["B", "RC 耐震壁を背面側に集約", "", "", ""],
           ["C", "境界梁付き耐震壁 (オープンウォール)", "", "", ""],
           ["D", "免震構造に切替 (1 階ピロティ全面開放)", "", "", ""]])
head(ws, r + 2, "■ 問題 3  意匠との合意形成")
body(ws, r + 3,
     "意匠から『窓を大きく』『天井を高く』『柱を細く』の3要望が出た。"
     "それぞれを満たすために構造側で何を譲歩・提案できるか、3 行で書け。",
     h=72)
head(ws, r + 5, "■ 問題 4  実務総合")
body(ws, r + 6,
     "あなたが設計担当する 5 階建てオフィスビル (S 造、平面 30m×20m) で、"
     "意匠から『1 階エントランスを 1 スパン抜きたい (柱だけ残す)』"
     "『屋上に重量設備 (受水槽 20t)』『南面は全面ガラスカーテンウォール』の 3 要望が"
     "同時に出た。構造計画上の優先順位と対応策をまとめよ (200字程度)。",
     h=100)

# ---- 解答 ----
ws = wb.create_sheet("解答")
setup(ws, [4, 22, 60, 14])
title_row(ws, 1, "解答・解説", span=4)
row = 2


def ah(t):
    global row
    head(ws, row, t, span=4); row += 1


def an(t, h=None):
    global row
    body(ws, row, t, span=4, ans=True, h=h); row += 1


ah("No.10-1  配置バランス")
an("問1：(a) NG  壁偏在で剛心が左下に寄り、捻れ振動でねじれモーメントが両端柱に集中。"
   "(b) NG  X 方向地震に抵抗する X 方向壁が無く、Y 方向壁だけでは X 方向の壁量・剛性不足。"
   "(c) GOOD  外周4辺に均等配置：剛心=重心、弾力半径も大きく捻れに強い。"
   "(d) 注意  偏心は無いが弾力半径 re が小さい→大地震時の捻れ変位が大きくなる。"
   "外周にも分散配置することが望ましい。", h=130)
an("問2：①両方向(X・Y)に耐震要素が存在  ②重心と剛心の距離 e が小さい (Re≤0.15)  "
   "③弾力半径 re が大きい (周辺配置)  ④外周と内側のバランス (片寄り回避)  "
   "⑤コア集中だけでなく分散配置で全体剛性確保。", h=72)
an("問3：①ピロティ階 (1階だけ壁/ブレース無し→Rs急減)  "
   "②中間階の壁抜け (連続性無く応力集中)  "
   "③逆三角形配置 (上階に壁多く下階に少ない→下階損傷集中)。", h=58)

ah("No.10-2  偏心率")
an("問1：①重心 G  ②剛心 R  ③弾力半径 re  ④0.15", h=32)
an("問2：W1 Ky·x = 3.0×0 = 0、W2 Ky·x = 2.0×6 = 12、Σ(Ky·x) = 12。"
   "剛心 X = 12/5.0 = 2.4 m。偏心 ex = 12 − 2.4 = 9.6 m。"
   "弾力半径 re_x ≒ √(Σ Ky·(x−X_R)²/ΣKy) = √((3·2.4² + 2·3.6²)/5) "
   "= √((17.3+25.9)/5) = √8.6 ≒ 2.9 m。"
   "Re = 9.6/2.9 ≒ 3.3 → 著しい NG（基準値 0.15 の 20 倍超）。",
   h=130)
an("問3：剛心 X = 96/9 = 10.67 m、偏心 ex = 12 − 10.67 = 1.33 m。"
   "弾力半径 re_x ≒ √(Σ Ky·(x−10.67)²/ΣKy) "
   "= √((3·10.67² + 2·4.67² + 2·7.33² + 2·13.33²)/9) "
   "= √((341.7+43.6+107.5+355.4)/9) = √94.2 ≒ 9.7 m。"
   "Re = 1.33/9.7 ≒ 0.14 → ぎりぎり OK。"
   "改善で偏心率は約 1/24 に低減。", h=130)
an("問4：例として、X4 通り (x=18) に長さ 21m の壁を 1 枚追加すれば、"
   "ΣKy=8、剛心X=(0+12+54)/8=8.25 となるが、これでは偏心が反対側に振れる。"
   "X3 と X4 の中間 (x=15) に長さ 14m を 1 枚追加すれば剛心X≒10付近に持っていける。"
   "壁配置の最適化は試行錯誤と感度解析が基本。", h=86)

ah("No.10-3  剛性率")
an("問1：①層間変形角 R_i (=1/層せん断剛性 K_i)  ②0.6", h=22)
an("問2：R̄=(90+95+100+40)/4=81.25。Rs：4F=90/81.25=1.11、3F=95/81.25=1.17、"
   "2F=100/81.25=1.23、1F=40/81.25=0.49。1F は Rs=0.49 < 0.6 → NG、Fs 割増し対象。",
   h=72)
an("問3：(a)RC 壁追加：確実に剛性確保、コスト中、駐車スペース減。"
   "(b)鉄骨ブレース：開放感維持、施工早い、見た目に出る・防火被覆要。"
   "(c)柱断面増：剛性増は限定的、壁・ブレースに比べ非経済的、駐車計画への影響少。"
   "実務的には (b) 鉄骨ブレースが最も柔軟・経済的。", h=86)
an("問4：1階のRが急減して Rs<0.6 → 大地震時に 1階に層間変形が集中"
   "→ 柱頭・柱脚に塑性ヒンジが先行形成 → せん断破壊・崩壊"
   "(層崩壊メカニズム)。新耐震では Rs・Re の規定強化、ピロティには Fs 割増しで対応。",
   h=86)

ah("No.10-4  RC マンション")
an("問1：(1) Y方向＝5枚×8m=40m、X方向＝2枚×5.5m=11m。"
   "(2) Y 方向壁量が圧倒的に多い → Y 方向地震に十分抵抗。"
   "X 方向は長手で壁量不足になりがち、追加対策が必要。", h=72)
an("問2：(b)EVシャフトのコア壁化＝最も効果大 (X・Y両方向効く)。"
   "(c)雁行配置で X 方向壁を追加＝意匠と調整可で実務多用。"
   "(a)妻壁厚＝補助的だが応力分布の改善に寄与。"
   "(d)バルコニー側壁＝意匠的に難しいが妻側で部分採用例あり。"
   "(e)鉄骨ブレース＝RC造マンションでは住戸内に出るため通常採用しない。",
   h=100)
an("問3：①開口位置は壁中央を避け端部から距離を取る (主筋を切らない)  "
   "②開口寸法は壁断面の 1/3 以下程度  "
   "③開口周りに補強筋 (斜め筋・周辺補強筋) を追加。",
   h=58)
an("問4：1 階だけ戸境壁を抜くと当該階の剛性・耐力が急減 (剛性率低下)、"
   "上階の壁の応力を受け止める基礎梁・柱の応力集中。"
   "代替：壁を抜く位置にラーメン枠 (太い柱・梁) を組む／鉄骨ブレース併用／"
   "境界梁付き連層壁 (Open Wall) で開口を確保しつつ壁を継続。", h=86)

ah("No.10-5  S 造工場")
an("問1：(1) 赤枠 (X1・X3・X6 通り) は Y方向に伸びるブレース構面で X 方向地震に抵抗。"
   "(2) 緑枠 (Y1・Y6 通り) は X 方向に伸びる構面で Y 方向地震に抵抗。"
   "(3) X 方向 (50m) は長く剛性を要するため 3 構面、Y 方向 (30m) は 2 構面で対応。"
   "ブレース構面の間隔は概ね 15-20m 以下を目安に配置。", h=86)
an("問2：(a) シャッターのある妻面 (X1・X6) には大開口がありブレースは妻面の上下端、"
   "もしくはシャッター枠の両側に分割して配置する。"
   "X2・X5 に持って行く案もあるが、剛心が中心寄りになって弾力半径が小さくなる "
   "→ 両妻面配置が捻れに対して有利。シャッターと干渉する箇所はブレースを"
   "上下の小スパンに退避させる、もしくは小開口側に偏らせるなどの調整が必要。",
   h=130)
an("問3：全階ブレース＝最も剛性確保、建方安定。1階のみ＝NG (Rs低下)。"
   "X 形＝最大耐力、引張専用で部材軽量化。V 形・逆 V 形＝大開口に向くが座屈が問題、"
   "BRB（座屈拘束ブレース）で対応するのが現代的。"
   "建方上は X 形は仮設安定性も高く有利、V 形は梁との接合が複雑。", h=100)
an("問4：①クレーン走行制動の繰返し荷重で疲労を考慮 (溶接部・接合部の検討)  "
   "②クレーン走行レール下のランウェイガーダ周りの剛性確保 (走行路の捻れ抑制)  "
   "③走行方向と直角のブレースは制動力を受けるため、許容応力度に余裕を見る。",
   h=72)

ah("No.10-6  建築要望")
an("問1：(1) 剛性率：1階のみ壁無く Rs<0.6 → NG。偏心率：1階で柱配置が"
   "対称なら偏心は小さいが、上階壁配置とのバランス次第。"
   "応力集中：上階壁直下の柱頂部にせん断力集中、層崩壊リスク。"
   "(2) 代替案では 1階に鉄骨ブレース追加で剛性回復、Rs ≥ 0.6 を満たす。"
   "(3) (c)柱断面増＝効果限定、見た目大柱で意匠悪化。"
   "(d)制振ダンパー＝高コストだが大開口維持可、振動低減効果大。"
   "(e)壁部分復活＝最確実だが意匠調整必要。", h=130)
an("問2：A)鉄骨ブレース→意匠表出だが効果確実、店舗内の制約。"
   "B)RC壁を背面側に集約→店舗ファサードは開放、背面の壁量で総剛性確保。"
   "C)境界梁付き耐震壁→開口を持ちつつ壁を継続、設計複雑。"
   "D)免震→1階全面開放可能、コスト最大だが意匠・防災ともに優れる、新築で要採用判断。"
   "推奨：低中層なら B + A 併用、超高層・耐震性能要求高なら D 検討。",
   h=130)
an("問3：①窓大→外周耐震要素を内側コアに集約、外周はラーメンのみで補強"
   "②天井高→階高UPは剛性低下・重心UPで偏心モーメント増、梁せい確保で対応"
   "③柱細→大スパン化困難なため、スパン縮小か梁強化、SC柱(コンクリ充填鋼管)で対応。",
   h=86)
an("問4：優先順位：①剛性率・偏心率の基準達成（一次的安全性）"
   "→②居住性・機能性（重量設備の振動・荷重） →③意匠要望（カーテンウォール）。"
   "対応：エントランス柱抜き→残柱を太く+鉄骨ブレースで補完、屋上設備→振動を抑える"
   "防振架台＋床補強、ガラスCW→外周柱は確保したまま方立・カーテンウォール納まりで対応。",
   h=130)

XLSX = os.path.join(OUT, "バランス配置問題集.xlsx")
wb.save(XLSX)
print("saved:", XLSX)
